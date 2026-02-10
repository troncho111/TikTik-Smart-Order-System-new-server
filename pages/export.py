"""
Export Page - TikTik Smart Order System
עמוד ייצוא דוחות
"""

import io
import streamlit as st
from ui_helpers import render_header
import pandas as pd
from datetime import datetime
from services.order_service import get_all_orders
from models import Order, OrderStatus, get_db


def page_export():
    """Export page for Excel reports"""
    render_header()
    
    st.markdown("### 📊 ייצוא דוחות")
    
    st.markdown("#### בחר טווח תאריכים")
    
    col1, col2 = st.columns(2)
    with col1:
        start_date = st.date_input("מתאריך", value=datetime.now().replace(day=1))
    with col2:
        end_date = st.date_input("עד תאריך", value=datetime.now())
    
    status_filter = st.multiselect(
        "סינון לפי סטטוס",
        ["טיוטה", "נשלח", "נצפה", "נחתם", "בוטל"],
        default=["נשלח", "נחתם"]
    )
    
    if st.button("📥 ייצא ל-Excel", type="primary", use_container_width=True):
        db = get_db()
        if db:
            try:
                from datetime import timedelta
                query = db.query(Order).filter(
                    Order.created_at >= start_date,
                    Order.created_at <= end_date + timedelta(days=1)
                )
                
                orders = query.all()
                
                if orders:
                    import openpyxl
                    from openpyxl.styles import Font, Alignment, PatternFill
                    
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = "הזמנות"
                    
                    headers = ["מספר הזמנה", "תאריך", "אירוע", "לקוח", "אימייל", "טלפון", 
                               "בלוק", "שורה", "מושבים", "כרטיסים", "סה\"כ יורו", "סה\"כ ש\"ח", "סטטוס"]
                    
                    header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
                    header_font = Font(bold=True, color="FFFFFF")
                    
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col, value=header)
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center")
                    
                    status_hebrew = {
                        OrderStatus.DRAFT: "טיוטה",
                        OrderStatus.SENT: "נשלח",
                        OrderStatus.VIEWED: "נצפה",
                        OrderStatus.SIGNED: "נחתם",
                        OrderStatus.CANCELLED: "בוטל"
                    }
                    
                    for row, order in enumerate(orders, 2):
                        ws.cell(row=row, column=1, value=order.order_number)
                        ws.cell(row=row, column=2, value=order.created_at.strftime('%d/%m/%Y') if order.created_at else '')
                        ws.cell(row=row, column=3, value=order.event_name)
                        ws.cell(row=row, column=4, value=order.customer_name)
                        ws.cell(row=row, column=5, value=order.customer_email)
                        ws.cell(row=row, column=6, value=order.customer_phone)
                        ws.cell(row=row, column=7, value=order.block)
                        ws.cell(row=row, column=8, value=order.row)
                        ws.cell(row=row, column=9, value=order.seats)
                        ws.cell(row=row, column=10, value=order.num_tickets)
                        ws.cell(row=row, column=11, value=order.total_euro)
                        ws.cell(row=row, column=12, value=order.total_nis)
                        ws.cell(row=row, column=13, value=status_hebrew.get(order.status, ''))
                    
                    for col in ws.columns:
                        max_length = 0
                        column = col[0].column_letter
                        for cell in col:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = (max_length + 2)
                        ws.column_dimensions[column].width = adjusted_width
                    
                    excel_buffer = io.BytesIO()
                    wb.save(excel_buffer)
                    excel_buffer.seek(0)
                    
                    st.download_button(
                        label="⬇️ הורד קובץ Excel",
                        data=excel_buffer,
                        file_name=f"הזמנות_TikTik_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                    
                    st.success(f"✅ נמצאו {len(orders)} הזמנות לייצוא")
                else:
                    st.warning("לא נמצאו הזמנות בטווח התאריכים שנבחר")
            except Exception as e:
                st.error(f"שגיאה בייצוא: {str(e)}")
            finally:
                db.close()
        else:
            st.error("לא ניתן להתחבר למסד הנתונים")

