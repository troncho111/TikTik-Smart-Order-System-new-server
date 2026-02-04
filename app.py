import streamlit as st
import os
import io
import json
import base64
import secrets
from datetime import datetime, timedelta
from PIL import Image
import tempfile
import uuid
import hashlib
from models import Order, OrderStatus, EventType, AtmosphereImage, User, UserSession, PackageTemplate, get_db, generate_order_number, init_db

def generate_session_token():
    """Generate a random session token"""
    return secrets.token_hex(32)

def create_user_session(user_id):
    """Create a new session in database and return token"""
    db = get_db()
    if not db:
        return None
    try:
        token = generate_session_token()
        expires_at = datetime.utcnow() + timedelta(days=30)
        
        session = UserSession(
            user_id=user_id,
            token=token,
            expires_at=expires_at
        )
        db.add(session)
        db.commit()
        return token
    except Exception as e:
        db.rollback()
        print(f"Error creating session: {e}")
        return None
    finally:
        db.close()

def validate_session_token(token):
    """Validate session token and return user if valid"""
    if not token:
        return None
    db = get_db()
    if not db:
        return None
    try:
        session = db.query(UserSession).filter(
            UserSession.token == token,
            UserSession.expires_at > datetime.utcnow()
        ).first()
        
        if session:
            user = db.query(User).filter(User.id == session.user_id, User.is_active == True).first()
            if user:
                session.last_seen = datetime.utcnow()
                db.commit()
                return {
                    'id': user.id,
                    'username': user.username,
                    'email': user.email,
                    'full_name': user.full_name,
                    'is_admin': user.is_admin
                }
        return None
    except Exception as e:
        print(f"Error validating session: {e}")
        return None
    finally:
        db.close()

def delete_user_session(token):
    """Delete a session from database"""
    if not token:
        return
    db = get_db()
    if not db:
        return
    try:
        db.query(UserSession).filter(UserSession.token == token).delete()
        db.commit()
    except:
        db.rollback()
    finally:
        db.close()

def restore_session_from_token():
    """Try to restore user session from query params with database validation"""
    params = st.query_params
    token = params.get('token')
    
    # New database-backed token system
    if token:
        user = validate_session_token(token)
        if user:
            return user
    
    # Fallback to old hash-based system for backward compatibility
    old_token = params.get('session')
    user_id = params.get('uid')
    
    if old_token and user_id:
        db = get_db()
        if db:
            try:
                secret = os.environ.get('SESSION_SECRET', 'tiktik-secret-key')
                data = f"{user_id}-{secret}"
                expected_token = hashlib.sha256(data.encode()).hexdigest()[:32]
                
                user = db.query(User).filter(User.id == int(user_id), User.is_active == True).first()
                if user and old_token == expected_token:
                    # Migrate to new database-backed session
                    new_token = create_user_session(user.id)
                    if new_token:
                        st.query_params['token'] = new_token
                        # Clean up old params
                        if 'session' in st.query_params:
                            del st.query_params['session']
                        if 'uid' in st.query_params:
                            del st.query_params['uid']
                    return {
                        'id': user.id,
                        'username': user.username,
                        'email': user.email,
                        'full_name': user.full_name,
                        'is_admin': user.is_admin
                    }
            except:
                pass
            finally:
                db.close()
    return None

def set_session_token(user):
    """Set session token in URL query params for persistence"""
    token = create_user_session(user['id'])
    if token:
        st.query_params['token'] = token

def clear_session_token():
    """Clear session token from query params"""
    token = st.query_params.get('token')
    if token:
        delete_user_session(token)
        del st.query_params['token']
    if 'session' in st.query_params:
        del st.query_params['session']
    if 'uid' in st.query_params:
        del st.query_params['uid']
import random
from passport_ocr import extract_passport_data
from hotel_resolver import resolve_hotel_safe
from airports import AIRPORTS, get_airport_options, get_airport_code, format_airport_display
from flight_ocr import extract_flight_data
from streamlit_paste_button import paste_image_button
from stadium_api import get_team_info, get_team_map_path, get_all_teams
from concerts_service import fetch_venue_map_from_ticketmaster, is_ticketmaster_url
from google import genai

def get_gemini_client():
    """Get Gemini client for AI chat"""
    api_key = os.environ.get("AI_INTEGRATIONS_GEMINI_API_KEY")
    base_url = os.environ.get("AI_INTEGRATIONS_GEMINI_BASE_URL")
    
    if not api_key:
        return None
    
    try:
        return genai.Client(
            api_key=api_key,
            http_options={
                'api_version': '',
                'base_url': base_url
            }
        )
    except Exception:
        return None

def ai_chat_response(question: str) -> str:
    """Generate AI response for user question about TikTik system"""
    client = get_gemini_client()
    if not client:
        return "שירות הצ'אט אינו זמין כרגע. פנה למנהל המערכת."
    
    system_prompt = """אתה עוזר וירטואלי של מערכת TikTik. ענה בעברית קצר וברור.

בעיות נפוצות ופתרונות:

❓ "העליתי דרכון ולא קורה כלום"
✅ תשובה: לאחר העלאת התמונה, חייבים ללחוץ על כפתור "🔍 סרוק דרכון"!

❓ "העליתי צילום טיסה ולא קורה כלום"  
✅ תשובה: לאחר העלאת התמונה, חייבים ללחוץ על כפתור "🔍 סרוק טיסה"!

❓ "איפה שער ההמרה?"
✅ תשובה: שער ההמרה מתעדכן אוטומטית (שער בנק ישראל + 5 אגורות). אין צורך להזין.

❓ "איך שולחים ללקוח?"
✅ תשובה: לחץ "צור PDF והורד", ושלח את הקובץ ללקוח דרך וואטסאפ.

מידע על המערכת:
- TikTik מוכרת כרטיסים למשחקי כדורגל והופעות באירופה
- "חבילה מלאה" = מלון + טיסות + העברות + כרטיסים
- "כרטיסים בלבד" = רק כרטיסים
- סריקת דרכון: העלה תמונה → לחץ "סרוק דרכון" → פרטים יתמלאו
- סריקת טיסה: העלה צילום מסך → לחץ "סרוק טיסה" → פרטים יתמלאו
- מלונות: הקלד שם → לחץ "חפש מלון" → פרטים יתמלאו
- מפות אצטדיון מופיעות אוטומטית לפי הקבוצה

ענה קצר וממוקד. אם לא יודע - הפנה לעמוד העזרה (כפתור ❓)."""
    
    try:
        response = client.models.generate_content(
            model="gemini-2.5-flash",
            contents=f"{system_prompt}\n\nשאלת המשתמש: {question}"
        )
        return response.text or "לא הצלחתי לענות. נסה שוב."
    except Exception as e:
        return f"שגיאה: לא הצלחתי לעבד את השאלה. נסה שוב מאוחר יותר."

def render_ai_chatbot():
    """Render AI chatbot widget in sidebar"""
    st.sidebar.markdown("---")
    st.sidebar.markdown("##### 🤖 עוזר AI")
    
    if 'ai_chat_history' not in st.session_state:
        st.session_state.ai_chat_history = []
    
    if 'ai_chat_input' not in st.session_state:
        st.session_state.ai_chat_input = ""
    
    with st.sidebar.expander("💬 שאל שאלה", expanded=False):
        user_question = st.text_input(
            "הקלד שאלה",
            key="ai_question_input",
            placeholder="איך יוצרים הזמנה חדשה?"
        )
        
        if st.button("שלח", key="send_ai_question", use_container_width=True):
            if user_question.strip():
                with st.spinner("חושב..."):
                    response = ai_chat_response(user_question)
                    st.session_state.ai_chat_history.append({
                        "question": user_question,
                        "answer": response
                    })
        
        if st.session_state.ai_chat_history:
            st.markdown("---")
            for i, chat in enumerate(reversed(st.session_state.ai_chat_history[-3:])):
                st.markdown(f"**🙋 שאלה:** {chat['question']}")
                st.markdown(f"**🤖 תשובה:** {chat['answer']}")
                if i < len(st.session_state.ai_chat_history[-3:]) - 1:
                    st.markdown("---")
        
        if st.session_state.ai_chat_history and st.button("🗑️ נקה היסטוריה", key="clear_ai_history"):
            st.session_state.ai_chat_history = []
            st.rerun()

init_db()

st.set_page_config(
    page_title="TikTik Smart Order System",
    page_icon="🎟️",
    layout="wide",
    initial_sidebar_state="expanded"
)

RTL_CSS = """
<style>
@import url('https://fonts.googleapis.com/css2?family=Heebo:wght@300;400;500;600;700&display=swap');

* {
    font-family: 'Heebo', sans-serif !important;
}

/* Fix icon text rendering issue - hide "ke" text from icon fonts */
[data-testid="stExpander"] summary span[data-testid="stMarkdownContainer"],
[data-testid="stExpander"] summary > span:first-child,
.stExpander summary,
details summary {
    font-size: inherit;
    unicode-bidi: isolate;
}

/* Hide any stray icon text before emojis */
[data-testid="stExpander"] summary::before,
[data-testid="stFileUploader"] label::before {
    content: none !important;
    display: none !important;
}

/* Ensure expander icon doesn't show fallback text */
[data-testid="stExpander"] svg + span,
[data-testid="stExpander"] details > summary > span:first-of-type {
    text-indent: 0;
}

/* Fix file uploader label */
[data-testid="stFileUploader"] > label {
    direction: rtl !important;
}

/* Hide stray icon-font characters that render as "ke" */
[class*="icon"]::before,
[class*="Icon"]::before {
    font-family: inherit !important;
}

.main .block-container {
    direction: rtl;
    text-align: right;
}

h1, h2, h3, h4, h5, h6, p, label, span, div {
    direction: rtl;
    text-align: right;
}

.stTextInput > div > div > input,
.stTextArea > div > div > textarea,
.stSelectbox > div > div > div,
.stNumberInput > div > div > input {
    direction: rtl;
    text-align: right;
}

.stButton > button {
    width: 100%;
    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
    color: white;
    border: none;
    padding: 0.75rem 1.5rem;
    font-size: 1.1rem;
    font-weight: 600;
    border-radius: 10px;
    transition: all 0.3s ease;
}

.stButton > button:hover {
    transform: translateY(-2px);
    box-shadow: 0 5px 20px rgba(102, 126, 234, 0.4);
}

.success-button > button {
    background: linear-gradient(135deg, #11998e 0%, #38ef7d 100%) !important;
}

.header-container {
    background: linear-gradient(135deg, #1a1a2e 0%, #16213e 100%);
    padding: 2rem;
    border-radius: 15px;
    margin-bottom: 2rem;
    text-align: center;
    box-shadow: 0 10px 30px rgba(0,0,0,0.3);
}

.header-container h1 {
    color: #fff;
    font-size: 2.5rem;
    margin-bottom: 0.5rem;
    text-align: center;
}

.header-container p {
    color: #a0a0a0;
    font-size: 1.1rem;
    text-align: center;
}

.form-section {
    background: #1e1e2e;
    padding: 1.5rem;
    border-radius: 15px;
    margin-bottom: 1.5rem;
    border: 1px solid #333;
}

.form-section h3 {
    color: #667eea;
    border-bottom: 2px solid #667eea;
    padding-bottom: 0.5rem;
    margin-bottom: 1rem;
}

.preview-box {
    background: linear-gradient(135deg, #1a1a2e 0%, #16213e 100%);
    padding: 2rem;
    border-radius: 15px;
    border: 2px solid #667eea;
    min-height: 400px;
}

.stFileUploader > div {
    direction: rtl;
}

/* Enhanced drag & drop styling */
[data-testid="stFileUploader"] > section > div {
    border: 2px dashed #667eea !important;
    border-radius: 12px !important;
    background: rgba(102, 126, 234, 0.05) !important;
    transition: all 0.3s ease !important;
    padding: 1.5rem !important;
}

[data-testid="stFileUploader"] > section > div:hover {
    border-color: #764ba2 !important;
    background: rgba(102, 126, 234, 0.1) !important;
}

[data-testid="stFileUploader"] > section > div[data-dragging="true"] {
    border-color: #38ef7d !important;
    background: rgba(56, 239, 125, 0.1) !important;
}

.passenger-item {
    background: #2d2d3d;
    padding: 0.5rem 1rem;
    border-radius: 8px;
    margin: 0.3rem 0;
    border-right: 4px solid #667eea;
    color: #ffffff !important;
}

.passenger-item strong,
.passenger-item small {
    color: #ffffff !important;
}

.price-display {
    background: linear-gradient(135deg, #11998e 0%, #38ef7d 100%);
    padding: 1.5rem;
    border-radius: 12px;
    text-align: center;
    color: white;
    font-size: 1.3rem;
    font-weight: 700;
    margin-top: 1rem;
}

.info-box {
    background: #2d2d3d;
    padding: 1rem;
    border-radius: 10px;
    border-right: 4px solid #ffc107;
    margin: 1rem 0;
}

.stDownloadButton > button {
    background: linear-gradient(135deg, #11998e 0%, #38ef7d 100%) !important;
    color: white !important;
    width: 100%;
    padding: 1rem;
    font-size: 1.2rem;
    font-weight: 600;
}

.status-badge {
    padding: 0.3rem 0.8rem;
    border-radius: 20px;
    font-size: 0.85rem;
    font-weight: 600;
    display: inline-block;
}

.status-draft { background: #6c757d; color: white; }
.status-sent { background: #007bff; color: white; }
.status-viewed { background: #ffc107; color: black; }
.status-signed { background: #28a745; color: white; }
.status-cancelled { background: #dc3545; color: white; }

.order-card {
    background: #1e1e2e;
    padding: 1.5rem;
    border-radius: 12px;
    margin-bottom: 1rem;
    border: 1px solid #333;
    transition: all 0.3s ease;
}

.order-card:hover {
    border-color: #667eea;
    transform: translateY(-2px);
}

/* ===== AGGRESSIVE FIX: Remove blue vertical line completely ===== */

/* Target ALL elements with any blue/purple border */
*[style*="border"],
*[style*="667eea"],
*[style*="764ba2"] {
    border-left: none !important;
    border-right: none !important;
}

/* Sidebar and ALL children - no borders at all */
[data-testid="stSidebar"],
[data-testid="stSidebar"] *,
section[data-testid="stSidebar"],
section[data-testid="stSidebar"] * {
    border: none !important;
    border-left: none !important;
    border-right: none !important;
    box-shadow: none !important;
    outline: none !important;
}

/* Sidebar navigation and resize elements */
[data-testid="stSidebarNav"],
[data-testid="stSidebarNavItems"],
[data-testid="stSidebarNavLink"],
[data-testid="stSidebarContent"],
[data-testid="stSidebarUserContent"],
[data-testid="stSidebarResizer"] {
    border: none !important;
    background: transparent !important;
}

/* Hide resize handle completely */
[data-testid="stSidebarResizer"],
[data-testid*="Resizer"] {
    display: none !important;
    width: 0 !important;
    height: 0 !important;
    visibility: hidden !important;
}

/* All pseudo-elements that might create the line */
[data-testid="stSidebar"]::after,
[data-testid="stSidebar"]::before,
section[data-testid="stSidebar"]::after,
section[data-testid="stSidebar"]::before,
[data-testid="stSidebarContent"]::after,
[data-testid="stSidebarContent"]::before {
    display: none !important;
    content: none !important;
    border: none !important;
    background: none !important;
    width: 0 !important;
}

/* Streamlit emotion cache classes that may create borders */
[class*="st-emotion-cache"] {
    border-left: transparent !important;
    border-right: transparent !important;
}

/* FORCE SIDEBAR ALWAYS VISIBLE - NO COLLAPSE */
[data-testid="stSidebar"] {
    transform: none !important;
    transition: none !important;
    width: 250px !important;
    min-width: 250px !important;
    visibility: visible !important;
    opacity: 1 !important;
}

/* HIDE ALL COLLAPSE/EXPAND BUTTONS COMPLETELY */
[data-testid="stSidebarCollapsedControl"],
[data-testid="collapsedControl"],
button[kind="header"],
[data-testid="stSidebarNavCollapseButton"],
button[title="Collapse sidebar"],
button[title="Expand sidebar"],
.stSidebarCollapsedControl,
[class*="SidebarCollapsed"],
svg[data-testid="stSidebarNavCollapseIcon"],
[data-testid="stSidebarCollapseButton"],
[class*="collapsedControl"],
button[data-testid*="collapse"],
button[data-testid*="Collapse"],
[aria-label*="Collapse"],
[aria-label*="Expand"],
.st-emotion-cache-1gwvy71,
.st-emotion-cache-eczf16 {
    display: none !important;
    visibility: hidden !important;
    opacity: 0 !important;
    width: 0 !important;
    height: 0 !important;
    pointer-events: none !important;
    position: absolute !important;
    left: -9999px !important;
}

/* Main app container - ensure no right border */
.main,
.main > div,
[data-testid="stAppViewContainer"],
[data-testid="stAppViewBlockContainer"] {
    border-left: none !important;
    border-right: none !important;
}

/* Block container */
.block-container {
    border: none !important;
}

/* ========== MOBILE RESPONSIVE STYLES ========== */

/* Mobile devices */
@media screen and (max-width: 768px) {
    /* COMPLETELY hide sidebar and all related elements */
    [data-testid="stSidebar"],
    [data-testid="stSidebarNav"],
    [data-testid="stSidebarNavLink"],
    [data-testid="collapsedControl"],
    section[data-testid="stSidebar"],
    div[data-testid="stSidebarCollapsedControl"],
    button[kind="header"],
    .css-1dp5vir,
    .css-1544g2n,
    .css-17ziqus {
        display: none !important;
        visibility: hidden !important;
        width: 0 !important;
        min-width: 0 !important;
        max-width: 0 !important;
        opacity: 0 !important;
        pointer-events: none !important;
        position: absolute !important;
        left: -9999px !important;
    }
    
    /* Hide Streamlit header completely */
    header[data-testid="stHeader"],
    .stDeployButton,
    #MainMenu,
    footer,
    header {
        display: none !important;
        visibility: hidden !important;
    }
    
    /* Main content takes FULL width */
    .main,
    .main > div,
    .block-container,
    [data-testid="stAppViewContainer"],
    [data-testid="stAppViewBlockContainer"] {
        width: 100% !important;
        max-width: 100% !important;
        min-width: 100% !important;
        margin: 0 !important;
        padding: 0 !important;
    }
    
    .main .block-container {
        padding: 0.5rem !important;
        padding-top: 0.5rem !important;
    }
    
    /* App container full width */
    .appview-container {
        width: 100vw !important;
        margin: 0 !important;
    }
    
    /* Remove any left margin from sidebar */
    .css-1d391kg,
    .css-18e3th9 {
        margin-left: 0 !important;
        padding-left: 0 !important;
    }
    
    /* Header styling */
    .header-container {
        padding: 0.75rem !important;
        margin-bottom: 0.75rem !important;
        border-radius: 10px !important;
    }
    
    .header-container h1 {
        font-size: 1.2rem !important;
        margin: 0 !important;
        line-height: 1.3 !important;
    }
    
    .header-container p {
        font-size: 0.8rem !important;
        margin: 0.25rem 0 0 0 !important;
    }
    
    /* Form sections */
    .form-section {
        padding: 0.75rem !important;
        margin-bottom: 0.75rem !important;
        border-radius: 10px !important;
    }
    
    .form-section h3 {
        font-size: 1rem !important;
        margin-bottom: 0.5rem !important;
    }
    
    /* Touch-friendly buttons */
    .stButton > button {
        padding: 0.75rem !important;
        font-size: 1rem !important;
        min-height: 48px !important;
        border-radius: 8px !important;
    }
    
    /* Input fields - prevent iOS zoom */
    .stTextInput > div > div > input,
    .stNumberInput > div > div > input,
    .stTextArea > div > div > textarea,
    .stSelectbox > div > div,
    .stMultiSelect > div > div {
        font-size: 16px !important;
        min-height: 48px !important;
    }
    
    /* Selectbox dropdown */
    .stSelectbox > div > div > div {
        font-size: 16px !important;
        min-height: 48px !important;
        display: flex !important;
        align-items: center !important;
    }
    
    /* File uploader */
    [data-testid="stFileUploader"] > section > div {
        padding: 0.5rem !important;
    }
    
    [data-testid="stFileUploader"] label {
        font-size: 0.85rem !important;
    }
    
    /* Price display */
    .price-display {
        padding: 0.75rem !important;
        font-size: 1rem !important;
    }
    
    /* Preview box */
    .preview-box {
        padding: 0.75rem !important;
        min-height: auto !important;
    }
    
    /* Passenger items */
    .passenger-item {
        padding: 0.5rem !important;
        font-size: 0.9rem !important;
    }
    
    /* Status badges */
    .status-badge {
        font-size: 0.75rem !important;
        padding: 0.25rem 0.5rem !important;
    }
    
    /* Order cards */
    .order-card {
        padding: 0.75rem !important;
    }
    
    /* No hover effects */
    .stButton > button:hover,
    .order-card:hover {
        transform: none !important;
        box-shadow: none !important;
    }
    
    /* Force columns to stack vertically */
    [data-testid="stHorizontalBlock"] {
        flex-direction: column !important;
        gap: 0.5rem !important;
    }
    
    [data-testid="stHorizontalBlock"] > div {
        width: 100% !important;
        flex: none !important;
    }
    
    [data-testid="column"] {
        width: 100% !important;
        flex: none !important;
    }
    
    /* Images responsive */
    [data-testid="stImage"] img {
        max-width: 100% !important;
        height: auto !important;
    }
    
    /* Radio buttons touch-friendly */
    [data-testid="stRadio"] > div {
        gap: 0.5rem !important;
    }
    
    [data-testid="stRadio"] label {
        padding: 0.5rem 0.75rem !important;
        min-height: 44px !important;
        font-size: 0.9rem !important;
        display: flex !important;
        align-items: center !important;
    }
    
    /* Checkbox touch-friendly */
    [data-testid="stCheckbox"] label {
        padding: 0.5rem !important;
        min-height: 44px !important;
        font-size: 0.9rem !important;
    }
    
    /* Date/time inputs */
    .stDateInput input,
    .stTimeInput input {
        font-size: 16px !important;
        min-height: 48px !important;
    }
    
    /* Tabs - horizontal scrollable */
    .stTabs [data-baseweb="tab-list"] {
        gap: 0.25rem !important;
        overflow-x: auto !important;
        flex-wrap: nowrap !important;
        -webkit-overflow-scrolling: touch !important;
    }
    
    .stTabs [data-baseweb="tab"] {
        font-size: 0.8rem !important;
        padding: 0.5rem 0.75rem !important;
        min-height: 40px !important;
        white-space: nowrap !important;
    }
    
    /* Expanders */
    [data-testid="stExpander"] summary {
        font-size: 0.9rem !important;
        padding: 0.75rem !important;
    }
    
    /* Tables scroll */
    [data-testid="stDataFrame"] {
        overflow-x: auto !important;
        -webkit-overflow-scrolling: touch !important;
    }
    
    /* Download button */
    .stDownloadButton > button {
        font-size: 1rem !important;
        padding: 0.75rem !important;
    }
    
    /* Metrics */
    [data-testid="stMetric"] {
        padding: 0.5rem !important;
    }
    
    [data-testid="stMetricValue"] {
        font-size: 1.1rem !important;
    }
    
    /* Info boxes */
    .info-box {
        padding: 0.75rem !important;
        font-size: 0.9rem !important;
    }
}

/* Extra small screens (iPhone SE, small phones) */
@media screen and (max-width: 400px) {
    .main .block-container {
        padding: 0.25rem !important;
    }
    
    .header-container h1 {
        font-size: 1rem !important;
    }
    
    .header-container p {
        font-size: 0.7rem !important;
    }
    
    .form-section {
        padding: 0.5rem !important;
    }
    
    .form-section h3 {
        font-size: 0.85rem !important;
    }
    
    .stButton > button {
        font-size: 0.85rem !important;
        padding: 0.6rem !important;
    }
}
</style>
"""

st.markdown(RTL_CSS, unsafe_allow_html=True)

def generate_pdf(order_data, stadium_image=None, hotel_image=None, hotel_image_2=None, stadium_photo=None, template_version=1):
    """Generate professional PDF using subprocess to avoid blocking Streamlit"""
    import subprocess
    import json
    
    stadium_image_path = None
    hotel_image_path = None
    hotel_image_2_path = None
    stadium_photo_path = None
    
    def save_image_safely(img, prefix="img"):
        """Safely save an image to temp file, handling various formats"""
        try:
            if img is None:
                return None
            
            # If it's bytes, try to load as PIL Image first
            if isinstance(img, bytes):
                try:
                    img = Image.open(io.BytesIO(img))
                except Exception:
                    return None
            
            # Convert to RGB if needed (for PNG with transparency, RGBA, etc.)
            if not isinstance(img, Image.Image):
                return None
            
            if img.mode in ('RGBA', 'LA', 'P'):
                # Create white background for transparency
                background = Image.new('RGB', img.size, (255, 255, 255))
                if img.mode == 'P':
                    img = img.convert('RGBA')
                background.paste(img, mask=img.split()[-1] if img.mode in ('RGBA', 'LA') else None)
                img = background
            elif img.mode != 'RGB':
                img = img.convert('RGB')
            
            with tempfile.NamedTemporaryFile(suffix='.png', delete=False, prefix=prefix) as tmp:
                img.save(tmp.name, 'PNG', optimize=True)
                return tmp.name
        except Exception as e:
            print(f"Error saving image: {e}")
            return None
    
    if stadium_image:
        stadium_image_path = save_image_safely(stadium_image, "stadium_")
    
    if stadium_photo:
        stadium_photo_path = save_image_safely(stadium_photo, "atmosphere_")
    
    if hotel_image:
        hotel_image_path = save_image_safely(hotel_image, "hotel_")
    
    if hotel_image_2:
        hotel_image_2_path = save_image_safely(hotel_image_2, "hotel2_")
    
    pdf_data = {
        'product_type': order_data.get('product_type', 'tickets'),
        'event_name': order_data['event_name'],
        'event_date': order_data['event_date'],
        'venue': order_data['venue'],
        'event_type': order_data.get('event_type', ''),
        'category': order_data.get('category', ''),
        'ticket_description': order_data.get('ticket_description', ''),
        'passengers': order_data.get('passengers', []),
        'price_per_ticket': order_data['price_per_ticket'],
        'price_nis': order_data['price_nis'],
        'total_euro': order_data['total_euro'],
        'total_nis': order_data['total_nis'],
        'num_tickets': order_data.get('num_tickets', 1),
        'exchange_rate': order_data.get('exchange_rate', 4.0),
        'order_number': order_data.get('order_number', ''),
        'customer_name': order_data['customer_name'],
        'customer_id': order_data['customer_id'],
        'customer_phone': order_data['customer_phone'],
        'customer_email': order_data['customer_email'],
        'hotel_name': order_data.get('hotel_name', ''),
        'hotel_nights': order_data.get('hotel_nights', 0),
        'hotel_stars': order_data.get('hotel_stars', ''),
        'hotel_meals': order_data.get('hotel_meals', ''),
        'hotel_address': order_data.get('hotel_address', ''),
        'hotel_website': order_data.get('hotel_website', ''),
        'hotel_rating': order_data.get('hotel_rating', ''),
        'flight_details': order_data.get('flight_details', ''),
        'flights': order_data.get('flights', []),
        'transfers': order_data.get('transfers', False),
        'bag_trolley': order_data.get('bag_trolley', False),
        'bag_checked': order_data.get('bag_checked', ''),
        'is_date_final': order_data.get('is_date_final', False),
        'seats_together': order_data.get('seats_together', False),
        'template_version': template_version,
        'stadium_image_path': stadium_image_path,
        'stadium_photo_path': stadium_photo_path,
        'hotel_image_path': order_data.get('hotel_image_path') or hotel_image_path,
        'hotel_image_2_path': order_data.get('hotel_image_path_2') or hotel_image_2_path
    }
    
    json_file = None
    try:
        with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False, encoding='utf-8') as jf:
            json.dump(pdf_data, jf)
            json_file = jf.name
        
        result = subprocess.run(
            ['python3', 'pdf_generator.py', json_file],
            capture_output=True,
            text=True,
            timeout=60
        )
        
        if result.returncode != 0:
            raise Exception(f"PDF generation failed: {result.stderr}")
        
        pdf_bytes = base64.b64decode(result.stdout.strip())
        return pdf_bytes
        
    finally:
        if json_file and os.path.exists(json_file):
            os.unlink(json_file)
        if stadium_image_path and os.path.exists(stadium_image_path):
            os.unlink(stadium_image_path)
        if stadium_photo_path and os.path.exists(stadium_photo_path):
            os.unlink(stadium_photo_path)
        if hotel_image_path and os.path.exists(hotel_image_path):
            os.unlink(hotel_image_path)
        if hotel_image_2_path and os.path.exists(hotel_image_2_path):
            os.unlink(hotel_image_2_path)

def get_event_type_from_hebrew(hebrew_type):
    """Map Hebrew event type to EventType enum"""
    type_map = {
        "כדורגל": EventType.FOOTBALL,
        "הופעה": EventType.CONCERT,
        "אחר": EventType.OTHER
    }
    return type_map.get(hebrew_type, EventType.OTHER)

def save_order_to_db(order_data, pdf_bytes=None):
    """Save order to database"""
    db = get_db()
    if not db:
        return None
    
    try:
        event_type = get_event_type_from_hebrew(order_data.get('event_type', 'אחר'))
        
        user_id = None
        if st.session_state.get('user'):
            user_id = st.session_state.user.get('id')
        
        order = Order(
            order_number=order_data.get('order_number') or generate_order_number(),
            user_id=user_id,
            event_name=order_data['event_name'],
            event_date=order_data.get('event_date_str', ''),
            event_time=order_data.get('event_time_str', ''),
            venue=order_data.get('venue', ''),
            event_type=event_type,
            customer_name=order_data['customer_name'],
            customer_id=order_data.get('customer_id', ''),
            customer_email=order_data.get('customer_email', ''),
            customer_phone=order_data.get('customer_phone', ''),
            ticket_description=order_data.get('ticket_description', ''),
            block=order_data.get('category', ''),
            row='',
            seats='',
            num_tickets=order_data.get('num_tickets', 1),
            price_per_ticket_euro=order_data.get('price_per_ticket', 0),
            exchange_rate=order_data.get('exchange_rate', 3.78),
            total_euro=order_data.get('total_euro', 0),
            total_nis=order_data.get('total_nis', 0),
            passengers=json.dumps(order_data.get('passengers', []), ensure_ascii=False),
            status=OrderStatus.DRAFT,
            signature_token=str(uuid.uuid4())
        )
        
        db.add(order)
        db.commit()
        db.refresh(order)
        return order
    except Exception as e:
        db.rollback()
        st.error(f"שגיאה בשמירת ההזמנה: {str(e)}")
        return None
    finally:
        db.close()

def update_order_status(order_id, new_status):
    """Update order status"""
    db = get_db()
    if not db:
        return False
    
    try:
        order = db.query(Order).filter(Order.id == order_id).first()
        if order:
            order.status = new_status
            if new_status == OrderStatus.SENT:
                order.sent_at = datetime.utcnow()
            elif new_status == OrderStatus.VIEWED:
                order.viewed_at = datetime.utcnow()
            elif new_status == OrderStatus.SIGNED:
                order.signed_at = datetime.utcnow()
            db.commit()
            return True
        return False
    except Exception as e:
        db.rollback()
        return False
    finally:
        db.close()

def delete_order(order_id):
    """Delete an order permanently"""
    db = get_db()
    if not db:
        return False
    
    try:
        order = db.query(Order).filter(Order.id == order_id).first()
        if order:
            db.delete(order)
            db.commit()
            return True
        return False
    except Exception as e:
        db.rollback()
        return False
    finally:
        db.close()

def get_saved_concerts():
    """Get all saved concerts for quick reuse"""
    from models import SavedConcert
    db = get_db()
    if not db:
        return []
    
    try:
        concerts = db.query(SavedConcert).filter(SavedConcert.is_active == True).order_by(SavedConcert.created_at.desc()).all()
        return [c.to_dict() for c in concerts]
    except Exception as e:
        print(f"Error getting saved concerts: {e}")
        return []
    finally:
        db.close()

def save_concert_to_favorites(artist_name, artist_name_he, venue_name, city, country, event_date=None, event_time=None, event_url=None, category=None, event_name=None, source=None, stadium_map_path=None, stadium_map_data=None, stadium_map_mime=None):
    """Save a manually entered concert to favorites for quick reuse"""
    from models import SavedConcert
    db = get_db()
    if not db:
        return False
    
    try:
        existing = db.query(SavedConcert).filter(
            SavedConcert.artist_name == artist_name,
            SavedConcert.venue_name == venue_name,
            SavedConcert.is_active == True
        ).first()
        
        if existing:
            existing.city = city
            existing.country = country
            existing.event_date = event_date
            existing.event_time = event_time
            existing.event_url = event_url
            existing.category = category
            existing.event_name = event_name
            existing.source = source or 'saved'
            if stadium_map_path:
                existing.stadium_map_path = stadium_map_path
            if stadium_map_data:
                existing.stadium_map_data = stadium_map_data
                existing.stadium_map_mime = stadium_map_mime or 'image/png'
        else:
            new_concert = SavedConcert(
                artist_name=artist_name,
                artist_name_he=artist_name_he,
                event_name=event_name,
                venue_name=venue_name,
                city=city,
                country=country,
                event_date=event_date,
                event_time=event_time,
                event_url=event_url,
                category=category,
                source=source or 'saved',
                stadium_map_path=stadium_map_path,
                stadium_map_data=stadium_map_data,
                stadium_map_mime=stadium_map_mime or 'image/png' if stadium_map_data else None
            )
            db.add(new_concert)
        
        db.commit()
        return True
    except Exception as e:
        db.rollback()
        print(f"Error saving concert: {e}")
        return False
    finally:
        db.close()

def delete_saved_concert(concert_id):
    """Delete a saved concert"""
    from models import SavedConcert
    db = get_db()
    if not db:
        return False
    
    try:
        concert = db.query(SavedConcert).filter(SavedConcert.id == concert_id).first()
        if concert:
            concert.is_active = False
            db.commit()
            return True
        return False
    except Exception as e:
        db.rollback()
        return False
    finally:
        db.close()

def get_saved_artists():
    """Get all saved artists for the dropdown"""
    from models import SavedArtist
    db = get_db()
    if not db:
        return []
    
    try:
        artists = db.query(SavedArtist).filter(SavedArtist.is_active == True).order_by(SavedArtist.name_en).all()
        return [a.to_dict() for a in artists]
    except Exception as e:
        print(f"Error getting saved artists: {e}")
        return []
    finally:
        db.close()

def save_artist_to_favorites(name_en, name_he=None, ticketmaster_id=None, genre=None, image_url=None):
    """Save an artist to favorites for quick access in dropdown"""
    from models import SavedArtist
    db = get_db()
    if not db:
        return False
    
    try:
        existing = db.query(SavedArtist).filter(
            SavedArtist.name_en == name_en,
            SavedArtist.is_active == True
        ).first()
        
        if existing:
            return True
        
        new_artist = SavedArtist(
            name_en=name_en,
            name_he=name_he or name_en,
            ticketmaster_id=ticketmaster_id,
            genre=genre,
            image_url=image_url
        )
        db.add(new_artist)
        db.commit()
        return True
    except Exception as e:
        db.rollback()
        print(f"Error saving artist: {e}")
        return False
    finally:
        db.close()

def delete_saved_artist(artist_id):
    """Delete a saved artist"""
    from models import SavedArtist
    db = get_db()
    if not db:
        return False
    
    try:
        artist = db.query(SavedArtist).filter(SavedArtist.id == artist_id).first()
        if artist:
            artist.is_active = False
            db.commit()
            return True
        return False
    except Exception as e:
        db.rollback()
        return False
    finally:
        db.close()

def get_all_orders(search_query=None, status_filter=None, user_id=None, is_admin=False):
    """Get all orders with optional filtering"""
    db = get_db()
    if not db:
        return []
    
    try:
        query = db.query(Order).order_by(Order.created_at.desc())
        
        if not is_admin and user_id:
            query = query.filter(Order.user_id == user_id)
        
        if search_query:
            search = f"%{search_query}%"
            query = query.filter(
                (Order.customer_name.ilike(search)) |
                (Order.event_name.ilike(search)) |
                (Order.order_number.ilike(search)) |
                (Order.customer_email.ilike(search))
            )
        
        if status_filter and status_filter != "הכל":
            status_map = {
                "טיוטה": OrderStatus.DRAFT,
                "נשלח": OrderStatus.SENT,
                "נצפה": OrderStatus.VIEWED,
                "נחתם": OrderStatus.SIGNED,
                "בוטל": OrderStatus.CANCELLED
            }
            if status_filter in status_map:
                query = query.filter(Order.status == status_map[status_filter])
        
        return query.all()
    except Exception as e:
        return []
    finally:
        db.close()

def get_status_badge(status):
    """Get HTML badge for status"""
    status_config = {
        OrderStatus.DRAFT: ("טיוטה", "status-draft"),
        OrderStatus.SENT: ("נשלח", "status-sent"),
        OrderStatus.VIEWED: ("נצפה", "status-viewed"),
        OrderStatus.SIGNED: ("נחתם", "status-signed"),
        OrderStatus.CANCELLED: ("בוטל", "status-cancelled")
    }
    
    label, css_class = status_config.get(status, ("לא ידוע", "status-draft"))
    return f'<span class="status-badge {css_class}">{label}</span>'

def render_header():
    """Render the header section"""
    st.markdown("""
    <div class="header-container">
        <h1>🎟️ TikTik Smart Order System</h1>
        <p>מערכת חכמה ליצירת הצעות מחיר והזמנות מקצועיות</p>
    </div>
    """, unsafe_allow_html=True)

def page_new_order():
    """New order page"""
    render_header()
    
    if st.session_state.get('package_loaded_success'):
        pkg_name = st.session_state['package_loaded_success']
        st.success(f"✅ נטענה חבילה: {pkg_name}")
        st.info("💡 עכשיו רק צריך להוסיף פרטי נוסעים ולקוח!")
        del st.session_state['package_loaded_success']
    
    if 'passengers' not in st.session_state:
        st.session_state.passengers = []
    if 'order_generated' not in st.session_state:
        st.session_state.order_generated = False
    
    if 'random_data' not in st.session_state:
        st.session_state.random_data = None
    
    st.markdown("""
        <a href="https://travel-agents-calculatornew.pages.dev/" target="_blank" style="
            display: inline-block;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 12px 24px;
            border-radius: 10px;
            text-decoration: none;
            font-weight: bold;
            font-size: 16px;
            box-shadow: 0 4px 15px rgba(102, 126, 234, 0.4);
            transition: all 0.3s ease;
            margin-bottom: 20px;
        ">
            🧮 מחשבון סוכנים
        </a>
    """, unsafe_allow_html=True)
    
    col1, col2 = st.columns([1.2, 0.8])
    
    with col1:
        st.markdown("### 📝 פרטי ההזמנה")
        
        rd = st.session_state.random_data or {}
        
        st.markdown('<div class="form-section"><h3>📦 סוג המוצר</h3></div>', unsafe_allow_html=True)
        product_options = ["tickets", "package"]
        product_default = 1 if rd.get('product_type') == 'package' else 0
        product_type = st.radio(
            "בחר סוג מוצר",
            options=product_options,
            index=product_default,
            format_func=lambda x: "🎫 כרטיסים בלבד" if x == "tickets" else "✈️ חבילה מלאה (טיסה + מלון + כרטיס)",
            horizontal=True
        )
        
        st.markdown('<div class="form-section"><h3>📦 טעינה מחבילה קבועה</h3></div>', unsafe_allow_html=True)
        
        db = get_db()
        saved_packages = []
        if db:
            try:
                saved_packages = db.query(PackageTemplate).filter(PackageTemplate.is_active == True).order_by(PackageTemplate.name).all()
            except:
                pass
            finally:
                db.close()
        
        if saved_packages:
            package_options = ["-- בחר חבילה --"] + [f"{pkg.name}" for pkg in saved_packages]
            package_ids = [None] + [pkg.id for pkg in saved_packages]
            
            selected_package_idx = st.selectbox(
                "📦 טען פרטים מחבילה שמורה",
                range(len(package_options)),
                format_func=lambda x: package_options[x],
                key="load_package_select"
            )
            
            if selected_package_idx and selected_package_idx > 0:
                pkg_id = package_ids[selected_package_idx]
                selected_pkg = next((p for p in saved_packages if p.id == pkg_id), None)
                
                if selected_pkg and st.button("📥 טען חבילה", type="primary", use_container_width=True):
                    pkg_data = selected_pkg.to_dict()
                    
                    product_val = "package" if pkg_data.get('product_type') == 'full_package' else "tickets"
                    event_type_map = {'concert': 'הופעה', 'football': 'כדורגל', 'other': 'אחר'}
                    
                    flights = pkg_data.get('flights', {})
                    hotel = pkg_data.get('hotel', {})
                    
                    st.session_state.random_data = {
                        'product_type': product_val,
                        'event_name': pkg_data.get('event_name', ''),
                        'event_date': pkg_data.get('event_date', ''),
                        'event_time': pkg_data.get('event_time', ''),
                        'venue': pkg_data.get('venue', ''),
                        'event_type': event_type_map.get(pkg_data.get('event_type'), 'הופעה'),
                        'category': pkg_data.get('ticket_category', ''),
                        'ticket_description': pkg_data.get('ticket_description', ''),
                        'price_euro': int(pkg_data.get('package_price_euro', 0) or 0),
                        'hotel_name': hotel.get('name', ''),
                        'hotel_checkin': hotel.get('check_in', ''),
                        'hotel_checkout': hotel.get('check_out', ''),
                        'outbound_from': flights.get('outbound', {}).get('from', ''),
                        'outbound_to': flights.get('outbound', {}).get('to', ''),
                        'outbound_date': flights.get('outbound', {}).get('date', ''),
                        'outbound_time': flights.get('outbound', {}).get('time', ''),
                        'outbound_flight': flights.get('outbound', {}).get('flight_number', ''),
                        'return_from': flights.get('return', {}).get('from', ''),
                        'return_to': flights.get('return', {}).get('to', ''),
                        'return_date': flights.get('return', {}).get('date', ''),
                        'return_time': flights.get('return', {}).get('time', ''),
                        'return_flight': flights.get('return', {}).get('flight_number', ''),
                        'loaded_from_package': pkg_data.get('name', ''),
                        'package_notes': pkg_data.get('notes', '')
                    }
                    
                    if flights.get('outbound'):
                        st.session_state['flight_outbound_from'] = flights['outbound'].get('from', '')
                        st.session_state['flight_outbound_to'] = flights['outbound'].get('to', '')
                        st.session_state['flight_outbound_date'] = flights['outbound'].get('date', '')
                        st.session_state['flight_outbound_time'] = flights['outbound'].get('time', '')
                        st.session_state['flight_outbound_no'] = flights['outbound'].get('flight_number', '')
                    if flights.get('return'):
                        st.session_state['flight_return_from'] = flights['return'].get('from', '')
                        st.session_state['flight_return_to'] = flights['return'].get('to', '')
                        st.session_state['flight_return_date'] = flights['return'].get('date', '')
                        st.session_state['flight_return_time'] = flights['return'].get('time', '')
                        st.session_state['flight_return_no'] = flights['return'].get('flight_number', '')
                    
                    if pkg_data.get('stadium_map_data'):
                        import base64
                        map_bytes = base64.b64decode(pkg_data['stadium_map_data'])
                        st.session_state['saved_stadium_map_bytes'] = map_bytes
                        st.session_state['package_stadium_map_loaded'] = True
                    
                    if hotel:
                        st.session_state.hotel_data = {
                            'hotel_name': hotel.get('name', ''),
                            'hotel_address': hotel.get('address', ''),
                            'hotel_website': hotel.get('website', ''),
                            'hotel_rating': hotel.get('rating', 0),
                            'hotel_stars': hotel.get('stars', '5 כוכבים'),
                            'hotel_image_path': hotel.get('image_path', ''),
                            'hotel_image_path_2': hotel.get('image_path_2', ''),
                            'hotel_checkin': hotel.get('check_in', ''),
                            'hotel_checkout': hotel.get('check_out', ''),
                            'from_package': True
                        }
                    
                    st.session_state['package_loaded_success'] = pkg_data.get('name')
                    st.rerun()
        else:
            st.caption("💡 אין חבילות שמורות. צור חבילות דרך 'חבילות קבועות' בתפריט.")
        
        st.markdown("---")
        
        if st.button("🎲 מילוי רנדומלי לבדיקה", type="secondary"):
            import random
            from sports_api import LEAGUES, get_teams_by_league, get_hebrew_name
            
            sample_football_matches = [
                {"league": "ליגה ספרדית", "home": "Real Madrid", "away": "Barcelona", "stadium": "Santiago Bernabeu", "city": "Madrid", "hotel": "Hotel Villa Magna Madrid"},
                {"league": "פרמיירליג", "home": "Manchester United", "away": "Liverpool", "stadium": "Old Trafford", "city": "Manchester", "hotel": "The Lowry Hotel Manchester"},
                {"league": "בונדסליגה", "home": "Bayern Munich", "away": "Borussia Dortmund", "stadium": "Allianz Arena", "city": "Munich", "hotel": "Mandarin Oriental Munich"},
                {"league": "סריה A", "home": "AC Milan", "away": "Inter", "stadium": "San Siro", "city": "Milan", "hotel": "Armani Hotel Milano"},
                {"league": "ליגה ספרדית", "home": "Atletico Madrid", "away": "Sevilla", "stadium": "Civitas Metropolitano", "city": "Madrid", "hotel": "Four Seasons Hotel Madrid"},
            ]
            
            sample_passengers = [
                [
                    {"first_name": "Israel", "last_name": "Israeli", "passport": "12345678", "birth_date": "15/03/1985", "passport_expiry": "20/05/2030", "ticket_type": "כרטיס רגיל"},
                    {"first_name": "Sarah", "last_name": "Israeli", "passport": "87654321", "birth_date": "22/07/1988", "passport_expiry": "18/09/2029", "ticket_type": "כרטיס רגיל"},
                ],
                [
                    {"first_name": "David", "last_name": "Cohen", "passport": "11223344", "birth_date": "01/01/1990", "passport_expiry": "01/01/2031", "ticket_type": "כרטיס VIP"},
                    {"first_name": "Rachel", "last_name": "Cohen", "passport": "44332211", "birth_date": "15/06/1992", "passport_expiry": "15/06/2032", "ticket_type": "כרטיס VIP"},
                    {"first_name": "Yosef", "last_name": "Cohen", "passport": "55667788", "birth_date": "10/10/2010", "passport_expiry": "10/10/2028", "ticket_type": "כרטיס ילד"},
                ],
            ]
            
            sample_structured_flights = {
                "Madrid": {
                    'outbound': {'from': 'TLV', 'to': 'MAD', 'date': '15/01/2025', 'time': '09:00', 'flight_number': 'LY315'},
                    'return': {'from': 'MAD', 'to': 'TLV', 'date': '18/01/2025', 'time': '22:00', 'flight_number': 'LY316'}
                },
                "Manchester": {
                    'outbound': {'from': 'TLV', 'to': 'MAN', 'date': '20/02/2025', 'time': '07:30', 'flight_number': 'LY317'},
                    'return': {'from': 'MAN', 'to': 'TLV', 'date': '24/02/2025', 'time': '21:00', 'flight_number': 'LY318'}
                },
                "Munich": {
                    'outbound': {'from': 'TLV', 'to': 'MUC', 'date': '10/03/2025', 'time': '08:00', 'flight_number': 'LH681'},
                    'return': {'from': 'MUC', 'to': 'TLV', 'date': '13/03/2025', 'time': '19:00', 'flight_number': 'LH682'}
                },
                "Milan": {
                    'outbound': {'from': 'TLV', 'to': 'MXP', 'date': '05/04/2025', 'time': '06:30', 'flight_number': 'LY381'},
                    'return': {'from': 'MXP', 'to': 'TLV', 'date': '08/04/2025', 'time': '20:00', 'flight_number': 'LY382'}
                },
            }
            
            passengers = random.choice(sample_passengers)
            currency = 'EUR'
            currency_symbols = {'EUR': '€', 'USD': '$', 'GBP': '£'}
            
            match = random.choice(sample_football_matches)
            home_heb = get_hebrew_name(match['home'])
            away_heb = get_hebrew_name(match['away'])
            event_name = f"{home_heb} נגד {away_heb}"
            venue = f"{match['stadium']}, {match['city']}"
            event_type = "כדורגל"
            hotel = match['hotel']
            
            st.session_state['football_league'] = match['league']
            st.session_state['football_team1'] = f"{home_heb} ({match['home']})"
            st.session_state['football_team2'] = f"{away_heb} ({match['away']})"
            
            league_eng = LEAGUES.get(match['league'], "")
            teams = get_teams_by_league(league_eng)
            
            home_team = next((t for t in teams if t['name'] == match['home']), None)
            away_team = next((t for t in teams if t['name'] == match['away']), None)
            
            if home_team:
                st.session_state['selected_team_data'] = home_team
                st.session_state['home_team_hebrew'] = home_heb
            if away_team:
                st.session_state['away_team_data'] = away_team
                st.session_state['away_team_hebrew'] = away_heb
            
            flights = sample_structured_flights.get(match['city'], sample_structured_flights['Madrid'])
            
            st.session_state.random_data = {
                'product_type': product_type,
                'event_name': event_name,
                'venue': venue,
                'event_type': event_type,
                'hotel_name': hotel,
                'hotel_nights': 3,
                'hotel_stars': "5 כוכבים",
                'hotel_meals': "ארוחת בוקר",
                'transfers': True,
                'outbound_from': flights['outbound']['from'],
                'outbound_to': flights['outbound']['to'],
                'outbound_date': flights['outbound']['date'],
                'outbound_time': flights['outbound']['time'],
                'outbound_flight': flights['outbound']['flight_number'],
                'return_from': flights['return']['from'],
                'return_to': flights['return']['to'],
                'return_date': flights['return']['date'],
                'return_time': flights['return']['time'],
                'return_flight': flights['return']['flight_number'],
                'bag_trolley': True,
                'bag_checked': '23kg',
                'customer_name': passengers[0]['first_name'] + " " + passengers[0]['last_name'],
                'customer_id': ''.join([str(random.randint(0, 9)) for _ in range(9)]),
                'customer_phone': f"052-{random.randint(1000000, 9999999)}",
                'customer_email': f"{passengers[0]['first_name'].lower()}.{passengers[0]['last_name'].lower()}@gmail.com",
                'category': random.choice(["CAT 1", "CAT 2", "CAT 3"]),
                'ticket_description': "כרטיסים בקטגוריה מול המגרש",
                'currency': currency,
                'currency_symbol': currency_symbols[currency],
                'price_euro': random.choice([350, 450, 550, 750]),
                'num_tickets': len(passengers),
                'passengers': passengers,
                'use_sample_images': True
            }
            
            st.session_state.passenger_list = passengers
            for i, p in enumerate(passengers):
                st.session_state[f"first_name_{i}"] = p['first_name']
                st.session_state[f"last_name_{i}"] = p['last_name']
                st.session_state[f"passport_{i}"] = p['passport']
                st.session_state[f"birth_date_{i}"] = p['birth_date']
                st.session_state[f"passport_expiry_{i}"] = p['passport_expiry']
            
            st.session_state['flight_outbound_from'] = flights['outbound']['from']
            st.session_state['flight_outbound_to'] = flights['outbound']['to']
            st.session_state['flight_outbound_date'] = flights['outbound']['date']
            st.session_state['flight_outbound_time'] = flights['outbound']['time']
            st.session_state['flight_outbound_no'] = flights['outbound']['flight_number']
            st.session_state['flight_return_from'] = flights['return']['from']
            st.session_state['flight_return_to'] = flights['return']['to']
            st.session_state['flight_return_date'] = flights['return']['date']
            st.session_state['flight_return_time'] = flights['return']['time']
            st.session_state['flight_return_no'] = flights['return']['flight_number']
            
            st.rerun()
        
        col_random, col_clear = st.columns(2)
        with col_clear:
            if st.button("🗑️ ניקוי טופס", type="secondary"):
                keys_to_clear = [
                    'random_data', 'passenger_list', 'order_generated', 'pdf_bytes',
                    'current_order_number', 'current_order_id', 'selected_team_data',
                    'away_team_data', 'home_team_hebrew', 'away_team_hebrew',
                    'football_league', 'hotel_data', 'pasted_passports', '_passport_paste_refresh', 'pasted_flight',
                    'worldcup_match', 'worldcup_venue', 'fixture_data', 'worldcup_stadium_map',
                    'pasted_stadium_map', 'saved_stadium_map_path', 'saved_stadium_map_bytes', '_selected_concert',
                    '_from_saved_concert', 'concert_venue_info', 'concert_artist_en',
                    'concert_artist_he', 'concert_venue_name', 'concert_venue_city',
                    'concert_selected_category', '_concert_venue_id'
                ]
                for key in keys_to_clear:
                    if key in st.session_state:
                        del st.session_state[key]
                for key in list(st.session_state.keys()):
                    if key.startswith(('first_name_', 'last_name_', 'passport_', 'birth_date_', 'passport_expiry_', 'flight_')):
                        del st.session_state[key]
                st.session_state.passenger_list = [{'first_name': '', 'last_name': '', 'passport': '', 'birth_date': '', 'passport_expiry': '', 'ticket_type': 'כרטיס רגיל'}]
                st.rerun()
        
        st.markdown('<div class="form-section"><h3>🎭 פרטי האירוע</h3></div>', unsafe_allow_html=True)
        
        event_types = ["כדורגל", "הופעה", "אחר"]
        default_type_idx = event_types.index(rd.get('event_type', 'כדורגל')) if rd.get('event_type') in event_types else 0
        event_type = st.selectbox("סוג אירוע", event_types, index=default_type_idx)
        
        if event_type == "כדורגל":
            from sports_api import LEAGUES, get_teams_by_league, get_hebrew_name, TEAM_HEBREW_NAMES, find_fixture
            
            st.markdown("##### ⚽ בחירת קבוצות (השלמה אוטומטית)")
            
            col_league = st.columns([1])[0]
            with col_league:
                league_options = ["-- בחר ליגה --"] + list(LEAGUES.keys())
                selected_league = st.selectbox("ליגה", league_options, key="football_league")
            
            is_worldcup = selected_league == "מונדיאל 2026"
            
            if is_worldcup:
                try:
                    with open('worldcup2026.json', 'r', encoding='utf-8') as f:
                        wc_data = json.load(f)
                    wc_matches = wc_data.get('matches', [])
                except:
                    wc_matches = []
                
                NATIONAL_TEAM_HEBREW = {
                    "Mexico": "מקסיקו", "South Africa": "דרום אפריקה", "South Korea": "דרום קוריאה",
                    "Canada": "קנדה", "USA": "ארה\"ב", "Paraguay": "פרגוואי", "Haiti": "האיטי",
                    "Scotland": "סקוטלנד", "Australia": "אוסטרליה", "Brazil": "ברזיל", "Morocco": "מרוקו",
                    "Qatar": "קטאר", "Switzerland": "שוויץ", "Ivory Coast": "חוף השנהב",
                    "Ecuador": "אקוודור", "Germany": "גרמניה", "Curacao": "קוראסאו",
                    "Netherlands": "הולנד", "Japan": "יפן", "Tunisia": "טוניסיה",
                    "Saudi Arabia": "סעודיה", "Uruguay": "אורוגוואי", "Spain": "ספרד",
                    "Cabo Verde": "קאבו ורדה", "Iran": "איראן", "New Zealand": "ניו זילנד",
                    "Belgium": "בלגיה", "Egypt": "מצרים", "France": "צרפת", "Senegal": "סנגל",
                    "Norway": "נורבגיה", "Argentina": "ארגנטינה", "Algeria": "אלג'יריה",
                    "Austria": "אוסטריה", "Jordan": "ירדן", "Ghana": "גאנה", "Panama": "פנמה",
                    "England": "אנגליה", "Croatia": "קרואטיה", "Portugal": "פורטוגל",
                    "Uzbekistan": "אוזבקיסטן", "Colombia": "קולומביה"
                }
                
                def get_team_heb(name):
                    return NATIONAL_TEAM_HEBREW.get(name, name)
                
                def format_date(date_str):
                    try:
                        from datetime import datetime as dt
                        d = dt.strptime(date_str, "%Y-%m-%d")
                        return d.strftime("%d/%m/%Y")
                    except:
                        return date_str
                
                match_options = ["-- בחר משחק --"]
                for m in wc_matches:
                    team1_heb = get_team_heb(m['team1'])
                    team2_heb = get_team_heb(m['team2'])
                    date_fmt = format_date(m['date'])
                    option = f"משחק {m['match_num']}: {team1_heb} נגד {team2_heb} ({date_fmt})"
                    match_options.append(option)
                
                selected_match = st.selectbox("🏆 בחר משחק מונדיאל", match_options, key="worldcup_match")
                
                if selected_match and selected_match != "-- בחר משחק --":
                    match_num = int(selected_match.split(":")[0].replace("משחק ", "").strip())
                    match_data = next((m for m in wc_matches if m['match_num'] == match_num), None)
                    
                    if match_data:
                        team1_heb = get_team_heb(match_data['team1'])
                        team2_heb = get_team_heb(match_data['team2'])
                        st.session_state['home_team_hebrew'] = team1_heb
                        st.session_state['away_team_hebrew'] = team2_heb
                        st.session_state['selected_team_data'] = {'name': match_data['team1']}
                        st.session_state['away_team_data'] = {'name': match_data['team2']}
                        st.session_state['fixture_data'] = {
                            'date': match_data['date'],
                            'time': match_data['time'],
                            'venue': match_data['venue'],
                            'city': match_data['city'],
                            'round': match_data['round']
                        }
                        st.session_state['worldcup_venue'] = f"{match_data['venue']}, {match_data['city']}"
                        
                        try:
                            with open('worldcup_stadiums_mapping.json', 'r', encoding='utf-8') as f:
                                wc_stadiums = json.load(f)
                            stadium_info = wc_stadiums.get('stadiums', {}).get(match_data['venue'], {})
                            if stadium_info.get('map_file'):
                                st.session_state['worldcup_stadium_map'] = stadium_info['map_file']
                            else:
                                st.session_state['worldcup_stadium_map'] = ''
                        except:
                            st.session_state['worldcup_stadium_map'] = ''
                        
                        st.info(f"🏆 **{match_data['round']}** | {team1_heb} נגד {team2_heb}")
                        st.caption(f"📍 {match_data['venue']}, {match_data['city']} | 📅 {format_date(match_data['date'])} {match_data['time']}")
                        
                        wc_categories = ["קטגוריה 3/4", "קטגוריה 3", "קטגוריה 2", "קטגוריה 1"]
                        st.selectbox("🎫 בחר קטגוריית כרטיסים", wc_categories, key="worldcup_category")
                else:
                    st.session_state['fixture_data'] = {}
                    st.session_state['selected_team_data'] = {}
                    st.session_state['away_team_data'] = {}
                    st.session_state['home_team_hebrew'] = ''
                    st.session_state['away_team_hebrew'] = ''
                    st.session_state['worldcup_venue'] = ''
            else:
                from sports_api import get_season_fixtures
                
                teams = []
                fixtures = []
                if selected_league and selected_league != "-- בחר ליגה --":
                    english_league = LEAGUES.get(selected_league, "")
                    teams = get_teams_by_league(english_league)
                    fixtures = get_season_fixtures(english_league)
                else:
                    st.session_state['fixture_data'] = {}
                    st.session_state['selected_team_data'] = {}
                    st.session_state['away_team_data'] = {}
                
                selection_modes = ["🎯 בחר משחק מרשימה", "✏️ בחר קבוצות ידנית"]
                selection_mode = st.radio("אופן בחירה", selection_modes, horizontal=True, key="football_selection_mode", label_visibility="collapsed")
                
                prev_mode = st.session_state.get('_prev_football_mode', '')
                if prev_mode != selection_mode:
                    st.session_state['_prev_football_mode'] = selection_mode
                    st.session_state['fixture_data'] = {}
                    st.session_state['selected_team_data'] = {}
                    st.session_state['away_team_data'] = {}
                    st.session_state['home_team_hebrew'] = ''
                    st.session_state['away_team_hebrew'] = ''
                
                if selection_mode == "🎯 בחר משחק מרשימה" and not fixtures:
                    st.warning("⚠️ אין נתוני משחקים זמינים לליגה זו. השתמש בבחירה ידנית.")
                
                if selection_mode == "🎯 בחר משחק מרשימה" and fixtures:
                    team_filter = st.text_input("🔍 חפש קבוצה (עברית/אנגלית)", placeholder="למשל: ברצלונה, Real Madrid", key="football_team_filter")
                    
                    def format_fixture_date(date_str):
                        try:
                            from datetime import datetime as dt
                            d = dt.strptime(date_str, "%Y-%m-%d")
                            return d.strftime("%d/%m/%Y")
                        except:
                            return date_str
                    
                    from datetime import datetime as dt
                    today = dt.now().date()
                    future_fixtures = []
                    for f in fixtures:
                        try:
                            match_date = dt.strptime(f['date'], "%Y-%m-%d").date()
                            if match_date >= today:
                                future_fixtures.append(f)
                        except:
                            future_fixtures.append(f)
                    
                    filtered_fixtures = future_fixtures
                    if team_filter:
                        filter_lower = team_filter.lower().strip()
                        eng_name = TEAM_HEBREW_NAMES.get(team_filter.strip(), '')
                        
                        filtered_fixtures = []
                        for f in future_fixtures:
                            home_lower = f['home_team'].lower()
                            away_lower = f['away_team'].lower()
                            home_heb = get_hebrew_name(f['home_team']).lower()
                            away_heb = get_hebrew_name(f['away_team']).lower()
                            
                            match_found = (
                                filter_lower in home_lower or filter_lower in away_lower or
                                filter_lower in home_heb or filter_lower in away_heb or
                                home_lower in filter_lower or away_lower in filter_lower or
                                (eng_name and (eng_name.lower() in home_lower or eng_name.lower() in away_lower))
                            )
                            if match_found:
                                filtered_fixtures.append(f)
                    
                    filtered_fixtures = sorted(filtered_fixtures, key=lambda x: x.get('date', ''))
                    
                    match_options = ["-- בחר משחק --"]
                    for f in filtered_fixtures:
                        home_heb = get_hebrew_name(f['home_team'])
                        away_heb = get_hebrew_name(f['away_team'])
                        date_fmt = format_fixture_date(f['date'])
                        time_str = f.get('time', '')[:5] if f.get('time') else ''
                        round_str = f.get('round', '')
                        
                        option = f"⚽ {home_heb} נגד {away_heb}"
                        option += f" | 📅 {date_fmt}"
                        if time_str and time_str != '00:00':
                            option += f" ⏰ {time_str}"
                        match_options.append((option, f))
                    
                    match_display_options = [m[0] if isinstance(m, tuple) else m for m in match_options]
                    filter_key = f"league_match_{len(filtered_fixtures)}_{team_filter[:10] if team_filter else 'all'}"
                    selected_match_idx = st.selectbox("⚽ בחר משחק", range(len(match_display_options)), 
                                                       format_func=lambda x: match_display_options[x], 
                                                       key=filter_key)
                    
                    if selected_match_idx > 0:
                        selected_fixture = match_options[selected_match_idx][1]
                        
                        home_team = next((t for t in teams if t['name'].lower() == selected_fixture['home_team'].lower() or 
                                          selected_fixture['home_team'].lower() in t['name'].lower() or
                                          t['name'].lower() in selected_fixture['home_team'].lower()), None)
                        away_team = next((t for t in teams if t['name'].lower() == selected_fixture['away_team'].lower() or
                                          selected_fixture['away_team'].lower() in t['name'].lower() or
                                          t['name'].lower() in selected_fixture['away_team'].lower()), None)
                        
                        if not home_team:
                            home_team = {'name': selected_fixture['home_team'], 'stadium': '', 'stadium_location': ''}
                        if not away_team:
                            away_team = {'name': selected_fixture['away_team'], 'stadium': '', 'stadium_location': ''}
                        
                        home_heb = get_hebrew_name(selected_fixture['home_team'])
                        away_heb = get_hebrew_name(selected_fixture['away_team'])
                        
                        st.session_state['selected_team_data'] = home_team
                        st.session_state['away_team_data'] = away_team
                        st.session_state['home_team_hebrew'] = home_heb
                        st.session_state['away_team_hebrew'] = away_heb
                        st.session_state['fixture_data'] = {
                            'date': selected_fixture['date'],
                            'time': selected_fixture.get('time', ''),
                            'round': selected_fixture.get('round', '')
                        }
                        
                        st.success(f"✅ **{home_heb}** נגד **{away_heb}**")
                        date_fmt = format_fixture_date(selected_fixture['date'])
                        time_str = selected_fixture.get('time', '')[:5] if selected_fixture.get('time') else ''
                        info_text = f"📅 {date_fmt}"
                        if time_str and time_str != '00:00':
                            info_text += f" | ⏰ {time_str}"
                        if home_team.get('stadium'):
                            info_text += f" | 🏟️ {home_team['stadium']}"
                        st.caption(info_text)
                    else:
                        st.session_state['fixture_data'] = {}
                        st.session_state['selected_team_data'] = {}
                        st.session_state['away_team_data'] = {}
                        st.session_state['home_team_hebrew'] = ''
                        st.session_state['away_team_hebrew'] = ''
                    
                    if not fixtures:
                        st.info("💡 לא נמצאו משחקים לליגה זו. נסה לבחור קבוצות ידנית.")
                
                else:
                    col_team1, col_team2 = st.columns(2)
                    with col_team1:
                        if teams:
                            team_options = ["-- קבוצה מארחת --"] + [f"{get_hebrew_name(t['name'])} ({t['name']})" for t in teams]
                            selected_team1 = st.selectbox("קבוצה מארחת 🏠", team_options, key="football_team1")
                            
                            if selected_team1 and selected_team1 != "-- קבוצה מארחת --":
                                team_name_eng = selected_team1.split("(")[-1].replace(")", "").strip()
                                selected_team = next((t for t in teams if t['name'] == team_name_eng), None)
                                if selected_team:
                                    st.session_state['selected_team_data'] = selected_team
                                    st.session_state['home_team_hebrew'] = selected_team1.split(" (")[0]
                            else:
                                st.session_state['selected_team_data'] = {}
                                st.session_state['home_team_hebrew'] = ''
                        else:
                            st.selectbox("קבוצה מארחת 🏠", ["-- בחר ליגה קודם --"], disabled=True, key="team1_disabled")
                    
                    with col_team2:
                        if teams:
                            team_options2 = ["-- קבוצה אורחת --"] + [f"{get_hebrew_name(t['name'])} ({t['name']})" for t in teams]
                            selected_team2 = st.selectbox("קבוצה אורחת ✈️", team_options2, key="football_team2")
                            
                            if selected_team2 and selected_team2 != "-- קבוצה אורחת --":
                                team_name_eng2 = selected_team2.split("(")[-1].replace(")", "").strip()
                                away_team = next((t for t in teams if t['name'] == team_name_eng2), None)
                                if away_team:
                                    st.session_state['away_team_data'] = away_team
                                st.session_state['away_team_hebrew'] = selected_team2.split(" (")[0]
                            else:
                                st.session_state['away_team_data'] = {}
                                st.session_state['away_team_hebrew'] = ''
                        else:
                            st.selectbox("קבוצה אורחת ✈️", ["-- בחר ליגה קודם --"], disabled=True, key="team2_disabled")
                
                team_data = st.session_state.get('selected_team_data', {})
                home_heb = st.session_state.get('home_team_hebrew', '')
                away_heb = st.session_state.get('away_team_hebrew', '')
                
                if team_data.get('badge') and home_heb:
                    col_badge, col_info = st.columns([1, 3])
                    with col_badge:
                        st.image(team_data['badge'], width=60)
                    with col_info:
                        match_text = f"**{home_heb}**" + (f" נגד **{away_heb}**" if away_heb else "")
                        st.markdown(match_text)
                        if team_data.get('stadium'):
                            st.caption(f"🏟️ {team_data['stadium']}")
                
                current_key = f"{selected_league}_{team_data.get('name', '')}_{st.session_state.get('away_team_data', {}).get('name', '')}"
                if st.session_state.get('fixture_lookup_key') != current_key:
                    st.session_state['fixture_data'] = {}
                    st.session_state['fixture_lookup_key'] = current_key
                
                if team_data.get('name') and st.session_state.get('away_team_data', {}).get('name'):
                    home_name = team_data['name']
                    away_name = st.session_state['away_team_data']['name']
                    english_league = LEAGUES.get(selected_league, "")
                    
                    if not st.session_state.get('fixture_data'):
                        fixture = find_fixture(home_name, away_name, english_league)
                        if fixture and fixture.get('date'):
                            time_str = fixture.get('time', '')
                            if time_str and time_str not in ('00:00:00', '00:00', ''):
                                st.session_state['fixture_data'] = fixture
                                st.success(f"📅 נמצא משחק: {fixture['date']} {time_str[:5]}")
                            elif fixture.get('date'):
                                st.session_state['fixture_data'] = {'date': fixture['date']}
                                st.success(f"📅 נמצא משחק: {fixture['date']}")
        
        elif event_type == "הופעה":
            from concerts_data import get_all_venues
            from concerts_service import search_artists, get_events_by_attraction_id, get_popular_artists, format_concert_for_display, search_events_combined, search_concerts_by_location
            
            st.markdown("##### 🎤 בחירת אמן")
            
            popular_artists = get_popular_artists()
            saved_concerts = get_saved_concerts()
            saved_artists = get_saved_artists()
            
            artist_options = ["-- בחר אמן --"]
            if saved_concerts:
                artist_options.append("⭐ הופעות שמורות")
            artist_options.append("📸 סריקת הופעה מתמונה")
            artist_options += [f"{a['name_he']} ({a['name_en']})" for a in popular_artists]
            if saved_artists:
                artist_options += [f"⭐ {a['name_he']} ({a['name_en']})" for a in saved_artists]
            artist_options.append("🔍 חיפוש אמן אחר...")
            
            selected_artist_option = st.selectbox("🎤 אמן", artist_options, key="concert_artist_select")
            
            artist_id = None
            artist_name_en = ''
            artist_name_he = ''
            
            if selected_artist_option == "⭐ הופעות שמורות":
                st.markdown("##### ⭐ הופעות שמורות")
                
                if saved_concerts:
                    saved_options = ["-- בחר הופעה שמורה --"]
                    for sc in saved_concerts:
                        date_str = sc.get('date', '')
                        if date_str:
                            try:
                                from datetime import datetime as dt
                                date_obj = dt.strptime(date_str, '%Y-%m-%d')
                                date_str = date_obj.strftime('%d/%m/%Y')
                            except:
                                pass
                        artist_name = sc.get('artist_he') or sc.get('artist', '')
                        venue = sc.get('venue', '')
                        city = sc.get('city', '')
                        display = f"{artist_name} @ {venue}"
                        if city:
                            display += f", {city}"
                        if date_str:
                            display += f" ({date_str})"
                        saved_options.append(display)
                    
                    selected_saved_idx = st.selectbox(
                        "⭐ בחר הופעה שמורה",
                        range(len(saved_options)),
                        format_func=lambda i: saved_options[i],
                        key="saved_concert_select"
                    )
                    
                    if selected_saved_idx and selected_saved_idx > 0:
                        selected_saved = saved_concerts[selected_saved_idx - 1]
                        artist_name_en = selected_saved.get('artist', '')
                        artist_name_he = selected_saved.get('artist_he') or artist_name_en
                        
                        st.session_state['concert_artist_en'] = artist_name_en
                        st.session_state['concert_artist_he'] = artist_name_he
                        st.session_state['concert_venue_name'] = selected_saved.get('venue', '')
                        st.session_state['concert_venue_city'] = selected_saved.get('city', '')
                        st.session_state['_selected_concert'] = selected_saved
                        st.session_state['concert_venue_info'] = {
                            'name_he': selected_saved.get('venue', ''),
                            'city_he': selected_saved.get('city', ''),
                            'country': selected_saved.get('country', '')
                        }
                        st.session_state['concert_selected_category'] = selected_saved.get('category', 'General Admission')
                        st.session_state['_from_saved_concert'] = True
                        
                        # Load stadium map from database (base64) or file path
                        if selected_saved.get('stadium_map_data'):
                            import base64
                            from io import BytesIO
                            img_data = base64.b64decode(selected_saved.get('stadium_map_data'))
                            st.session_state['saved_stadium_map_bytes'] = img_data
                        elif selected_saved.get('stadium_map_path') and os.path.exists(selected_saved.get('stadium_map_path')):
                            with open(selected_saved.get('stadium_map_path'), 'rb') as f:
                                st.session_state['saved_stadium_map_bytes'] = f.read()
                        
                        has_map = selected_saved.get('stadium_map_data') or (selected_saved.get('stadium_map_path') and os.path.exists(selected_saved.get('stadium_map_path', '')))
                        st.success(f"✅ נטען: {artist_name_he} @ {selected_saved.get('venue', '')}" + (" (כולל תרשים מושבים)" if has_map else ""))
                        st.caption(f"📍 {selected_saved.get('city', '')}, {selected_saved.get('country', '')} | 🎫 {selected_saved.get('category', 'General Admission')}")
                        
                        if selected_saved.get('date'):
                            st.session_state['_extracted_date'] = selected_saved.get('date', '')
                            st.session_state['_extracted_time'] = selected_saved.get('time', '')
                        
                        if selected_saved.get('url'):
                            st.markdown(f"🔗 [קישור לאירוע]({selected_saved.get('url')})")
                        
                        categories = ['VIP', 'Golden Circle', 'Floor', 'Lower Tier', 'Upper Tier', 'General Admission']
                        default_cat_idx = categories.index(selected_saved.get('category', 'General Admission')) if selected_saved.get('category') in categories else 5
                        selected_cat = st.selectbox("🎫 קטגוריית כרטיסים", categories, index=default_cat_idx, key="saved_concert_category")
                        st.session_state['concert_selected_category'] = selected_cat
                else:
                    st.info("אין הופעות שמורות. שמור הופעות מ'הזנה ידנית' כדי לראות אותן כאן.")
            
            elif selected_artist_option == "📸 סריקת הופעה מתמונה":
                from concert_ocr import extract_concert_data
                
                st.markdown("##### 📸 סריקת הופעה מתמונה")
                st.info("💡 העלה צילום מסך של דף ההופעה והמערכת תחלץ את הפרטים אוטומטית")
                
                col_upload, col_paste = st.columns([3, 1])
                with col_upload:
                    concert_screenshot = st.file_uploader(
                        "📷 העלה צילום מסך של דף האירוע",
                        type=['png', 'jpg', 'jpeg'],
                        key="concert_ocr_upload",
                        help="צלם מסך מאתר המכירות והעלה כאן"
                    )
                with col_paste:
                    concert_paste = paste_image_button("📋 הדבק", key="concert_ocr_paste")
                    if concert_paste.image_data:
                        st.session_state['concert_pasted_image'] = concert_paste.image_data
                        st.image(concert_paste.image_data, caption="תמונה שהודבקה", width=100)
                
                concert_image_to_scan = concert_screenshot or st.session_state.get('concert_pasted_image')
                
                scan_concert_btn = st.button("🔍 סרוק פרטי הופעה", type="primary", use_container_width=True, key="scan_concert_btn")
                
                if st.session_state.get('concert_ocr_result'):
                    ocr_result = st.session_state['concert_ocr_result']
                    st.success("✅ הסריקה הושלמה! הפרטים מולאו בטופס למטה.")
                    
                    st.markdown("**פרטים שזוהו:**")
                    if ocr_result.get('artist_name'):
                        st.write(f"🎤 אמן: {ocr_result.get('artist_name')}")
                    if ocr_result.get('event_name'):
                        st.write(f"🎭 אירוע: {ocr_result.get('event_name')}")
                    if ocr_result.get('event_date'):
                        st.write(f"📅 תאריך: {ocr_result.get('event_date')} {ocr_result.get('event_time', '')}")
                    if ocr_result.get('venue_name'):
                        st.write(f"📍 מקום: {ocr_result.get('venue_name')}, {ocr_result.get('venue_city', '')}")
                    if ocr_result.get('categories'):
                        cats = ocr_result.get('categories', [])
                        if cats:
                            st.write("🎫 קטגוריות:")
                            for cat in cats[:5]:
                                price_str = f" - €{cat.get('price')}" if cat.get('price') else ""
                                st.write(f"  • {cat.get('name', 'כללי')}{price_str}")
                    
                    st.session_state['concert_artist_en'] = ocr_result.get('artist_name', '')
                    st.session_state['concert_artist_he'] = ocr_result.get('artist_name', '')
                    st.session_state['concert_venue_name'] = ocr_result.get('venue_name', '')
                    st.session_state['concert_venue_city'] = ocr_result.get('venue_city', '')
                    st.session_state['concert_venue_info'] = {
                        'name_he': ocr_result.get('venue_name', ''),
                        'city_he': ocr_result.get('venue_city', ''),
                        'country': ocr_result.get('venue_country', '')
                    }
                    st.session_state['_ocr_event_name'] = ocr_result.get('event_name', '')
                    st.session_state['_ocr_event_date'] = ocr_result.get('event_date', '')
                    st.session_state['_ocr_event_time'] = ocr_result.get('event_time', '')
                    st.session_state['_ocr_categories'] = ocr_result.get('categories', [])
                    
                    if ocr_result.get('categories'):
                        cat_names = [c.get('name', 'General') for c in ocr_result.get('categories', [])]
                        selected_ocr_cat = st.selectbox("🎫 בחר קטגוריה", cat_names, key="ocr_category_select")
                        st.session_state['concert_selected_category'] = selected_ocr_cat
                    
                    st.markdown("---")
                    if st.button("⭐ שמור הופעה לשימוש חוזר", use_container_width=True, key="save_ocr_concert"):
                        from models import SavedConcert as SavedConcertModel
                        db = get_db()
                        if db:
                            try:
                                date_str = ocr_result.get('event_date', '')
                                try:
                                    from datetime import datetime as dt
                                    if '/' in date_str:
                                        parsed = dt.strptime(date_str, '%d/%m/%Y')
                                        date_str = parsed.strftime('%Y-%m-%d')
                                except:
                                    pass
                                
                                new_concert = SavedConcertModel(
                                    artist_name=ocr_result.get('artist_name', ''),
                                    artist_name_he=ocr_result.get('artist_name', ''),
                                    event_name=ocr_result.get('event_name', ''),
                                    venue_name=ocr_result.get('venue_name', ''),
                                    city=ocr_result.get('venue_city', ''),
                                    country=ocr_result.get('venue_country', ''),
                                    event_date=date_str,
                                    event_time=ocr_result.get('event_time', ''),
                                    category=st.session_state.get('concert_selected_category', 'General Admission'),
                                    source='ocr'
                                )
                                db.add(new_concert)
                                db.commit()
                                st.success("✅ ההופעה נשמרה! תוכל לבחור אותה מ'הופעות שמורות'.")
                            except Exception as e:
                                db.rollback()
                                st.error(f"❌ שגיאה בשמירה: {str(e)}")
                            finally:
                                db.close()
                
                if scan_concert_btn:
                    if concert_image_to_scan:
                        with st.spinner("🔍 סורק את פרטי ההופעה..."):
                            if concert_screenshot:
                                image_bytes = concert_screenshot.read()
                            else:
                                pasted_img = st.session_state['concert_pasted_image']
                                img_byte_arr = io.BytesIO()
                                pasted_img.save(img_byte_arr, format='PNG')
                                image_bytes = img_byte_arr.getvalue()
                            
                            result = extract_concert_data(image_bytes)
                            
                            if result.get('success'):
                                st.session_state['concert_ocr_result'] = result
                                st.rerun()
                            else:
                                st.error(f"❌ לא הצלחנו לזהות פרטי הופעה: {result.get('error', 'נסה תמונה ברורה יותר')}")
                    else:
                        st.warning("⚠️ יש להעלות צילום מסך לפני הסריקה")
            
            elif selected_artist_option == "🔍 חיפוש אמן אחר...":
                artist_search = st.text_input(
                    "🔍 חיפוש באנגלית", 
                    value=st.session_state.get('_artist_search_query', ''),
                    placeholder="Type artist name in English...",
                    key="artist_search_input"
                )
                st.session_state['_artist_search_query'] = artist_search
                
                if artist_search and len(artist_search.strip()) >= 2:
                    search_key = f"search_{artist_search.strip().lower()}"
                    
                    if st.session_state.get('_last_artist_search') != search_key:
                        st.session_state['_last_artist_search'] = search_key
                        st.session_state['_artist_results'] = []
                        
                        with st.spinner("🔍 מחפש אמנים ב-Ticketmaster..."):
                            result = search_artists(artist_search.strip())
                            if result.get('error'):
                                st.warning(f"⚠️ שגיאה בחיפוש: {result['error']}")
                            elif result.get('artists'):
                                st.session_state['_artist_results'] = result['artists']
                            else:
                                st.info("לא נמצאו אמנים. נסה חיפוש אחר.")
                    
                    artist_results = st.session_state.get('_artist_results', [])
                    
                    if artist_results:
                        search_options = ["-- בחר מתוצאות החיפוש --"]
                        for a in artist_results:
                            events_txt = f" ({a.get('upcoming_events', 0)} הופעות)" if a.get('upcoming_events', 0) > 0 else ""
                            genre_txt = f" • {a.get('genre', '')}" if a.get('genre') else ""
                            search_options.append(f"{a['name']}{genre_txt}{events_txt}")
                        
                        selected_search_idx = st.selectbox(
                            "🎤 בחר אמן מתוצאות החיפוש",
                            range(len(search_options)),
                            format_func=lambda i: search_options[i],
                            key="concert_search_result_select"
                        )
                        
                        if selected_search_idx and selected_search_idx > 0:
                            selected = artist_results[selected_search_idx - 1]
                            artist_id = selected.get('id', '')
                            artist_name_en = selected.get('name', '')
                            artist_name_he = artist_name_en
                            selected_genre = selected.get('genre', '')
                            selected_image = selected.get('image_url', '')
                            
                            st.session_state['_search_selected_artist_id'] = artist_id
                            st.session_state['_search_selected_artist_name'] = artist_name_en
                            st.session_state['_search_selected_artist_genre'] = selected_genre
                            st.session_state['_search_selected_artist_image'] = selected_image
                            
                            if st.button("⭐ הוסף לאמנים שלי", key="save_artist_btn", use_container_width=True):
                                success = save_artist_to_favorites(
                                    name_en=artist_name_en,
                                    name_he=artist_name_en,
                                    ticketmaster_id=artist_id,
                                    genre=selected_genre,
                                    image_url=selected_image
                                )
                                if success:
                                    st.success(f"✅ האמן {artist_name_en} נוסף לרשימה שלך!")
                                    st.rerun()
                                else:
                                    st.error("❌ שגיאה בהוספת האמן")
                        elif st.session_state.get('_search_selected_artist_id'):
                            artist_id = st.session_state.get('_search_selected_artist_id', '')
                            artist_name_en = st.session_state.get('_search_selected_artist_name', '')
                            artist_name_he = artist_name_en
            
            elif selected_artist_option and selected_artist_option not in ["-- בחר אמן --", "🔍 חיפוש אמן אחר..."]:
                is_saved_artist = selected_artist_option.startswith("⭐ ")
                clean_option = selected_artist_option[2:] if is_saved_artist else selected_artist_option
                
                artist_name_he = clean_option.split(" (")[0]
                artist_name_en = clean_option.split("(")[-1].replace(")", "").strip()
                
                if is_saved_artist:
                    artist_info = next((a for a in saved_artists if a['name_en'] == artist_name_en), None)
                else:
                    artist_info = next((a for a in popular_artists if a['name_en'] == artist_name_en), None)
                
                if artist_info:
                    artist_id = artist_info['id']
                    st.session_state['_artist_results'] = []
                    st.session_state['_artist_search_query'] = ''
                    st.session_state['_search_selected_artist_id'] = ''
                    st.session_state['_search_selected_artist_name'] = ''
            
            # For search mode, always use session state values if available
            if selected_artist_option == "🔍 חיפוש אמן אחר..." and st.session_state.get('_search_selected_artist_id'):
                artist_id = st.session_state.get('_search_selected_artist_id', '')
                artist_name_en = st.session_state.get('_search_selected_artist_name', '')
                artist_name_he = artist_name_en
            
            if artist_id or (artist_name_en and selected_artist_option == "🔍 חיפוש אמן אחר..."):
                if artist_name_en:
                    st.session_state['concert_artist_en'] = artist_name_en
                    st.session_state['concert_artist_he'] = artist_name_he or artist_name_en
                    st.session_state['_selected_artist_id'] = artist_id or ''
                
                events_key = f"events_combined_{artist_id or artist_name_en}"
                
                if st.session_state.get('_last_events_fetch') != events_key:
                    st.session_state['_last_events_fetch'] = events_key
                    
                    with st.spinner(f"🎫 מחפש הופעות של {artist_name_he or artist_name_en} (Ticketmaster + מקורות נוספים)..."):
                        events_result = search_events_combined(artist_name_en, artist_id or '', size=50)
                        
                        if events_result.get('error'):
                            st.warning(f"⚠️ שגיאה: {events_result['error']}")
                            st.session_state['_live_concerts'] = []
                        elif events_result.get('concerts'):
                            st.session_state['_live_concerts'] = events_result['concerts']
                            sources_text = ""
                            if events_result.get('sources'):
                                sources_text = " (ממספר מקורות)"
                            st.success(f"🎫 נמצאו {events_result['total']} הופעות קרובות של {artist_name_he or artist_name_en} באירופה{sources_text}")
                        else:
                            st.session_state['_live_concerts'] = []
                            st.info(f"🎤 {artist_name_he or artist_name_en} - לא נמצאו הופעות קרובות באירופה")
            else:
                if selected_artist_option == "-- בחר אמן --":
                    st.session_state['_live_concerts'] = []
                    st.session_state['_selected_artist_id'] = ''
                    st.session_state['concert_artist_en'] = ''
                    st.session_state['concert_artist_he'] = ''
                    st.session_state['_last_events_fetch'] = ''
            
            live_concerts = st.session_state.get('_live_concerts', [])
            
            if live_concerts:
                concert_options = ["-- בחר הופעה --"]
                for i, c in enumerate(live_concerts):
                    date_str = c.get('date', '')
                    if date_str:
                        try:
                            from datetime import datetime as dt
                            date_obj = dt.strptime(date_str, '%Y-%m-%d')
                            date_str = date_obj.strftime('%d/%m/%Y')
                        except:
                            pass
                    time_str = c.get('time', '')
                    venue = c.get('venue', '')
                    city = c.get('city', '')
                    country = c.get('country', '')
                    display = f"{date_str} {time_str} - {venue}, {city} ({country})"
                    concert_options.append(display)
                
                concert_options.append("✏️ הזנה ידנית...")
                manual_entry_idx = len(concert_options) - 1
                
                prev_venue = st.session_state.get('_prev_concert_venue', '')
                selected_concert_idx = st.selectbox(
                    "🏟️ מקום ההופעה", 
                    range(len(concert_options)),
                    format_func=lambda i: concert_options[i],
                    key="concert_venue_dropdown"
                )
                
                if selected_concert_idx == manual_entry_idx:
                    st.session_state['_manual_concert_entry'] = True
                    
                    # Check if coming from saved concert - preserve data
                    from_saved = st.session_state.get('_from_saved_concert', False)
                    saved_concert_data = st.session_state.get('_selected_concert', {}) if from_saved else {}
                    
                    if not from_saved:
                        st.session_state['_selected_concert'] = {}
                        st.session_state['concert_venue_name'] = ''
                        st.session_state['concert_venue_city'] = ''
                        st.session_state['_concert_venue_id'] = ''
                        st.session_state['concert_venue_info'] = {}
                    else:
                        # Pre-populate extracted_concert with saved concert data for form fields
                        if saved_concert_data and '_extracted_concert' not in st.session_state:
                            st.session_state['_extracted_concert'] = {
                                'venue': saved_concert_data.get('venue', ''),
                                'city': saved_concert_data.get('city', ''),
                                'country': saved_concert_data.get('country', ''),
                                'date': saved_concert_data.get('date', ''),
                                'time': saved_concert_data.get('time', ''),
                                'url': saved_concert_data.get('url', ''),
                                'source': 'saved'
                            }
                    
                    st.markdown("---")
                    st.markdown("##### ✏️ הזנה ידנית של פרטי ההופעה")
                    
                    st.markdown("**🔗 יש לך לינק לאירוע?** הדבק אותו ונחלץ את הפרטים אוטומטית:")
                    
                    url_col1, url_col2 = st.columns([4, 1])
                    with url_col1:
                        event_url = st.text_input("🔗 לינק לאירוע", key="manual_event_url", placeholder="https://www.ticketmaster.com/...", label_visibility="collapsed")
                    with url_col2:
                        extract_btn = st.button("🔍 חלץ", key="extract_url_btn", use_container_width=True)
                    
                    if extract_btn and event_url:
                        from concerts_service import extract_concert_from_url
                        with st.spinner("מחלץ פרטי אירוע..."):
                            result = extract_concert_from_url(event_url)
                            if result.get('error'):
                                st.error(f"❌ {result['error']}")
                            elif result.get('concert'):
                                extracted = result['concert']
                                st.session_state['_extracted_concert'] = extracted
                                st.session_state['_from_saved_concert'] = False  # Reset flag since we got new data
                                st.success(f"✅ נמצא! מקור: {extracted.get('source', 'Unknown')}")
                                st.rerun()
                    
                    extracted = st.session_state.get('_extracted_concert', {})
                    
                    manual_venue = st.text_input("🏟️ שם מקום ההופעה *", 
                        value=extracted.get('venue', ''),
                        key="manual_venue_name", 
                        placeholder="לדוגמה: O2 Arena")
                    
                    mcol1, mcol2 = st.columns(2)
                    with mcol1:
                        manual_city = st.text_input("🌆 עיר", 
                            value=extracted.get('city', ''),
                            key="manual_venue_city", 
                            placeholder="לדוגמה: לונדון")
                    with mcol2:
                        manual_country = st.text_input("🌍 מדינה", 
                            value=extracted.get('country', ''),
                            key="manual_venue_country", 
                            placeholder="לדוגמה: אנגליה")
                    
                    if extracted.get('date') or extracted.get('time'):
                        st.caption(f"📅 תאריך שחולץ: {extracted.get('date', '')} {extracted.get('time', '')}")
                        st.session_state['_extracted_date'] = extracted.get('date', '')
                        st.session_state['_extracted_time'] = extracted.get('time', '')
                    
                    if manual_venue:
                        st.session_state['concert_venue_name'] = manual_venue
                        st.session_state['concert_venue_city'] = manual_city or ''
                        st.session_state['_concert_venue_id'] = ''
                        st.session_state['concert_venue_info'] = {
                            'name_he': manual_venue,
                            'city_he': manual_city or '',
                            'country': manual_country or ''
                        }
                        
                        categories = ['VIP', 'Golden Circle', 'Floor', 'Lower Tier', 'Upper Tier', 'General Admission']
                        selected_cat = st.selectbox("🎫 קטגוריית כרטיסים", categories, key="concert_category_dropdown")
                        st.session_state['concert_selected_category'] = selected_cat
                        
                        st.markdown("---")
                        if st.button("⭐ שמור להופעות קבועות", key="save_concert_btn_1", use_container_width=True):
                            artist_en = st.session_state.get('concert_artist_en', '')
                            artist_he = st.session_state.get('concert_artist_he', artist_en)
                            if artist_en and manual_venue:
                                map_data = None
                                map_mime = None
                                if 'pasted_stadium_map' in st.session_state and st.session_state['pasted_stadium_map']:
                                    try:
                                        from io import BytesIO
                                        img_buffer = BytesIO()
                                        st.session_state['pasted_stadium_map'].save(img_buffer, format='PNG')
                                        map_data = img_buffer.getvalue()
                                        map_mime = 'image/png'
                                    except Exception as e:
                                        st.warning(f"⚠️ לא הצלחתי לשמור את התרשים: {e}")
                                
                                success = save_concert_to_favorites(
                                    artist_name=artist_en,
                                    artist_name_he=artist_he,
                                    venue_name=manual_venue,
                                    city=manual_city,
                                    country=manual_country,
                                    event_date=extracted.get('date'),
                                    event_time=extracted.get('time'),
                                    event_url=event_url if event_url else extracted.get('url'),
                                    category=selected_cat,
                                    source=extracted.get('source', 'manual'),
                                    stadium_map_data=map_data,
                                    stadium_map_mime=map_mime
                                )
                                if success:
                                    st.success("✅ ההופעה נשמרה לקבועות!" + (" (כולל תרשים מושבים)" if map_data else ""))
                                else:
                                    st.error("❌ שגיאה בשמירת ההופעה")
                            else:
                                st.warning("⚠️ נא לבחור אמן ולהזין שם מקום ההופעה")
                    else:
                        st.warning("⚠️ נא להזין שם מקום ההופעה")
                    
                elif selected_concert_idx and selected_concert_idx > 0:
                    st.session_state['_manual_concert_entry'] = False
                    selected_concert = live_concerts[selected_concert_idx - 1]
                    st.session_state['concert_venue_name'] = selected_concert.get('venue', '')
                    st.session_state['concert_venue_city'] = selected_concert.get('city', '')
                    st.session_state['_concert_venue_id'] = selected_concert.get('id', '')
                    st.session_state['_selected_concert'] = selected_concert
                    
                    st.caption(f"📍 {selected_concert.get('venue', '')}, {selected_concert.get('city', '')} ({selected_concert.get('country', '')})")
                    
                    if selected_concert.get('address'):
                        st.caption(f"📮 כתובת: {selected_concert.get('address', '')}")
                    
                    if selected_concert.get('capacity'):
                        st.caption(f"👥 קיבולת: {selected_concert.get('capacity', ''):,} אנשים")
                    
                    if selected_concert.get('price_min') or selected_concert.get('price_max'):
                        price_info = f"💰 מחירים: {selected_concert.get('price_min', 'N/A')} - {selected_concert.get('price_max', 'N/A')} {selected_concert.get('currency', 'EUR')}"
                        st.caption(price_info)
                    
                    if selected_concert.get('url'):
                        st.markdown(f"🎫 [מעבר לעמוד ההזמנה של Ticketmaster]({selected_concert.get('url')})")
                    
                    categories = ['VIP', 'Golden Circle', 'Floor', 'Lower Tier', 'Upper Tier', 'General Admission']
                    selected_cat = st.selectbox("🎫 קטגוריית כרטיסים", categories, key="concert_category_dropdown")
                    st.session_state['concert_selected_category'] = selected_cat
                    
                    # Save to favorites button for API concerts
                    if st.button("⭐ שמור להופעות קבועות", key="save_concert_btn_api", use_container_width=True):
                        artist_en = st.session_state.get('concert_artist_en', '')
                        artist_he = st.session_state.get('concert_artist_he', artist_en)
                        if artist_en and selected_concert.get('venue'):
                            success = save_concert_to_favorites(
                                artist_name=artist_en,
                                artist_name_he=artist_he,
                                venue_name=selected_concert.get('venue', ''),
                                city=selected_concert.get('city', ''),
                                country=selected_concert.get('country', ''),
                                event_date=selected_concert.get('date'),
                                event_time=selected_concert.get('time'),
                                event_url=selected_concert.get('url'),
                                category=selected_cat,
                                source='ticketmaster'
                            )
                            if success:
                                st.success("✅ ההופעה נשמרה לקבועות!")
                            else:
                                st.error("❌ שגיאה בשמירת ההופעה")
                        else:
                            st.warning("⚠️ נא לבחור אמן ומקום הופעה")
                elif prev_venue != selected_concert_idx:
                    st.session_state['concert_venue_info'] = {}
                    st.session_state['concert_venue_name'] = ''
                    st.session_state['concert_venue_city'] = ''
                    st.session_state['concert_selected_category'] = ''
                    st.session_state['_concert_venue_id'] = ''
                    st.session_state['_selected_concert'] = {}
                    st.session_state['_manual_concert_entry'] = False
                st.session_state['_prev_concert_venue'] = selected_concert_idx
            else:
                # Check if we have a selected artist from search with no European concerts
                has_selected_search_artist = (
                    selected_artist_option == "🔍 חיפוש אמן אחר..." and 
                    st.session_state.get('_search_selected_artist_id')
                )
                
                if has_selected_search_artist:
                    # Artist from search has no European concerts - go directly to manual entry
                    st.session_state['_manual_concert_entry'] = True
                    
                    st.markdown("---")
                    st.markdown("##### ✏️ הזנה ידנית של פרטי ההופעה")
                    st.info("💡 לא נמצאו הופעות אירופאיות. ניתן להזין פרטים ידנית או להדביק לינק לאירוע.")
                    
                    st.markdown("**🔗 יש לך לינק לאירוע?** הדבק אותו ונחלץ את הפרטים אוטומטית:")
                    
                    url_col1, url_col2 = st.columns([4, 1])
                    with url_col1:
                        event_url = st.text_input("🔗 לינק לאירוע", key="manual_event_url_search", placeholder="https://www.ticketmaster.com/...", label_visibility="collapsed")
                    with url_col2:
                        extract_btn = st.button("🔍 חלץ", key="extract_url_btn_search", use_container_width=True)
                    
                    if extract_btn and event_url:
                        from concerts_service import extract_concert_from_url
                        with st.spinner("מחלץ פרטי אירוע..."):
                            result = extract_concert_from_url(event_url)
                            if result.get('error'):
                                st.error(f"❌ {result['error']}")
                            elif result.get('concert'):
                                extracted = result['concert']
                                st.session_state['_extracted_concert'] = extracted
                                st.success(f"✅ נמצא! מקור: {extracted.get('source', 'Unknown')}")
                                st.rerun()
                    
                    extracted = st.session_state.get('_extracted_concert', {})
                    
                    manual_venue = st.text_input("🏟️ שם מקום ההופעה *", 
                        value=extracted.get('venue', ''),
                        key="manual_venue_name_search", 
                        placeholder="לדוגמה: O2 Arena")
                    
                    mcol1, mcol2 = st.columns(2)
                    with mcol1:
                        manual_city = st.text_input("🌆 עיר", 
                            value=extracted.get('city', ''),
                            key="manual_venue_city_search", 
                            placeholder="לדוגמה: לונדון")
                    with mcol2:
                        manual_country = st.text_input("🌍 מדינה", 
                            value=extracted.get('country', ''),
                            key="manual_venue_country_search", 
                            placeholder="לדוגמה: אנגליה")
                    
                    if extracted.get('date') or extracted.get('time'):
                        st.caption(f"📅 תאריך שחולץ: {extracted.get('date', '')} {extracted.get('time', '')}")
                        st.session_state['_extracted_date'] = extracted.get('date', '')
                        st.session_state['_extracted_time'] = extracted.get('time', '')
                    
                    if manual_venue:
                        st.session_state['concert_venue_name'] = manual_venue
                        st.session_state['concert_venue_city'] = manual_city or ''
                        st.session_state['_concert_venue_id'] = ''
                        st.session_state['concert_venue_info'] = {
                            'name_he': manual_venue,
                            'city_he': manual_city or '',
                            'country': manual_country or ''
                        }
                        
                        categories = ['VIP', 'Golden Circle', 'Floor', 'Lower Tier', 'Upper Tier', 'General Admission']
                        selected_cat = st.selectbox("🎫 קטגוריית כרטיסים", categories, key="concert_category_dropdown_search")
                        st.session_state['concert_selected_category'] = selected_cat
                        
                        st.markdown("---")
                        if st.button("⭐ שמור להופעות קבועות", key="save_concert_btn_2", use_container_width=True):
                            artist_en = st.session_state.get('concert_artist_en', '') or st.session_state.get('_search_selected_artist_name', '')
                            artist_he = st.session_state.get('concert_artist_he', artist_en)
                            if artist_en and manual_venue:
                                map_data = None
                                map_mime = None
                                if 'pasted_stadium_map' in st.session_state and st.session_state['pasted_stadium_map']:
                                    try:
                                        from io import BytesIO
                                        img_buffer = BytesIO()
                                        st.session_state['pasted_stadium_map'].save(img_buffer, format='PNG')
                                        map_data = img_buffer.getvalue()
                                        map_mime = 'image/png'
                                    except:
                                        pass
                                success = save_concert_to_favorites(
                                    artist_name=artist_en,
                                    artist_name_he=artist_he,
                                    venue_name=manual_venue,
                                    city=manual_city,
                                    country=manual_country,
                                    event_date=extracted.get('date'),
                                    event_time=extracted.get('time'),
                                    event_url=event_url if event_url else extracted.get('url'),
                                    category=selected_cat,
                                    source=extracted.get('source', 'manual'),
                                    stadium_map_data=map_data,
                                    stadium_map_mime=map_mime
                                )
                                if success:
                                    st.success("✅ ההופעה נשמרה לקבועות!" + (" (כולל תרשים מושבים)" if map_data else ""))
                                else:
                                    st.error("❌ שגיאה בשמירת ההופעה")
                            else:
                                st.warning("⚠️ נא לבחור אמן ולהזין שם מקום ההופעה")
                    else:
                        st.warning("⚠️ נא להזין שם מקום ההופעה")
                else:
                    # No artist selected - show generic venue list
                    venues = get_all_venues()
                    venue_options = ["-- בחר מקום הופעה --"] + [f"{v['name_he']} - {v['city_he']}" for v in venues] + ["✏️ הזנה ידנית..."]
                    manual_venue_idx = len(venue_options) - 1
                    
                    prev_venue = st.session_state.get('_prev_concert_venue', '')
                    selected_venue = st.selectbox("🏟️ מקום ההופעה", venue_options, key="concert_venue_dropdown")
                
                    if selected_venue == "✏️ הזנה ידנית...":
                        st.session_state['_manual_concert_entry'] = True
                        st.session_state['_selected_concert'] = {}
                        st.session_state['concert_venue_name'] = ''
                        st.session_state['concert_venue_city'] = ''
                        st.session_state['_concert_venue_id'] = ''
                        st.session_state['concert_venue_info'] = {}
                        
                        st.markdown("---")
                        st.markdown("##### ✏️ הזנה ידנית של פרטי ההופעה")
                        
                        st.markdown("**🔗 יש לך לינק לאירוע?** הדבק אותו ונחלץ את הפרטים אוטומטית:")
                        
                        url_col1, url_col2 = st.columns([4, 1])
                        with url_col1:
                            event_url = st.text_input("🔗 לינק לאירוע", key="manual_event_url_fallback", placeholder="https://www.ticketmaster.com/...", label_visibility="collapsed")
                        with url_col2:
                            extract_btn = st.button("🔍 חלץ", key="extract_url_btn_fallback", use_container_width=True)
                        
                        if extract_btn and event_url:
                            from concerts_service import extract_concert_from_url
                            with st.spinner("מחלץ פרטי אירוע..."):
                                result = extract_concert_from_url(event_url)
                                if result.get('error'):
                                    st.error(f"❌ {result['error']}")
                                elif result.get('concert'):
                                    extracted = result['concert']
                                    st.session_state['_extracted_concert'] = extracted
                                    st.success(f"✅ נמצא! מקור: {extracted.get('source', 'Unknown')}")
                                    st.rerun()
                        
                        extracted = st.session_state.get('_extracted_concert', {})
                        
                        manual_venue = st.text_input("🏟️ שם מקום ההופעה *", 
                            value=extracted.get('venue', ''),
                            key="manual_venue_name_fallback", 
                            placeholder="לדוגמה: O2 Arena")
                        
                        mcol1, mcol2 = st.columns(2)
                        with mcol1:
                            manual_city = st.text_input("🌆 עיר", 
                                value=extracted.get('city', ''),
                                key="manual_venue_city_fallback", 
                                placeholder="לדוגמה: לונדון")
                        with mcol2:
                            manual_country = st.text_input("🌍 מדינה", 
                                value=extracted.get('country', ''),
                                key="manual_venue_country_fallback", 
                                placeholder="לדוגמה: אנגליה")
                        
                        if extracted.get('date') or extracted.get('time'):
                            st.caption(f"📅 תאריך שחולץ: {extracted.get('date', '')} {extracted.get('time', '')}")
                            st.session_state['_extracted_date'] = extracted.get('date', '')
                            st.session_state['_extracted_time'] = extracted.get('time', '')
                        
                        if manual_venue:
                            st.session_state['concert_venue_name'] = manual_venue
                            st.session_state['concert_venue_city'] = manual_city or ''
                            st.session_state['_concert_venue_id'] = ''
                            st.session_state['concert_venue_info'] = {
                                'name_he': manual_venue,
                                'city_he': manual_city or '',
                                'country': manual_country or ''
                            }
                            
                            categories = ['VIP', 'Golden Circle', 'Floor', 'Lower Tier', 'Upper Tier', 'General Admission']
                            selected_cat = st.selectbox("🎫 קטגוריית כרטיסים", categories, key="concert_category_dropdown")
                            st.session_state['concert_selected_category'] = selected_cat
                            
                            st.markdown("---")
                            manual_artist_name = st.text_input("🎤 שם אמן (באנגלית)", key="manual_artist_fallback", placeholder="לדוגמה: Ed Sheeran")
                            if st.button("⭐ שמור להופעות קבועות", key="save_concert_btn_3", use_container_width=True):
                                if manual_artist_name and manual_venue:
                                    map_data = None
                                    map_mime = None
                                    if 'pasted_stadium_map' in st.session_state and st.session_state['pasted_stadium_map']:
                                        try:
                                            from io import BytesIO
                                            img_buffer = BytesIO()
                                            st.session_state['pasted_stadium_map'].save(img_buffer, format='PNG')
                                            map_data = img_buffer.getvalue()
                                            map_mime = 'image/png'
                                        except:
                                            pass
                                    success = save_concert_to_favorites(
                                        artist_name=manual_artist_name,
                                        artist_name_he=manual_artist_name,
                                        venue_name=manual_venue,
                                        city=manual_city,
                                        country=manual_country,
                                        event_date=extracted.get('date'),
                                        event_time=extracted.get('time'),
                                        event_url=event_url if event_url else extracted.get('url'),
                                        category=selected_cat,
                                        source=extracted.get('source', 'manual'),
                                        stadium_map_data=map_data,
                                        stadium_map_mime=map_mime
                                    )
                                    if success:
                                        st.success("✅ ההופעה נשמרה לקבועות!" + (" (כולל תרשים מושבים)" if map_data else ""))
                                    else:
                                        st.error("❌ שגיאה בשמירת ההופעה")
                                else:
                                    st.warning("⚠️ נא להזין שם אמן ושם מקום ההופעה")
                        else:
                            st.warning("⚠️ נא להזין שם מקום ההופעה")
                            
                    elif selected_venue and selected_venue != "-- בחר מקום הופעה --":
                        venue_he = selected_venue.split(" - ")[0]
                        venue_info = next((v for v in venues if v['name_he'] == venue_he), None)
                        if venue_info:
                            st.session_state['concert_venue_info'] = venue_info
                            st.session_state['concert_venue_name'] = venue_info['name_he']
                            st.session_state['concert_venue_city'] = venue_info['city_he']
                            st.session_state['_concert_venue_id'] = venue_info.get('id', '')
                            
                            st.caption(f"📍 {venue_info['city_he']}, {venue_info['country']} | 👥 קיבולת: {venue_info['capacity']:,}")
                            
                            categories = venue_info.get('categories', ['General Admission'])
                            selected_cat = st.selectbox("🎫 קטגוריית כרטיסים", categories, key="concert_category_dropdown")
                            st.session_state['concert_selected_category'] = selected_cat
                    elif prev_venue != selected_venue:
                        st.session_state['concert_venue_info'] = {}
                        st.session_state['concert_venue_name'] = ''
                        st.session_state['concert_venue_city'] = ''
                        st.session_state['concert_selected_category'] = ''
                        st.session_state['_concert_venue_id'] = ''
                    st.session_state['_prev_concert_venue'] = selected_venue
        
        default_event_name = rd.get('event_name', '')
        team_data = st.session_state.get('selected_team_data', {})
        home_heb = st.session_state.get('home_team_hebrew', '')
        away_heb = st.session_state.get('away_team_hebrew', '')
        if event_type == "כדורגל" and home_heb and not default_event_name:
            if away_heb:
                default_event_name = f"{home_heb} נגד {away_heb}"
            else:
                default_event_name = f"{home_heb} נגד "
        elif event_type == "הופעה" and not default_event_name:
            ocr_event_name = st.session_state.get('_ocr_event_name', '')
            if ocr_event_name:
                default_event_name = ocr_event_name
            else:
                artist_he = st.session_state.get('concert_artist_he', '')
                venue_name = st.session_state.get('concert_venue_name', '')
                if artist_he and venue_name:
                    default_event_name = f"הופעה של {artist_he} ב{venue_name}"
                elif artist_he:
                    default_event_name = f"הופעה של {artist_he}"
        
        event_name = st.text_input("שם האירוע", value=default_event_name, placeholder="לדוגמה: Real Madrid vs Barcelona")
        
        fixture_data = st.session_state.get('fixture_data', {})
        selected_concert = st.session_state.get('_selected_concert', {})
        default_date = None
        default_time = None
        
        if fixture_data.get('date'):
            try:
                from datetime import datetime as dt
                default_date = dt.strptime(fixture_data['date'], "%Y-%m-%d").date()
            except:
                pass
        if fixture_data.get('time'):
            try:
                from datetime import datetime as dt
                time_str = fixture_data['time'][:5] if len(fixture_data['time']) >= 5 else fixture_data['time']
                default_time = dt.strptime(time_str, "%H:%M").time()
            except:
                pass
        
        if selected_concert.get('date') and not default_date:
            try:
                from datetime import datetime as dt
                default_date = dt.strptime(selected_concert['date'], "%Y-%m-%d").date()
            except:
                pass
        if selected_concert.get('time') and not default_time:
            try:
                from datetime import datetime as dt
                time_str = selected_concert['time'][:5] if len(selected_concert['time']) >= 5 else selected_concert['time']
                default_time = dt.strptime(time_str, "%H:%M").time()
            except:
                pass
        
        extracted_date = st.session_state.get('_extracted_date', '')
        extracted_time = st.session_state.get('_extracted_time', '')
        if extracted_date and not default_date:
            try:
                from datetime import datetime as dt
                default_date = dt.strptime(extracted_date, "%Y-%m-%d").date()
            except:
                pass
        if extracted_time and not default_time:
            try:
                from datetime import datetime as dt
                time_str = extracted_time[:5] if len(extracted_time) >= 5 else extracted_time
                default_time = dt.strptime(time_str, "%H:%M").time()
            except:
                pass
        
        ocr_date = st.session_state.get('_ocr_event_date', '')
        ocr_time = st.session_state.get('_ocr_event_time', '')
        if ocr_date and not default_date:
            try:
                from datetime import datetime as dt
                default_date = dt.strptime(ocr_date, "%d/%m/%Y").date()
            except:
                try:
                    default_date = dt.strptime(ocr_date, "%Y-%m-%d").date()
                except:
                    pass
        if ocr_time and not default_time:
            try:
                from datetime import datetime as dt
                time_str = ocr_time[:5] if len(ocr_time) >= 5 else ocr_time
                default_time = dt.strptime(time_str, "%H:%M").time()
            except:
                pass
        
        col_date, col_time = st.columns(2)
        with col_date:
            if default_date:
                event_date = st.date_input("תאריך האירוע", value=default_date)
            else:
                event_date = st.date_input("תאריך האירוע")
        with col_time:
            if default_time:
                event_time = st.time_input("שעת האירוע", value=default_time)
            else:
                event_time = st.time_input("שעת האירוע")
        
        date_status = st.radio(
            "סטטוס התאריך",
            options=["התאריך אינו סופי", "התאריך הינו סופי"],
            index=0,
            horizontal=True,
            key="date_status_radio"
        )
        is_date_final = (date_status == "התאריך הינו סופי")
        
        seats_together = st.checkbox("🪑 ישיבה 3 יחד", value=False, key="seats_together_checkbox")
        
        default_venue = rd.get('venue', '')
        if st.session_state.get('worldcup_venue'):
            default_venue = st.session_state['worldcup_venue']
        elif event_type == "כדורגל" and team_data.get('stadium') and not default_venue:
            default_venue = f"{team_data['stadium']}, {team_data.get('stadium_location', '')}"
        elif event_type == "הופעה" and not default_venue:
            venue_name = st.session_state.get('concert_venue_name', '')
            venue_city = st.session_state.get('concert_venue_city', '')
            if venue_name and venue_city:
                default_venue = f"{venue_name}, {venue_city}"
        
        venue = st.text_input("מקום האירוע / אצטדיון", value=default_venue, placeholder="לדוגמה: Santiago Bernabeu, Madrid")
        
        st.markdown("**🗺️ תרשים מושבים (Seat Map)**")
        
        auto_stadium_map = None
        
        wc_map = st.session_state.get('worldcup_stadium_map', '')
        if wc_map and os.path.exists(wc_map):
            auto_stadium_map = wc_map
            venue_name = st.session_state.get('fixture_data', {}).get('venue', '')
            st.success(f"✅ נמצאה מפת אצטדיון FIFA עבור {venue_name}")
            st.image(wc_map, caption="מפת קטגוריות מושבים - FIFA World Cup 2026", use_container_width=True)
        elif event_type == "כדורגל" and team_data.get('name'):
            team_name_eng = team_data.get('name', '')
            map_path = get_team_map_path(team_name_eng)
            if map_path and os.path.exists(map_path):
                auto_stadium_map = map_path
                st.success(f"✅ נמצאה מפת אצטדיון אוטומטית עבור {st.session_state.get('home_team_hebrew', team_name_eng)}")
                st.image(map_path, caption="מפת אצטדיון", use_container_width=True)
        elif event_type == "הופעה":
            from concerts_data import get_venue_map_path, CONCERT_DEFAULT_BG
            
            concert_venue_info = st.session_state.get('concert_venue_info', {})
            concert_venue_id = st.session_state.get('_concert_venue_id', '')
            selected_concert = st.session_state.get('_selected_concert', {})
            
            if selected_concert:
                venue_name = selected_concert.get('venue', '')
                venue_city = selected_concert.get('city', '')
                capacity = selected_concert.get('capacity', 0)
                concert_url = selected_concert.get('url', '')
                
                concert_map_path = None
                if concert_venue_id:
                    # Check for existing maps with any extension
                    for map_ext in ['png', 'jpg', 'jpeg', 'gif', 'webp']:
                        possible_map = f"attached_assets/concert_venue_maps/{concert_venue_id}.{map_ext}"
                        if os.path.exists(possible_map):
                            concert_map_path = possible_map
                            break
                
                if concert_map_path:
                    auto_stadium_map = concert_map_path
                    st.success(f"✅ נמצאה מפת מושבים שמורה עבור {venue_name}")
                    st.image(concert_map_path, caption=f"מפת מושבים - {venue_name}", use_container_width=True)
                    
                    if st.button("🗑️ מחק מפה שמורה", key=f"delete_map_{concert_venue_id}"):
                        try:
                            os.remove(concert_map_path)
                            st.success("✅ המפה נמחקה!")
                            st.rerun()
                        except Exception as e:
                            st.error(f"❌ שגיאה במחיקה: {str(e)}")
                else:
                    if capacity:
                        st.info(f"🎤 **{venue_name}, {venue_city}** - קיבולת: {capacity:,} מושבים")
                    else:
                        st.info(f"🎤 **{venue_name}, {venue_city}**")
                    
                    # Try to auto-fetch map from Ticketmaster CDN
                    venue_id_tm = selected_concert.get('venue_id', '')
                    auto_map_found = False
                    
                    if venue_id_tm and not concert_map_path:
                        # Try common Ticketmaster seatmap URL patterns
                        map_patterns = [
                            f"https://media.ticketmaster.co.uk/tm/en-gb/tmimages/venue/maps/uk2/{venue_id_tm}s.gif",
                            f"https://media.ticketmaster.eu/tm/en-eu/tmimages/venue/maps/eu/{venue_id_tm}s.gif",
                            f"https://s1.ticketm.net/tm/en-us/tmimages/venue/maps/nyc/{venue_id_tm}s.gif",
                        ]
                        
                        for pattern_url in map_patterns:
                            try:
                                headers = {'User-Agent': 'Mozilla/5.0'}
                                test_resp = requests.head(pattern_url, headers=headers, timeout=5)
                                if test_resp.status_code == 200:
                                    # Found a map! Download it
                                    img_resp = requests.get(pattern_url, headers=headers, timeout=15)
                                    if img_resp.status_code == 200:
                                        os.makedirs('attached_assets/concert_venue_maps', exist_ok=True)
                                        ext = 'gif' if 'gif' in pattern_url else 'png'
                                        save_path = f'attached_assets/concert_venue_maps/{concert_venue_id}.{ext}'
                                        with open(save_path, 'wb') as f:
                                            f.write(img_resp.content)
                                        st.success(f"✅ נמצאה והורדה מפת מושבים אוטומטית!")
                                        auto_map_found = True
                                        concert_map_path = save_path
                                        st.image(save_path, caption=f"מפת מושבים - {venue_name}", use_container_width=True)
                                        break
                            except:
                                continue
                    
                    if not auto_map_found:
                        st.markdown("**📥 הדבק מפת מושבים** (תישמר לשימוש עתידי)")
                        
                        venue_url = selected_concert.get('venue_url', '')
                        if venue_url:
                            st.markdown(f"🔗 [לחץ כאן לפתוח את עמוד האולם ב-Ticketmaster]({venue_url})")
                        elif concert_url:
                            st.markdown(f"🔗 [לחץ כאן לפתוח את דף האירוע ב-Ticketmaster]({concert_url})")
                        
                        st.info("📋 **צלם את מפת המושבים עם מספריים (Win+Shift+S) והדבק כאן:**")
                        
                        from streamlit_paste_button import paste_image_button as pbutton
                        
                        paste_result = pbutton(
                            label="📋 הדבק מפה מהמספריים (Ctrl+V)",
                            key=f"paste_map_{concert_venue_id}"
                        )
                        
                        if paste_result and paste_result.image_data:
                            try:
                                os.makedirs('attached_assets/concert_venue_maps', exist_ok=True)
                                save_path = f'attached_assets/concert_venue_maps/{concert_venue_id}.png'
                                paste_result.image_data.save(save_path, 'PNG')
                                st.success("✅ המפה נשמרה! היא תופיע אוטומטית בפעם הבאה שתבחר את ההופעה הזו.")
                                st.rerun()
                            except Exception as e:
                                st.error(f"❌ שגיאה בשמירה: {str(e)}")
            
            elif concert_venue_info:
                venue_name_he = concert_venue_info.get('name_he', '')
                capacity = concert_venue_info.get('capacity', 0)
                categories = concert_venue_info.get('categories', [])
                
                specific_venue_map = get_venue_map_path(concert_venue_id, use_fallback=False)
                if specific_venue_map and os.path.exists(specific_venue_map):
                    auto_stadium_map = specific_venue_map
                    st.success(f"✅ נמצאה מפת מושבים אוטומטית עבור {venue_name_he}")
                    st.image(specific_venue_map, caption=f"מפת מושבים - {venue_name_he}", use_container_width=True)
                else:
                    venue_map_fallback = get_venue_map_path(concert_venue_id, use_fallback=True)
                    if venue_map_fallback and os.path.exists(venue_map_fallback):
                        auto_stadium_map = venue_map_fallback
                        st.info(f"🎤 **{venue_name_he}** - קיבולת: {capacity:,} מושבים\n\n📍 קטגוריות זמינות: {', '.join(categories)}")
                        st.caption("תמונת אווירה כללית תופיע במסמך. ניתן להעלות מפת מושבים ספציפית:")
                    else:
                        st.info(f"🎤 **{venue_name_he}** - קיבולת: {capacity:,} מושבים\n\n📍 קטגוריות זמינות: {', '.join(categories)}")
                        st.markdown("*העלה תרשים מושבים של האולם להצגה במסמך ההזמנה:*")
        
        stadium_image = None
        
        # Check for saved concert map bytes first (from database)
        saved_map_bytes = st.session_state.get('saved_stadium_map_bytes')
        if saved_map_bytes:
            from io import BytesIO
            st.success("✅ תרשים מושבים מההופעה השמורה")
            st.image(saved_map_bytes, caption="תרשים מושבים (מההופעה השמורה)", use_container_width=True)
            stadium_image = Image.open(BytesIO(saved_map_bytes))
            if st.button("🗑️ הסר תרשים", key="remove_saved_map"):
                del st.session_state['saved_stadium_map_bytes']
                st.rerun()
        elif 'pasted_stadium_map' in st.session_state and st.session_state['pasted_stadium_map']:
            stadium_image = st.session_state['pasted_stadium_map']
            st.image(stadium_image, caption="תרשים מושבים (מהלוח)", use_container_width=True)
            if st.button("🗑️ הסר תרשים", key="remove_pasted_map"):
                del st.session_state['pasted_stadium_map']
                st.rerun()
        else:
            col_paste, col_upload = st.columns([1, 1])
            with col_paste:
                paste_map = paste_image_button(
                    label="📋 הדבק תרשים מהלוח",
                    key="paste_stadium_map",
                    background_color="#667eea",
                    hover_background_color="#5a6fd6"
                )
                if paste_map and paste_map.image_data:
                    st.session_state['pasted_stadium_map'] = paste_map.image_data
                    st.rerun()
            
            with col_upload:
                uploaded_map = st.file_uploader("📁 העלה קובץ", type=['png', 'jpg', 'jpeg'], key="upload_stadium_map")
                if uploaded_map:
                    st.session_state['pasted_stadium_map'] = Image.open(uploaded_map)
                    st.rerun()
            
            if not auto_stadium_map:
                st.caption("💡 ניתן גם להשתמש בכלי **🗺️ הורדת מפות** בסרגל הכלים")
        
        stadium_photo = None
        
        hotel_image = None
        hotel_image_2 = None
        
        if product_type == "package":
            st.markdown('<div class="form-section"><h3>🏨 פרטי המלון והטיסה</h3></div>', unsafe_allow_html=True)
            
            if 'hotel_data' not in st.session_state:
                st.session_state.hotel_data = {}
            
            hd = st.session_state.hotel_data
            
            if hd.get('hotel_image_path') and os.path.exists(hd.get('hotel_image_path', '')):
                hotel_image = hd.get('hotel_image_path')
            if hd.get('hotel_image_path_2') and os.path.exists(hd.get('hotel_image_path_2', '')):
                hotel_image_2 = hd.get('hotel_image_path_2')
            
            if hd.get('from_package') and hd.get('hotel_address'):
                st.success(f"✅ פרטי המלון נטענו מהחבילה: {hd.get('hotel_name', '')}")
            
            col_hotel, col_btn = st.columns([3, 1])
            with col_hotel:
                hotel_name = st.text_input("שם המלון", value=hd.get('hotel_name') or rd.get('hotel_name', ''), placeholder="לדוגמה: Hilton Madrid")
            with col_btn:
                st.markdown("<br>", unsafe_allow_html=True)
                lookup_hotel = st.button("🔍 חפש מלון", use_container_width=True)
            
            if lookup_hotel and hotel_name:
                city = venue.split(',')[-1].strip() if ',' in venue else ''
                query = f"{hotel_name}, {city}" if city else hotel_name
                
                with st.spinner("מחפש פרטי מלון..."):
                    result = resolve_hotel_safe(query)
                    
                    if result.get('error'):
                        st.error(f"❌ {result['error']}")
                    else:
                        if result.get('hotel_rating'):
                            rating = float(result['hotel_rating'])
                            if rating >= 4.5:
                                result['hotel_stars'] = "5 כוכבים"
                            elif rating >= 3.5:
                                result['hotel_stars'] = "4 כוכבים"
                            else:
                                result['hotel_stars'] = "3 כוכבים"
                        st.session_state.hotel_data = result
                        if result.get('from_cache'):
                            st.success(f"✅ נמצא (מהזיכרון): {result.get('hotel_name', hotel_name)}")
                        else:
                            st.success(f"✅ נמצא: {result.get('hotel_name', hotel_name)}")
                        if result.get('hotel_rating'):
                            st.info(f"⭐ דירוג: {result['hotel_rating']} ({result.get('hotel_stars', '')})")
                        st.rerun()
            
            if hd.get('hotel_address'):
                st.markdown(f"""
                <div style="background: #f8f9fa; padding: 1rem; border-radius: 10px; border-right: 4px solid #667eea; margin: 0.5rem 0;">
                    <p style="color: #333; margin: 5px 0;"><strong>📍 כתובת:</strong> {hd.get('hotel_address', '')}</p>
                    <p style="color: #333; margin: 5px 0;"><strong>🌐 אתר:</strong> {hd.get('hotel_website', 'לא זמין')}</p>
                    <p style="color: #333; margin: 5px 0;"><strong>⭐ דירוג:</strong> {hd.get('hotel_rating', 'לא זמין')}</p>
                </div>
                """, unsafe_allow_html=True)
            
            col_nights, col_stars = st.columns(2)
            with col_nights:
                hotel_nights = st.number_input("מספר לילות", min_value=1, value=rd.get('hotel_nights', 3))
            with col_stars:
                stars_options = ["3 כוכבים", "4 כוכבים", "5 כוכבים"]
                default_stars = hd.get('hotel_stars') or rd.get('hotel_stars', '5 כוכבים')
                stars_default = stars_options.index(default_stars) if default_stars in stars_options else 2
                hotel_stars = st.selectbox("דירוג המלון", stars_options, index=stars_default)
            
            meals_options = ["ללא ארוחות", "ארוחת בוקר", "חצי פנסיון", "פנסיון מלא"]
            meals_default = meals_options.index(rd.get('hotel_meals', 'ארוחת בוקר')) if rd.get('hotel_meals') in meals_options else 1
            hotel_meals = st.selectbox("ארוחות", meals_options, index=meals_default)
            
            st.markdown("**✈️ פרטי טיסות**")
            
            if 'flights_data' not in st.session_state:
                if rd.get('outbound_from'):
                    st.session_state.flights_data = {
                        'outbound': {'from': rd.get('outbound_from', 'TLV'), 'to': rd.get('outbound_to', ''), 'date': rd.get('outbound_date', ''), 'time': rd.get('outbound_time', ''), 'flight_no': rd.get('outbound_flight', '')},
                        'return': {'from': rd.get('return_from', ''), 'to': rd.get('return_to', 'TLV'), 'date': rd.get('return_date', ''), 'time': rd.get('return_time', ''), 'flight_no': rd.get('return_flight', '')}
                    }
                else:
                    st.session_state.flights_data = {
                        'outbound': {'from': 'TLV', 'to': '', 'date': '', 'time': '', 'flight_no': ''},
                        'return': {'from': '', 'to': 'TLV', 'date': '', 'time': '', 'flight_no': ''}
                    }
            
            fd = st.session_state.flights_data
            airport_options = [""] + get_airport_options()
            
            st.markdown("""
            <div style="background: linear-gradient(135deg, #667eea22, #764ba222); padding: 15px; border-radius: 10px; margin-bottom: 15px; border: 1px solid #667eea44;">
                <p style="margin: 0; font-size: 14px;">📷 <strong>סריקת טיסות אוטומטית:</strong> העלה צילום מסך של פרטי הטיסות והמערכת תמלא את השדות!</p>
            </div>
            """, unsafe_allow_html=True)
            
            col_flight_upload, col_flight_paste = st.columns([3, 1])
            with col_flight_upload:
                flight_screenshot = st.file_uploader(
                    "📷 העלה צילום מסך של טיסות",
                    type=['png', 'jpg', 'jpeg'],
                    key="flight_scan_upload",
                    help="צלם מסך מאתר הזמנת הטיסות והעלה כאן"
                )
            with col_flight_paste:
                flight_paste = paste_image_button("📋 הדבק טיסות", key="flight_paste")
                if flight_paste.image_data:
                    st.session_state['pasted_flight'] = flight_paste.image_data
                    st.image(flight_paste.image_data, caption="צילום מסך שהודבק", width=100)
            
            scan_flights_btn = st.button("🔍 סרוק פרטי טיסות", type="secondary", use_container_width=True)
            
            flight_image_to_scan = flight_screenshot or st.session_state.get('pasted_flight')
            if scan_flights_btn and flight_image_to_scan:
                with st.spinner("סורק פרטי טיסות..."):
                    if flight_screenshot:
                        image_bytes = flight_screenshot.read()
                    else:
                        pasted_img = st.session_state['pasted_flight']
                        img_byte_arr = io.BytesIO()
                        pasted_img.save(img_byte_arr, format='PNG')
                        image_bytes = img_byte_arr.getvalue()
                    result = extract_flight_data(image_bytes)
                    
                    if result.get('success') and result.get('flights'):
                        flights = result['flights']
                        
                        for f in flights:
                            direction = f.get('direction', 'outbound')
                            if direction in ['outbound', 'return']:
                                for key in [f"flight_{direction}_from", f"flight_{direction}_to", 
                                           f"flight_{direction}_date", f"flight_{direction}_time", 
                                           f"flight_{direction}_no"]:
                                    if key in st.session_state:
                                        del st.session_state[key]
                                
                                from_code = f.get('from', '').upper().strip()
                                to_code = f.get('to', '').upper().strip()
                                
                                st.session_state.flights_data[direction] = {
                                    'from': from_code,
                                    'to': to_code,
                                    'date': f.get('date', ''),
                                    'time': f.get('time', ''),
                                    'flight_no': f.get('flight_no', '')
                                }
                                
                                from_display = format_airport_display(from_code) if from_code else ""
                                to_display = format_airport_display(to_code) if to_code else ""
                                
                                st.session_state[f"flight_{direction}_from"] = from_display
                                st.session_state[f"flight_{direction}_to"] = to_display
                                st.session_state[f"flight_{direction}_date"] = f.get('date', '')
                                st.session_state[f"flight_{direction}_time"] = f.get('time', '')
                                st.session_state[f"flight_{direction}_no"] = f.get('flight_no', '')
                        
                        st.success(f"✅ נסרקו {len(flights)} טיסות!")
                        st.rerun()
                    else:
                        st.error(f"❌ לא הצלחנו לזהות פרטי טיסות: {result.get('error', 'נסה תמונה ברורה יותר')}")
            elif scan_flights_btn and not flight_screenshot:
                st.warning("⚠️ יש להעלות צילום מסך לפני הסריקה")
            
            st.markdown("**טיסת הלוך:**")
            col_out1, col_out2 = st.columns(2)
            with col_out1:
                if "flight_outbound_from" not in st.session_state:
                    out_from_default = format_airport_display(fd['outbound'].get('from', 'TLV')) if fd['outbound'].get('from') else airport_options[0]
                    st.session_state["flight_outbound_from"] = out_from_default
                outbound_from = st.selectbox("מ:", airport_options, key="flight_outbound_from")
            with col_out2:
                if "flight_outbound_to" not in st.session_state:
                    out_to_default = format_airport_display(fd['outbound'].get('to', '')) if fd['outbound'].get('to') else airport_options[0]
                    st.session_state["flight_outbound_to"] = out_to_default
                outbound_to = st.selectbox("אל:", airport_options, key="flight_outbound_to")
            
            col_out3, col_out4, col_out5 = st.columns(3)
            with col_out3:
                if "flight_outbound_date" not in st.session_state:
                    st.session_state["flight_outbound_date"] = fd['outbound'].get('date', '')
                outbound_date = st.text_input("תאריך", placeholder="15/01", key="flight_outbound_date")
            with col_out4:
                if "flight_outbound_time" not in st.session_state:
                    st.session_state["flight_outbound_time"] = fd['outbound'].get('time', '')
                outbound_time = st.text_input("שעה", placeholder="09:00", key="flight_outbound_time")
            with col_out5:
                if "flight_outbound_no" not in st.session_state:
                    st.session_state["flight_outbound_no"] = fd['outbound'].get('flight_no', '')
                outbound_no = st.text_input("מס' טיסה", placeholder="LY315", key="flight_outbound_no")
            
            st.markdown("**טיסת חזור:**")
            col_ret1, col_ret2 = st.columns(2)
            with col_ret1:
                if "flight_return_from" not in st.session_state:
                    ret_from_default = format_airport_display(fd['return'].get('from', '')) if fd['return'].get('from') else airport_options[0]
                    st.session_state["flight_return_from"] = ret_from_default
                return_from = st.selectbox("מ:", airport_options, key="flight_return_from")
            with col_ret2:
                if "flight_return_to" not in st.session_state:
                    ret_to_default = format_airport_display(fd['return'].get('to', 'TLV')) if fd['return'].get('to') else airport_options[0]
                    st.session_state["flight_return_to"] = ret_to_default
                return_to = st.selectbox("אל:", airport_options, key="flight_return_to")
            
            col_ret3, col_ret4, col_ret5 = st.columns(3)
            with col_ret3:
                if "flight_return_date" not in st.session_state:
                    st.session_state["flight_return_date"] = fd['return'].get('date', '')
                return_date = st.text_input("תאריך", placeholder="18/01", key="flight_return_date")
            with col_ret4:
                if "flight_return_time" not in st.session_state:
                    st.session_state["flight_return_time"] = fd['return'].get('time', '')
                return_time = st.text_input("שעה", placeholder="22:00", key="flight_return_time")
            with col_ret5:
                if "flight_return_no" not in st.session_state:
                    st.session_state["flight_return_no"] = fd['return'].get('flight_no', '')
                return_no = st.text_input("מס' טיסה", placeholder="LY316", key="flight_return_no")
            
            out_from_code = get_airport_code(outbound_from)
            out_to_code = get_airport_code(outbound_to)
            ret_from_code = get_airport_code(return_from)
            ret_to_code = get_airport_code(return_to)
            
            flights_list = []
            if out_from_code and out_to_code:
                flights_list.append({
                    'direction': 'הלוך',
                    'from': out_from_code,
                    'to': out_to_code,
                    'date': outbound_date,
                    'time': outbound_time,
                    'flight_no': outbound_no
                })
            if ret_from_code and ret_to_code:
                flights_list.append({
                    'direction': 'חזור',
                    'from': ret_from_code,
                    'to': ret_to_code,
                    'date': return_date,
                    'time': return_time,
                    'flight_no': return_no
                })
            
            flight_details = ""
            if flights_list:
                lines = []
                for f in flights_list:
                    line = f"{f['direction']}: {f['from']}-{f['to']}"
                    if f['date']:
                        line += f" {f['date']}"
                    if f['time']:
                        line += f" {f['time']}"
                    if f['flight_no']:
                        line += f" ({f['flight_no']})"
                    lines.append(line)
                flight_details = "\n".join(lines)
            
            st.markdown("**🧳 כבודה:**")
            col_bag1, col_bag2 = st.columns(2)
            with col_bag1:
                bag_trolley = st.checkbox("כולל טרולי עד 7 ק\"ג", value=rd.get('bag_trolley', True), key="bag_trolley")
            with col_bag2:
                bag_options = ["ללא כבודה רשומה", "כולל כבודה עד 20 ק\"ג", "כולל כבודה עד 23 ק\"ג", "כולל כבודה עד 25 ק\"ג"]
                bag_default_idx = 0
                if rd.get('bag_checked'):
                    for i, opt in enumerate(bag_options):
                        if rd.get('bag_checked') in opt:
                            bag_default_idx = i
                            break
                bag_checked = st.selectbox("כבודה רשומה:", bag_options, index=bag_default_idx, key="bag_checked")
            
            transfers = st.checkbox("כולל העברות משדה התעופה", value=rd.get('transfers', True))
            
        else:
            hotel_name = ""
            hotel_nights = 0
            hotel_stars = ""
            hotel_meals = ""
            flight_details = ""
            flights_list = []
            transfers = False
            bag_trolley = False
            bag_checked = ""
        
        st.markdown('<div class="form-section"><h3>👤 פרטי הלקוח</h3></div>', unsafe_allow_html=True)
        
        customer_name = st.text_input("שם מלא", value=rd.get('customer_name', ''), placeholder="ישראל ישראלי")
        
        col_id, col_phone = st.columns(2)
        with col_id:
            customer_id = st.text_input("תעודת זהות", value=rd.get('customer_id', ''), placeholder="123456789")
        with col_phone:
            customer_phone = st.text_input("טלפון", value=rd.get('customer_phone', ''), placeholder="050-1234567")
        
        customer_email = st.text_input("אימייל", value=rd.get('customer_email', ''), placeholder="example@email.com")
        
        st.markdown('<div class="form-section"><h3>🎫 פרטי הכרטיסים</h3></div>', unsafe_allow_html=True)
        
        ticket_description = st.text_area(
            "תיאור הכרטיסים",
            value=rd.get('ticket_description', ''),
            placeholder="לדוגמה: שני כרטיסים בישיבה בטבעת הרביעית מאחורי השער",
            height=80
        )
        
        default_category = rd.get('category', '')
        if not default_category:
            if event_type == "כדורגל" and st.session_state.get('worldcup_category'):
                default_category = st.session_state.get('worldcup_category', '')
            elif event_type == "הופעה" and st.session_state.get('concert_selected_category'):
                default_category = st.session_state.get('concert_selected_category', '')
        
        category = st.text_input("קטגוריה", value=default_category, placeholder="CAT 1 / VIP / Premium")
        
        from exchange_rates import fetch_exchange_rates, get_currency_symbol, get_currency_name_hebrew
        
        currency_options = ['EUR', 'USD', 'GBP']
        currency_labels = ['€ יורו', '$ דולר', '£ פאונד']
        
        col_currency, col_price, col_qty = st.columns([1, 1, 1])
        with col_currency:
            selected_currency = st.selectbox(
                "מטבע", 
                currency_options,
                format_func=lambda x: currency_labels[currency_options.index(x)]
            )
        with col_price:
            currency_symbol = get_currency_symbol(selected_currency)
            currency_name = get_currency_name_hebrew(selected_currency)
            price_foreign = st.number_input(f"מחיר לאדם ({currency_symbol})", min_value=0, value=rd.get('price_euro', 330))
        with col_qty:
            num_tickets = st.number_input("מספר כרטיסים", min_value=1, value=rd.get('num_tickets', 2))
        
        rates = fetch_exchange_rates()
        exchange_rate = rates.get(selected_currency, 3.78)
        
        st.markdown(f"""
        <div style="background: #f0f2f6; padding: 10px; border-radius: 8px; margin: 10px 0;">
            📊 שער המרה ({currency_name} לשקל): <strong>{exchange_rate}</strong> ₪ 
            <span style="color: #666; font-size: 12px;">(כולל מרווח 5 אג')</span>
        </div>
        """, unsafe_allow_html=True)
        
        price_nis = int(price_foreign * exchange_rate)
        total_foreign = price_foreign * num_tickets
        total_nis = int(total_foreign * exchange_rate)
        
        st.markdown(f"""
        <div class="price-display">
            💰 סה"כ: {total_foreign} {currency_symbol} = {total_nis:,} ש"ח
        </div>
        """, unsafe_allow_html=True)
        
        price_euro = price_foreign if selected_currency == 'EUR' else 0
        total_euro = total_foreign if selected_currency == 'EUR' else 0
        
        st.markdown('<div class="form-section"><h3>✈️ פרטי הנוסעים</h3></div>', unsafe_allow_html=True)
        
        if 'passenger_list' not in st.session_state:
            st.session_state.passenger_list = [{
                'first_name': '', 
                'last_name': '', 
                'passport': '', 
                'birth_date': '',
                'passport_expiry': '',
                'ticket_type': 'כרטיס רגיל'
            }]
        
        st.markdown("""
        <div style="background: linear-gradient(135deg, #667eea22, #764ba222); padding: 15px; border-radius: 10px; margin-bottom: 20px; border: 1px solid #667eea44;">
            <p style="margin: 0; font-size: 14px;">📷 <strong>סריקת דרכונים אוטומטית:</strong> העלה תמונות דרכונים - כל דרכון יוסיף נוסע חדש עם הפרטים שנסרקו!</p>
        </div>
        """, unsafe_allow_html=True)
        
        col_passport_upload, col_passport_paste = st.columns([3, 1])
        with col_passport_upload:
            passport_uploads = st.file_uploader(
                "📷 העלה תמונות דרכונים (ניתן להעלות מספר קבצים)",
                type=['png', 'jpg', 'jpeg'],
                key="passport_scan_upload",
                accept_multiple_files=True,
                help="העלה צילומים ברורים של דרכונים - כל דרכון יוסיף נוסע חדש"
            )
        with col_passport_paste:
            # Initialize
            if 'pasted_passports' not in st.session_state:
                st.session_state['pasted_passports'] = []
            if '_passport_paste_refresh' not in st.session_state:
                st.session_state['_passport_paste_refresh'] = 0
            
            # Dynamic key with refresh counter
            paste_key = f"passport_paste_{st.session_state['_passport_paste_refresh']}"
            passport_paste = paste_image_button("📋 הדבק דרכון", key=paste_key)
            
            if passport_paste.image_data:
                # Add to list and force component refresh
                st.session_state['pasted_passports'].append(passport_paste.image_data)
                st.session_state['_passport_paste_refresh'] += 1
                st.success(f"✅ דרכון {len(st.session_state['pasted_passports'])} נוסף!")
                st.rerun()
        
        # Display all pasted passports
        if st.session_state.get('pasted_passports'):
            st.info(f"📋 {len(st.session_state['pasted_passports'])} דרכונים מהלוח מוכנים לסריקה")
            cols = st.columns(min(len(st.session_state['pasted_passports']), 5))
            for i, pasted_img in enumerate(st.session_state['pasted_passports']):
                with cols[i % 5]:
                    st.image(pasted_img, caption=f"#{i+1}", width=80)
            if st.button("🗑️ מחק הכל", key="clear_all_pasted"):
                st.session_state['pasted_passports'] = []
                st.session_state['_passport_paste_refresh'] += 1
                st.rerun()
        
        has_passport_input = passport_uploads or st.session_state.get('pasted_passports')
        if passport_uploads:
            st.info(f"📁 {len(passport_uploads)} דרכונים הועלו")
        
        scan_button = st.button("🔍 סרוק דרכונים והוסף נוסעים", type="primary", use_container_width=True)
        
        if scan_button and has_passport_input:
            progress_bar = st.progress(0)
            status_text = st.empty()
            scanned_passengers = []
            
            images_to_scan = []
            if passport_uploads:
                for pf in passport_uploads:
                    images_to_scan.append(('file', pf))
            if st.session_state.get('pasted_passports'):
                for pasted_img in st.session_state['pasted_passports']:
                    images_to_scan.append(('pasted', pasted_img))
            
            for idx, (source_type, passport_data) in enumerate(images_to_scan):
                source_name = f"דרכון מהלוח" if source_type == 'pasted' else passport_data.name
                status_text.text(f"🔄 סורק {source_name} ({idx + 1} מתוך {len(images_to_scan)})...")
                progress_bar.progress((idx + 1) / len(images_to_scan))
                
                if source_type == 'file':
                    image_bytes = passport_data.read()
                else:
                    img_byte_arr = io.BytesIO()
                    passport_data.save(img_byte_arr, format='PNG')
                    image_bytes = img_byte_arr.getvalue()
                
                result = extract_passport_data(image_bytes)
                
                if result.get('success'):
                    scanned_passengers.append({
                        'first_name': result.get('first_name', ''),
                        'last_name': result.get('last_name', ''),
                        'passport': result.get('passport_number', ''),
                        'birth_date': result.get('birth_date', ''),
                        'passport_expiry': result.get('passport_expiry', ''),
                        'ticket_type': 'כרטיס רגיל'
                    })
                else:
                    st.error(f"❌ שגיאה בסריקת {source_name}: {result.get('error', 'לא ניתן לקרוא')}")
            
            if scanned_passengers:
                is_first_empty = len(st.session_state.passenger_list) == 1 and not st.session_state.passenger_list[0].get('first_name')
                
                for i, passenger in enumerate(scanned_passengers):
                    if is_first_empty and i == 0:
                        passenger_idx = 0
                        st.session_state.passenger_list[0] = passenger
                    else:
                        passenger_idx = len(st.session_state.passenger_list)
                        st.session_state.passenger_list.append(passenger)
                    
                    for key in [f"first_name_{passenger_idx}", f"last_name_{passenger_idx}", 
                               f"passport_{passenger_idx}", f"birth_date_{passenger_idx}", 
                               f"passport_expiry_{passenger_idx}"]:
                        if key in st.session_state:
                            del st.session_state[key]
                    
                    st.session_state[f"first_name_{passenger_idx}"] = passenger['first_name']
                    st.session_state[f"last_name_{passenger_idx}"] = passenger['last_name']
                    st.session_state[f"passport_{passenger_idx}"] = passenger['passport']
                    st.session_state[f"birth_date_{passenger_idx}"] = passenger['birth_date']
                    st.session_state[f"passport_expiry_{passenger_idx}"] = passenger['passport_expiry']
                
                status_text.text(f"✅ סריקה הושלמה! {len(scanned_passengers)} נוסעים נוספו.")
                st.rerun()
            else:
                status_text.text("❌ לא הצלחנו לסרוק אף דרכון")
        elif scan_button and not has_passport_input:
            st.warning("⚠️ יש להעלות או להדביק תמונות דרכונים לפני הסריקה")
        
        st.markdown("---")
        
        ticket_type_options = ['כרטיס רגיל', 'כרטיס VIP', 'כרטיס ילד', 'כרטיס מלווה']
        
        for i, passenger in enumerate(st.session_state.passenger_list):
            st.markdown(f"**נוסע {i+1}**")
            
            fn_key = f"first_name_{i}"
            ln_key = f"last_name_{i}"
            type_key = f"type_{i}"
            passport_key = f"passport_{i}"
            birth_key = f"birth_date_{i}"
            exp_key = f"passport_expiry_{i}"
            
            # Always sync from passenger_list to session_state on each render
            # This ensures OCR-scanned data is properly displayed
            stored_first = passenger.get('first_name', '')
            stored_last = passenger.get('last_name', '')
            stored_passport = passenger.get('passport', '')
            stored_birth = passenger.get('birth_date', '')
            stored_exp = passenger.get('passport_expiry', '')
            
            # Initialize if not in session state, or if passenger_list has newer data
            if fn_key not in st.session_state or (stored_first and not st.session_state.get(fn_key)):
                st.session_state[fn_key] = stored_first
            if ln_key not in st.session_state or (stored_last and not st.session_state.get(ln_key)):
                st.session_state[ln_key] = stored_last
            if passport_key not in st.session_state or (stored_passport and not st.session_state.get(passport_key)):
                st.session_state[passport_key] = stored_passport
            if birth_key not in st.session_state or (stored_birth and not st.session_state.get(birth_key)):
                st.session_state[birth_key] = stored_birth
            if exp_key not in st.session_state or (stored_exp and not st.session_state.get(exp_key)):
                st.session_state[exp_key] = stored_exp
            
            col_fn, col_ln, col_type, col_del = st.columns([1.3, 1.3, 1.2, 0.3])
            with col_fn:
                first_name = st.text_input("שם פרטי", key=fn_key, placeholder="John")
                st.session_state.passenger_list[i]['first_name'] = first_name
            with col_ln:
                last_name = st.text_input("שם משפחה", key=ln_key, placeholder="Doe")
                st.session_state.passenger_list[i]['last_name'] = last_name
            with col_type:
                current_type = passenger.get('ticket_type', 'כרטיס רגיל')
                type_index = ticket_type_options.index(current_type) if current_type in ticket_type_options else 0
                ticket_type = st.selectbox("סוג כרטיס", options=ticket_type_options, index=type_index, key=type_key)
                st.session_state.passenger_list[i]['ticket_type'] = ticket_type
            with col_del:
                if len(st.session_state.passenger_list) > 1:
                    if st.button("🗑️", key=f"del_{i}"):
                        st.session_state.passenger_list.pop(i)
                        st.rerun()
            
            col_passport, col_birth, col_exp = st.columns(3)
            with col_passport:
                passport = st.text_input("מספר דרכון", key=passport_key, placeholder="12345678")
                st.session_state.passenger_list[i]['passport'] = passport
            with col_birth:
                birth_date = st.text_input("תאריך לידה (DD/MM/YYYY)", key=birth_key, placeholder="15/03/1990")
                st.session_state.passenger_list[i]['birth_date'] = birth_date
            with col_exp:
                passport_expiry = st.text_input("תוקף דרכון (DD/MM/YYYY)", key=exp_key, placeholder="15/03/2030")
                st.session_state.passenger_list[i]['passport_expiry'] = passport_expiry
        
        if st.button("➕ הוסף נוסע"):
            st.session_state.passenger_list.append({
                'first_name': '', 
                'last_name': '', 
                'passport': '', 
                'birth_date': '',
                'passport_expiry': '',
                'ticket_type': 'כרטיס רגיל'
            })
            st.rerun()
        
        passengers = [p for p in st.session_state.passenger_list if (p.get('first_name', '').strip() or p.get('last_name', '').strip() or p.get('name', '').strip())]
    
    with col2:
        st.markdown("### 👁️ תצוגה מקדימה")
        
        if stadium_image:
            st.image(stadium_image, caption="תרשים מושבים", use_container_width=True)
        elif auto_stadium_map and os.path.exists(auto_stadium_map):
            st.image(auto_stadium_map, caption="תרשים מושבים (אוטומטי)", use_container_width=True)
        if hotel_image or hotel_image_2:
            if hotel_image and hotel_image_2:
                col_h1, col_h2 = st.columns(2)
                with col_h1:
                    st.image(hotel_image, caption="תמונת המלון 1", use_container_width=True)
                with col_h2:
                    st.image(hotel_image_2, caption="תמונת המלון 2", use_container_width=True)
            elif hotel_image:
                st.image(hotel_image, caption="תמונת המלון", use_container_width=True)
        
        st.markdown("#### 📋 סיכום ההזמנה")
        
        if event_name:
            st.info(f"**אירוע:** {event_name}")
        if event_date:
            st.write(f"**תאריך:** {event_date.strftime('%d/%m/%Y')} {event_time.strftime('%H:%M')}")
        if venue:
            st.write(f"**מקום:** {venue}")
        if customer_name:
            st.write(f"**לקוח:** {customer_name}")
        if category:
            st.write(f"**קטגוריה:** {category}")
        
        if passengers:
            st.write("**נוסעים:**")
            for p in passengers:
                if isinstance(p, dict):
                    name = f"{p.get('first_name', '')} {p.get('last_name', '')}".strip()
                    if not name:
                        name = p.get('name', '')
                    st.write(f"- {name} ({p.get('ticket_type', 'כרטיס רגיל')})")
                else:
                    st.write(f"- {p}")
        
        if total_foreign > 0:
            st.success(f"**סה\"כ:** {total_foreign} {currency_symbol} = {total_nis:,} ש\"ח")
        
        st.markdown("---")
        
        can_generate = all([event_name, customer_name, customer_email, category])
        
        if can_generate:
            hd = st.session_state.get('hotel_data', {})
            
            team_data = st.session_state.get('selected_team_data', {})
            away_team_data = st.session_state.get('away_team_data', {})
            
            order_data = {
                'product_type': product_type,
                'event_name': event_name,
                'event_type': event_type,
                'event_date': f"{event_date.strftime('%d/%m/%Y')} {event_time.strftime('%H:%M')}",
                'event_date_str': event_date.strftime('%d/%m/%Y'),
                'event_time_str': event_time.strftime('%H:%M'),
                'venue': venue or '',
                'customer_name': customer_name,
                'customer_id': customer_id or '',
                'customer_phone': customer_phone or '',
                'customer_email': customer_email,
                'ticket_description': ticket_description or '',
                'category': category,
                'currency': selected_currency,
                'currency_symbol': currency_symbol,
                'price_per_ticket': price_foreign,
                'price_nis': price_nis,
                'total_foreign': total_foreign,
                'total_euro': total_foreign,
                'total_nis': total_nis,
                'num_tickets': num_tickets,
                'passengers': passengers,
                'exchange_rate': exchange_rate,
                'home_team_badge': team_data.get('badge', ''),
                'away_team_badge': away_team_data.get('badge', ''),
                'home_team_name': st.session_state.get('home_team_hebrew', ''),
                'away_team_name': st.session_state.get('away_team_hebrew', ''),
                'hotel_name': hd.get('hotel_name') or hotel_name,
                'hotel_nights': hotel_nights,
                'hotel_stars': hotel_stars,
                'hotel_meals': hotel_meals,
                'hotel_address': hd.get('hotel_address', ''),
                'hotel_website': hd.get('hotel_website', ''),
                'hotel_rating': hd.get('hotel_rating', ''),
                'hotel_image_path': hd.get('hotel_image_path', ''),
                'hotel_image_path_2': hd.get('hotel_image_path_2', ''),
                'flight_details': flight_details,
                'flights': flights_list if product_type == 'package' else [],
                'transfers': transfers,
                'bag_trolley': bag_trolley if product_type == 'package' else False,
                'bag_checked': bag_checked if product_type == 'package' else '',
                'is_date_final': is_date_final,
                'seats_together': seats_together
            }
            
            stadium_img = None
            stadium_photo_img = None
            hotel_img = None
            hotel_img_2 = None
            
            def safe_open_image(path_or_image):
                """Safely open image, skipping SVG files that PIL can't handle"""
                try:
                    if path_or_image is None:
                        return None
                    if isinstance(path_or_image, Image.Image):
                        return path_or_image
                    if isinstance(path_or_image, str):
                        # Skip SVG files - PIL can't handle them
                        if path_or_image.lower().endswith('.svg'):
                            return None
                        if os.path.exists(path_or_image):
                            return Image.open(path_or_image)
                    return None
                except Exception as e:
                    print(f"Error opening image: {e}")
                    return None
            
            if stadium_image:
                stadium_img = safe_open_image(stadium_image)
            elif auto_stadium_map:
                stadium_img = safe_open_image(auto_stadium_map)
            
            if not stadium_img and rd.get('use_sample_images') and os.path.exists('attached_assets/stock_images/football_stadium_int_9fde699a.jpg'):
                stadium_img = Image.open('attached_assets/stock_images/football_stadium_int_9fde699a.jpg')
            
            random_atmosphere = get_random_atmosphere_image(event_type)
            if random_atmosphere:
                stadium_photo_img = safe_open_image(random_atmosphere)
            
            if hotel_image:
                hotel_img = safe_open_image(hotel_image)
            elif rd.get('use_sample_images') and os.path.exists('attached_assets/stock_images/luxury_hotel_exterio_3264e2db.jpg'):
                hotel_img = Image.open('attached_assets/stock_images/luxury_hotel_exterio_3264e2db.jpg')
            
            if hotel_image_2:
                hotel_img_2 = safe_open_image(hotel_image_2)
            
            template_version = 2
            
            st.markdown("### 📤 פעולות")
            
            col_btn1, col_btn2 = st.columns(2)
            with col_btn1:
                if st.button("📦 שמור כחבילה קבועה", type="secondary", use_container_width=True):
                    st.session_state['show_save_package_form'] = True
            with col_btn2:
                generate_pdf_btn = st.button("📄 צור PDF ושמור הזמנה", type="primary", use_container_width=True)
            
            if st.session_state.get('package_saved_success'):
                st.success(f"✅ החבילה '{st.session_state['package_saved_success']}' נשמרה בהצלחה!")
                st.info("💡 תוכל למצוא אותה ב'חבילות קבועות' בתפריט או לטעון אותה מהרשימה למעלה.")
                del st.session_state['package_saved_success']
            
            if st.session_state.get('show_save_package_form'):
                st.markdown("---")
                st.markdown("#### 📦 שמירה כחבילה קבועה")
                
                default_pkg_name = f"{event_name} - {category}" if event_name else ""
                pkg_name_input = st.text_input("📝 שם החבילה", value=default_pkg_name, placeholder="למשל: סטינג לימסול 2026 - גולדן", key="save_pkg_name")
                
                col_save, col_cancel = st.columns(2)
                with col_save:
                    confirm_save = st.button("💾 אשר שמירת חבילה", use_container_width=True, type="primary")
                with col_cancel:
                    if st.button("❌ ביטול", use_container_width=True):
                        st.session_state['show_save_package_form'] = False
                        st.rerun()
                
                if confirm_save:
                    if not pkg_name_input:
                        st.error("❌ יש להזין שם לחבילה")
                    else:
                        db = get_db()
                        if db:
                            try:
                                event_type_map = {'הופעה': EventType.CONCERT, 'כדורגל': EventType.FOOTBALL, 'אחר': EventType.OTHER}
                                event_type_enum = event_type_map.get(event_type, EventType.OTHER)
                                product_type_val = "full_package" if product_type == "package" else "tickets_only"
                                
                                hotel_data_json = {}
                                if product_type == "package":
                                    hd_save = st.session_state.get('hotel_data', {})
                                    hotel_data_json = {
                                        'name': hd_save.get('hotel_name') or hotel_name,
                                        'check_in': hd_save.get('check_in', ''),
                                        'check_out': hd_save.get('check_out', ''),
                                        'address': hd_save.get('hotel_address', ''),
                                        'website': hd_save.get('hotel_website', ''),
                                        'rating': hd_save.get('hotel_rating', ''),
                                        'stars': hotel_stars,
                                        'nights': hotel_nights,
                                        'meals': hotel_meals
                                    }
                                
                                flight_data_json = {}
                                if product_type == "package" and flights_list:
                                    outbound_flight = next((f for f in flights_list if f.get('direction') == 'הלוך'), {})
                                    return_flight = next((f for f in flights_list if f.get('direction') == 'חזור'), {})
                                    flight_data_json = {
                                        'outbound': {
                                            'from': outbound_flight.get('from', ''),
                                            'to': outbound_flight.get('to', ''),
                                            'date': outbound_flight.get('date', ''),
                                            'time': outbound_flight.get('time', ''),
                                            'flight_number': outbound_flight.get('flight_no', '')
                                        },
                                        'return': {
                                            'from': return_flight.get('from', ''),
                                            'to': return_flight.get('to', ''),
                                            'date': return_flight.get('date', ''),
                                            'time': return_flight.get('time', ''),
                                            'flight_number': return_flight.get('flight_no', '')
                                        }
                                    }
                                
                                stadium_map_bytes = None
                                if st.session_state.get('saved_stadium_map_bytes'):
                                    stadium_map_bytes = st.session_state.get('saved_stadium_map_bytes')
                                elif st.session_state.get('pasted_stadium_map'):
                                    pasted_img = st.session_state.get('pasted_stadium_map')
                                    img_byte_arr = io.BytesIO()
                                    pasted_img.save(img_byte_arr, format='PNG')
                                    stadium_map_bytes = img_byte_arr.getvalue()
                                elif auto_stadium_map and os.path.exists(auto_stadium_map):
                                    with open(auto_stadium_map, 'rb') as f:
                                        stadium_map_bytes = f.read()
                                
                                if hotel_data_json and product_type == "package":
                                    hotel_data_json['image_path'] = hd_save.get('hotel_image_path', '')
                                    hotel_data_json['image_path_2'] = hd_save.get('hotel_image_path_2', '')
                                
                                new_pkg = PackageTemplate(
                                    name=pkg_name_input,
                                    event_type=event_type_enum,
                                    product_type=product_type_val,
                                    event_name=event_name,
                                    event_date=event_date.strftime('%d/%m/%Y'),
                                    event_time=event_time.strftime('%H:%M'),
                                    venue=venue,
                                    ticket_description=ticket_description,
                                    ticket_category=category,
                                    price_per_ticket_euro=float(price_foreign),
                                    hotel_data=json.dumps(hotel_data_json) if hotel_data_json else None,
                                    flight_data=json.dumps(flight_data_json) if flight_data_json else None,
                                    package_price_euro=float(price_foreign),
                                    stadium_map_data=stadium_map_bytes,
                                    stadium_map_mime='image/png' if stadium_map_bytes else None,
                                    notes=""
                                )
                                
                                db.add(new_pkg)
                                db.commit()
                                st.session_state['package_saved_success'] = pkg_name_input
                                st.session_state['show_save_package_form'] = False
                                st.rerun()
                            except Exception as e:
                                db.rollback()
                                st.error(f"❌ שגיאה בשמירה: {str(e)}")
                            finally:
                                db.close()
            
            if generate_pdf_btn:
                order_number = generate_order_number()
                order_data['order_number'] = order_number
                
                with st.spinner("יוצר PDF..."):
                    pdf_bytes = generate_pdf(order_data, stadium_img, hotel_img, hotel_img_2, stadium_photo_img, template_version)
                    st.session_state.pdf_bytes = pdf_bytes
                    st.session_state.order_generated = True
                    st.session_state.current_order_number = order_number
                
                st.success("✅ ה-PDF נוצר בהצלחה!")
                st.info(f"📋 מספר הזמנה: {order_number}")
                
                # Try to save to database (may be slow, but PDF is already ready)
                try:
                    saved_order = save_order_to_db(order_data, pdf_bytes)
                    if saved_order:
                        st.session_state.current_order_id = saved_order.id
                except Exception as e:
                    st.warning("⚠️ ההזמנה לא נשמרה במסד הנתונים, אך ה-PDF זמין להורדה.")
            
            if st.session_state.get('order_generated') and st.session_state.get('pdf_bytes'):
                filename = f"הזמנה_{customer_name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.pdf"
                
                st.download_button(
                    label="⬇️ הורד PDF",
                    data=st.session_state.pdf_bytes,
                    file_name=filename,
                    mime="application/pdf",
                    use_container_width=True
                )
                
                st.markdown("---")
                
                st.markdown("### 📧 שלח ללקוח באימייל")
                
                email_subject = st.text_input(
                    "נושא המייל",
                    value=f"הצעת מחיר - {event_name}"
                )
                
                email_body = st.text_area(
                    "תוכן ההודעה",
                    value=f"""שלום {customer_name},

מצורפת הצעת המחיר שלך לאירוע:
🎟️ {event_name}
📅 {event_date.strftime('%d/%m/%Y')} בשעה {event_time.strftime('%H:%M')}
📍 {venue}

סה"כ: {total_euro} יורו ({total_nis:,} ש"ח)

לאישור ההזמנה, אנא השב למייל זה או צור קשר בטלפון.

בברכה,
צוות TikTik
972-732726000
""",
                    height=200
                )
                
                if st.button("📧 שלח מייל ללקוח", use_container_width=True):
                    resend_api_key = os.environ.get('RESEND_API_KEY')
                    if resend_api_key:
                        try:
                            import resend
                            resend.api_key = resend_api_key
                            
                            pdf_base64 = base64.b64encode(st.session_state.pdf_bytes).decode()
                            
                            resend.Emails.send({
                                "from": "TikTik <orders@tiktik.co.il>",
                                "to": [customer_email],
                                "subject": email_subject,
                                "text": email_body,
                                "attachments": [{
                                    "filename": filename,
                                    "content": pdf_base64
                                }]
                            })
                            
                            if st.session_state.get('current_order_id'):
                                update_order_status(st.session_state.current_order_id, OrderStatus.SENT)
                            
                            st.success(f"✅ המייל נשלח בהצלחה ל-{customer_email}!")
                        except Exception as e:
                            st.error(f"❌ שגיאה בשליחת המייל: {str(e)}")
                    else:
                        st.warning("⚠️ לא הוגדר מפתח API לשליחת מיילים. הורד את ה-PDF ושלח ידנית.")
                        st.info("💡 טיפ: הוסף את מפתח ה-RESEND_API_KEY בהגדרות הסביבה כדי להפעיל שליחת מיילים אוטומטית.")
        else:
            st.markdown('<div class="info-box">', unsafe_allow_html=True)
            st.info("📝 מלא את הפרטים הנדרשים: שם אירוע, שם לקוח, אימייל ובלוק")
            st.markdown('</div>', unsafe_allow_html=True)

def page_order_history():
    """Order history page"""
    render_header()
    
    st.markdown("### 📋 היסטוריית הזמנות")
    
    col1, col2 = st.columns([2, 1])
    with col1:
        search_query = st.text_input("🔍 חיפוש", placeholder="חפש לפי שם לקוח, אירוע, מספר הזמנה...")
    with col2:
        status_filter = st.selectbox("סינון לפי סטטוס", ["הכל", "טיוטה", "נשלח", "נצפה", "נחתם", "בוטל"])
    
    user = st.session_state.get('user', {})
    user_id = user.get('id')
    is_admin = user.get('is_admin', False)
    orders = get_all_orders(search_query, status_filter, user_id, is_admin)
    
    if not orders:
        st.info("לא נמצאו הזמנות")
        return
    
    st.markdown(f"**נמצאו {len(orders)} הזמנות**")
    
    for order in orders:
        with st.container():
            st.markdown(f"""
            <div class="order-card">
                <div style="display: flex; justify-content: space-between; align-items: center;">
                    <div>
                        <h4 style="margin: 0; color: #667eea;">{order.order_number}</h4>
                        <p style="margin: 0.5rem 0;">{order.event_name}</p>
                        <p style="margin: 0; color: #888;">{order.customer_name} | {order.customer_email}</p>
                    </div>
                    <div style="text-align: left;">
                        {get_status_badge(order.status)}
                        <p style="margin: 0.5rem 0; font-size: 0.9rem; color: #888;">
                            {order.created_at.strftime('%d/%m/%Y %H:%M') if order.created_at else ''}
                        </p>
                        <p style="margin: 0; font-weight: bold; color: #38ef7d;">
                            {order.total_euro:.0f}€ = {order.total_nis:.0f}₪
                        </p>
                    </div>
                </div>
            </div>
            """, unsafe_allow_html=True)
            
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                if st.button("📄 צפה בפרטים", key=f"view_{order.id}"):
                    st.session_state.selected_order = order.id
            with col2:
                if order.status == OrderStatus.DRAFT:
                    if st.button("📧 שלח ללקוח", key=f"send_{order.id}"):
                        update_order_status(order.id, OrderStatus.SENT)
                        st.rerun()
            with col3:
                if order.status != OrderStatus.CANCELLED:
                    if st.button("❌ בטל", key=f"cancel_{order.id}"):
                        update_order_status(order.id, OrderStatus.CANCELLED)
                        st.rerun()
            with col4:
                delete_key = f"delete_{order.id}"
                confirm_key = f"confirm_delete_{order.id}"
                if st.session_state.get(confirm_key):
                    st.warning("בטוח למחוק?")
                    c1, c2 = st.columns(2)
                    with c1:
                        if st.button("✅ כן", key=f"yes_{order.id}"):
                            if delete_order(order.id):
                                st.session_state[confirm_key] = False
                                st.rerun()
                    with c2:
                        if st.button("❌ לא", key=f"no_{order.id}"):
                            st.session_state[confirm_key] = False
                            st.rerun()
                else:
                    if st.button("🗑️ מחק", key=delete_key):
                        st.session_state[confirm_key] = True
                        st.rerun()

def page_export():
    """Export page for Excel reports"""
    render_header()
    
    st.markdown("### 📊 ייצוא דוחות")
    
    st.markdown("#### בחר טווח תאריכים")
    
    col1, col2 = st.columns(2)
    with col1:
        start_date = st.date_input("מתאריך", value=datetime.now().replace(day=1))
    with col2:
        end_date = st.date_input("עד תאריך", value=datetime.now())
    
    status_filter = st.multiselect(
        "סינון לפי סטטוס",
        ["טיוטה", "נשלח", "נצפה", "נחתם", "בוטל"],
        default=["נשלח", "נחתם"]
    )
    
    if st.button("📥 ייצא ל-Excel", type="primary", use_container_width=True):
        db = get_db()
        if db:
            try:
                from datetime import timedelta
                query = db.query(Order).filter(
                    Order.created_at >= start_date,
                    Order.created_at <= end_date + timedelta(days=1)
                )
                
                orders = query.all()
                
                if orders:
                    import openpyxl
                    from openpyxl.styles import Font, Alignment, PatternFill
                    
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = "הזמנות"
                    
                    headers = ["מספר הזמנה", "תאריך", "אירוע", "לקוח", "אימייל", "טלפון", 
                               "בלוק", "שורה", "מושבים", "כרטיסים", "סה\"כ יורו", "סה\"כ ש\"ח", "סטטוס"]
                    
                    header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
                    header_font = Font(bold=True, color="FFFFFF")
                    
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col, value=header)
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center")
                    
                    status_hebrew = {
                        OrderStatus.DRAFT: "טיוטה",
                        OrderStatus.SENT: "נשלח",
                        OrderStatus.VIEWED: "נצפה",
                        OrderStatus.SIGNED: "נחתם",
                        OrderStatus.CANCELLED: "בוטל"
                    }
                    
                    for row, order in enumerate(orders, 2):
                        ws.cell(row=row, column=1, value=order.order_number)
                        ws.cell(row=row, column=2, value=order.created_at.strftime('%d/%m/%Y') if order.created_at else '')
                        ws.cell(row=row, column=3, value=order.event_name)
                        ws.cell(row=row, column=4, value=order.customer_name)
                        ws.cell(row=row, column=5, value=order.customer_email)
                        ws.cell(row=row, column=6, value=order.customer_phone)
                        ws.cell(row=row, column=7, value=order.block)
                        ws.cell(row=row, column=8, value=order.row)
                        ws.cell(row=row, column=9, value=order.seats)
                        ws.cell(row=row, column=10, value=order.num_tickets)
                        ws.cell(row=row, column=11, value=order.total_euro)
                        ws.cell(row=row, column=12, value=order.total_nis)
                        ws.cell(row=row, column=13, value=status_hebrew.get(order.status, ''))
                    
                    for col in ws.columns:
                        max_length = 0
                        column = col[0].column_letter
                        for cell in col:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = (max_length + 2)
                        ws.column_dimensions[column].width = adjusted_width
                    
                    excel_buffer = io.BytesIO()
                    wb.save(excel_buffer)
                    excel_buffer.seek(0)
                    
                    st.download_button(
                        label="⬇️ הורד קובץ Excel",
                        data=excel_buffer,
                        file_name=f"הזמנות_TikTik_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                    
                    st.success(f"✅ נמצאו {len(orders)} הזמנות לייצוא")
                else:
                    st.warning("לא נמצאו הזמנות בטווח התאריכים שנבחר")
            except Exception as e:
                st.error(f"שגיאה בייצוא: {str(e)}")
            finally:
                db.close()
        else:
            st.error("לא ניתן להתחבר למסד הנתונים")

def get_random_atmosphere_image(event_type_str):
    """Get a random atmosphere image for the given event type"""
    # Default fallback images based on event type
    fallback_images = {
        "הופעה": "assets/concert_bg.jpg",
        "כדורגל": "assets/cover_page.jpg",
        "אחר": "assets/cover_page.jpg"
    }
    fallback_path = fallback_images.get(event_type_str, "assets/cover_page.jpg")
    
    db = get_db()
    if not db:
        # Return fallback if database unavailable
        if os.path.exists(fallback_path):
            return fallback_path
        return None
    try:
        if event_type_str == "כדורגל":
            category = EventType.FOOTBALL
        elif event_type_str == "הופעה":
            category = EventType.CONCERT
        else:
            category = EventType.OTHER
        
        images = db.query(AtmosphereImage).filter(
            AtmosphereImage.category == category,
            AtmosphereImage.is_active == True
        ).all()
        
        if images:
            selected = random.choice(images)
            return selected.file_path
        
        # No images in database - use fallback
        if os.path.exists(fallback_path):
            return fallback_path
        return None
    except Exception as e:
        print(f"Error getting atmosphere image: {e}")
        # Return fallback on error
        if os.path.exists(fallback_path):
            return fallback_path
        return None
    finally:
        db.close()

def page_image_gallery():
    """Admin page for managing atmosphere images"""
    render_header()
    
    st.markdown("""
    <style>
    .image-card {
        background: white;
        border-radius: 12px;
        box-shadow: 0 2px 8px rgba(0,0,0,0.1);
        padding: 10px;
        margin-bottom: 15px;
        transition: transform 0.2s, box-shadow 0.2s;
    }
    .image-card:hover {
        transform: translateY(-3px);
        box-shadow: 0 4px 15px rgba(0,0,0,0.15);
    }
    .image-card img {
        border-radius: 8px;
        width: 100%;
        height: 150px;
        object-fit: cover;
    }
    .image-card-name {
        font-size: 12px;
        color: #666;
        margin-top: 8px;
        text-align: center;
        white-space: nowrap;
        overflow: hidden;
        text-overflow: ellipsis;
    }
    .gallery-stats {
        background: linear-gradient(135deg, #667eea22, #764ba222);
        border-radius: 10px;
        padding: 15px;
        margin: 15px 0;
    }
    .stat-item {
        display: inline-block;
        margin: 0 20px;
        text-align: center;
    }
    .stat-number {
        font-size: 24px;
        font-weight: bold;
        color: #667eea;
    }
    .stat-label {
        font-size: 12px;
        color: #666;
    }
    </style>
    """, unsafe_allow_html=True)
    
    if st.button("⬅️ חזרה לתפריט הראשי"):
        st.session_state.admin_page = None
        st.rerun()
    
    st.markdown("### 🖼️ ניהול תמונות אווירה")
    st.markdown("העלה תמונות אווירה לפי קטגוריה. התמונות ישמשו באופן אוטומטי ביצירת הזמנות.")
    
    category_options = {
        "⚽ כדורגל": EventType.FOOTBALL,
        "🎵 הופעה": EventType.CONCERT,
        "🎭 אחר": EventType.OTHER
    }
    
    db = get_db()
    if db:
        stats_html = '<div class="gallery-stats" style="text-align: center;">'
        for cat_name, cat_enum in category_options.items():
            count = db.query(AtmosphereImage).filter(
                AtmosphereImage.category == cat_enum,
                AtmosphereImage.is_active == True
            ).count()
            stats_html += f'<div class="stat-item"><div class="stat-number">{count}</div><div class="stat-label">{cat_name}</div></div>'
        stats_html += '</div>'
        st.markdown(stats_html, unsafe_allow_html=True)
        db.close()
    
    st.markdown("---")
    
    selected_category = st.selectbox("בחר קטגוריה להעלאה/צפייה", list(category_options.keys()))
    category = category_options[selected_category]
    
    st.markdown("#### ➕ העלאת תמונות חדשות")
    uploaded_files = st.file_uploader(
        "גרור תמונות לכאן או לחץ לבחירה",
        type=['png', 'jpg', 'jpeg', 'webp'],
        accept_multiple_files=True,
        key="atmosphere_uploader"
    )
    
    if uploaded_files:
        st.info(f"📎 נבחרו {len(uploaded_files)} תמונות - לחץ על 'שמור תמונות' לשמירה במערכת")
        
        # Show preview of uploaded images with error handling
        preview_cols = st.columns(min(len(uploaded_files), 4))
        for idx, f in enumerate(uploaded_files[:4]):
            with preview_cols[idx]:
                try:
                    st.image(f, caption=f.name[:15] + "..." if len(f.name) > 15 else f.name, use_container_width=True)
                except Exception:
                    st.warning(f"⚠️ {f.name[:10]}...")
        if len(uploaded_files) > 4:
            st.caption(f"...ועוד {len(uploaded_files) - 4} תמונות נוספות")
        
        if st.button("💾 שמור תמונות", type="primary", use_container_width=True):
            db = get_db()
            if db:
                saved_count = 0
                for uploaded_file in uploaded_files:
                    folder = f"attached_assets/atmosphere_images/{category.value}"
                    os.makedirs(folder, exist_ok=True)
                    
                    filename = f"{uuid.uuid4().hex[:8]}_{uploaded_file.name}"
                    file_path = os.path.join(folder, filename)
                    
                    with open(file_path, "wb") as f:
                        f.write(uploaded_file.getbuffer())
                    
                    new_image = AtmosphereImage(
                        filename=uploaded_file.name,
                        category=category,
                        file_path=file_path,
                        is_active=True
                    )
                    db.add(new_image)
                    saved_count += 1
                
                db.commit()
                db.close()
                st.success(f"✅ הועלו {saved_count} תמונות בהצלחה!")
                st.rerun()
    
    st.markdown("---")
    st.markdown("#### 🖼️ תמונות קיימות")
    
    db = get_db()
    if db:
        images = db.query(AtmosphereImage).filter(
            AtmosphereImage.category == category,
            AtmosphereImage.is_active == True
        ).order_by(AtmosphereImage.created_at.desc()).all()
        
        if images:
            cols = st.columns(4)
            for idx, img in enumerate(images):
                with cols[idx % 4]:
                    if os.path.exists(img.file_path):
                        st.markdown('<div class="image-card">', unsafe_allow_html=True)
                        st.image(img.file_path, use_container_width=True)
                        short_name = img.filename[:20] + "..." if len(img.filename) > 20 else img.filename
                        st.markdown(f'<div class="image-card-name">{short_name}</div>', unsafe_allow_html=True)
                        if st.button("🗑️ מחק", key=f"del_{img.id}", use_container_width=True):
                            img.is_active = False
                            db.commit()
                            st.rerun()
                        st.markdown('</div>', unsafe_allow_html=True)
        else:
            st.info("🖼️ אין תמונות בקטגוריה זו. העלה תמונות חדשות למעלה.")
        
        db.close()

def login_user(username, password):
    """Authenticate user and return user object if successful"""
    db = get_db()
    if not db:
        return None
    try:
        user = db.query(User).filter(User.username == username, User.is_active == True).first()
        if user and user.check_password(password):
            user.last_login = datetime.utcnow()
            db.commit()
            return {
                'id': user.id,
                'username': user.username,
                'full_name': user.full_name,
                'is_admin': user.is_admin,
                'email': user.email
            }
        return None
    except Exception as e:
        print(f"Login error: {e}")
        return None
    finally:
        db.close()

def reset_user_password(email_or_username):
    """Reset user password and send email with temporary password"""
    import secrets
    db = get_db()
    if not db:
        return False, "שגיאה בחיבור למסד הנתונים"
    
    try:
        user = db.query(User).filter(
            ((User.email == email_or_username) | (User.username == email_or_username)),
            User.is_active == True
        ).first()
        
        if not user:
            return False, "משתמש לא נמצא"
        
        if not user.email:
            return False, "למשתמש זה אין כתובת אימייל מוגדרת"
        
        temp_password = secrets.token_urlsafe(8)
        user.set_password(temp_password)
        db.commit()
        
        try:
            import resend
            resend_api_key = os.environ.get('RESEND_API_KEY')
            if not resend_api_key:
                return False, "שירות האימייל לא מוגדר. פנה למנהל המערכת."
            
            resend.api_key = resend_api_key
            
            email_html = f"""
            <div dir="rtl" style="font-family: Arial, sans-serif; padding: 20px;">
                <h2>🔐 איפוס סיסמה - TikTik</h2>
                <p>שלום {user.full_name},</p>
                <p>הסיסמה שלך אופסה בהצלחה.</p>
                <p><strong>הסיסמה הזמנית החדשה שלך:</strong></p>
                <p style="font-size: 24px; background: #f0f0f0; padding: 15px; border-radius: 8px; text-align: center;">
                    <code>{temp_password}</code>
                </p>
                <p>אנא התחבר והחלף את הסיסמה בהקדם.</p>
                <hr>
                <p style="color: #666; font-size: 12px;">מערכת TikTik Smart Order System</p>
            </div>
            """
            
            resend.Emails.send({
                "from": "TikTik System <onboarding@resend.dev>",
                "to": [user.email],
                "subject": "🔐 איפוס סיסמה - TikTik",
                "html": email_html
            })
            
            return True, f"סיסמה זמנית נשלחה לאימייל: {user.email[:3]}***{user.email.split('@')[1]}"
            
        except Exception as e:
            print(f"Email send error: {e}")
            return False, "לא ניתן לשלוח אימייל. אנא פנה למנהל המערכת לאיפוס הסיסמה."
            
    except Exception as e:
        db.rollback()
        print(f"Password reset error: {e}")
        return False, f"שגיאה באיפוס הסיסמה: {str(e)}"
    finally:
        db.close()

def page_login():
    """Login page with quick user selection"""
    st.markdown("""
    <div class="header-container">
        <h1>🎟️ TikTik Smart Order System</h1>
        <p>מערכת חכמה ליצירת הצעות מחיר והזמנות מקצועיות</p>
    </div>
    """, unsafe_allow_html=True)
    
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        st.markdown('<div class="form-section">', unsafe_allow_html=True)
        st.markdown("### 🔐 התחברות למערכת")
        
        saved_users_display = []
        db = get_db()
        if db:
            try:
                users = db.query(User).filter(User.is_active == True).all()
                saved_users_display = [(u.username, u.full_name) for u in users]
            except:
                pass
            finally:
                db.close()
        
        if saved_users_display:
            st.markdown("##### ⚡ כניסה מהירה")
            user_options = ["-- בחר משתמש --"] + [f"{u[1]} ({u[0]})" for u in saved_users_display]
            selected_quick = st.selectbox("בחר משתמש", user_options, key="quick_user", label_visibility="collapsed")
            
            if selected_quick and selected_quick != "-- בחר משתמש --":
                quick_username = selected_quick.split("(")[-1].replace(")", "").strip()
                quick_password = st.text_input("סיסמה", type="password", key="quick_password")
                
                if st.button("🚀 התחבר", use_container_width=True, key="quick_login_btn"):
                    if quick_password:
                        user = login_user(quick_username, quick_password)
                        if user:
                            st.session_state.user = user
                            st.session_state.logged_in = True
                            set_session_token(user)
                            st.success(f"👋 שלום {user['full_name']}!")
                            st.rerun()
                        else:
                            st.error("❌ סיסמה שגויה")
                    else:
                        st.warning("⚠️ נא להזין סיסמה")
            
            st.markdown("---")
            with st.expander("📝 התחברות ידנית"):
                username = st.text_input("שם משתמש", key="login_username")
                password = st.text_input("סיסמה", type="password", key="login_password")
                
                if st.button("🚀 התחבר", use_container_width=True, key="manual_login_btn"):
                    if username and password:
                        user = login_user(username, password)
                        if user:
                            st.session_state.user = user
                            st.session_state.logged_in = True
                            set_session_token(user)
                            st.success(f"👋 שלום {user['full_name']}!")
                            st.rerun()
                        else:
                            st.error("❌ שם משתמש או סיסמה שגויים")
                    else:
                        st.warning("⚠️ נא למלא שם משתמש וסיסמה")
        else:
            username = st.text_input("שם משתמש", key="login_username")
            password = st.text_input("סיסמה", type="password", key="login_password")
            
            if st.button("🚀 התחבר", use_container_width=True):
                if username and password:
                    user = login_user(username, password)
                    if user:
                        st.session_state.user = user
                        st.session_state.logged_in = True
                        set_session_token(user)
                        st.success(f"👋 שלום {user['full_name']}!")
                        st.rerun()
                    else:
                        st.error("❌ שם משתמש או סיסמה שגויים")
                else:
                    st.warning("⚠️ נא למלא שם משתמש וסיסמה")
        
        st.markdown("---")
        with st.expander("🔑 שכחתי סיסמה"):
            st.markdown("הזן את שם המשתמש או האימייל שלך ונשלח לך סיסמה זמנית:")
            reset_identifier = st.text_input("שם משתמש / אימייל", key="reset_identifier")
            
            if st.button("📧 שלח סיסמה זמנית", use_container_width=True, key="reset_password_btn"):
                if reset_identifier:
                    with st.spinner("מאפס סיסמה..."):
                        success, message = reset_user_password(reset_identifier)
                        if success:
                            st.success(f"✅ {message}")
                        else:
                            st.error(f"❌ {message}")
                else:
                    st.warning("⚠️ נא להזין שם משתמש או אימייל")
        
        st.markdown('</div>', unsafe_allow_html=True)

def page_user_management():
    """Admin page for user management"""
    st.markdown("""
    <div class="header-container">
        <h1>👥 ניהול משתמשים</h1>
        <p>צור וערוך משתמשים במערכת</p>
    </div>
    """, unsafe_allow_html=True)
    
    if st.button("⬅️ חזרה לתפריט"):
        st.session_state.admin_page = None
        st.rerun()
    
    st.markdown("---")
    
    col1, col2 = st.columns([1, 1])
    
    with col1:
        st.markdown('<div class="form-section">', unsafe_allow_html=True)
        st.markdown("### ➕ הוספת משתמש חדש")
        
        new_username = st.text_input("שם משתמש", key="new_user_username")
        new_fullname = st.text_input("שם מלא", key="new_user_fullname")
        new_email = st.text_input("אימייל", key="new_user_email")
        new_password = st.text_input("סיסמה", type="password", key="new_user_password")
        new_is_admin = st.checkbox("מנהל מערכת", key="new_user_admin")
        
        if st.button("✅ צור משתמש", use_container_width=True):
            if new_username and new_fullname and new_email and new_password:
                db = get_db()
                if db:
                    try:
                        existing = db.query(User).filter(
                            (User.username == new_username) | (User.email == new_email)
                        ).first()
                        if existing:
                            st.error("❌ שם משתמש או אימייל כבר קיימים")
                        else:
                            user = User(
                                username=new_username,
                                email=new_email,
                                full_name=new_fullname,
                                is_admin=new_is_admin,
                                is_active=True
                            )
                            user.set_password(new_password)
                            db.add(user)
                            db.commit()
                            st.success(f"✅ המשתמש {new_fullname} נוצר בהצלחה!")
                            st.rerun()
                    except Exception as e:
                        db.rollback()
                        st.error(f"❌ שגיאה: {str(e)}")
                    finally:
                        db.close()
            else:
                st.warning("⚠️ נא למלא את כל השדות")
        
        st.markdown('</div>', unsafe_allow_html=True)
    
    with col2:
        st.markdown('<div class="form-section">', unsafe_allow_html=True)
        st.markdown("### 📋 משתמשים קיימים")
        
        db = get_db()
        if db:
            users = db.query(User).order_by(User.created_at.desc()).all()
            for user in users:
                status = "🟢" if user.is_active else "🔴"
                admin_badge = " 👑" if user.is_admin else ""
                st.markdown(f"""
                <div class="passenger-item">
                    <strong>{status} {user.full_name}{admin_badge}</strong><br>
                    <small>@{user.username} | {user.email}</small>
                </div>
                """, unsafe_allow_html=True)
                
                if user.username != "admin":
                    with st.expander("🔧 פעולות"):
                        new_pass = st.text_input("סיסמה חדשה", type="password", key=f"new_pass_{user.id}", placeholder="השאר ריק לאיפוס ל-123456")
                        col_a, col_b, col_c = st.columns(3)
                        with col_a:
                            if st.button("🔄 שנה סיסמה", key=f"reset_{user.id}"):
                                password_to_set = new_pass if new_pass else "123456"
                                user.set_password(password_to_set)
                                db.commit()
                                if new_pass:
                                    st.success(f"✅ הסיסמה שונתה!")
                                else:
                                    st.success(f"✅ סיסמה אופסה ל-123456")
                        with col_b:
                            if user.is_active:
                                if st.button("🚫 השבת", key=f"disable_{user.id}"):
                                    user.is_active = False
                                    db.commit()
                                    st.rerun()
                            else:
                                if st.button("✅ הפעל", key=f"enable_{user.id}"):
                                    user.is_active = True
                                    db.commit()
                                    st.rerun()
                        with col_c:
                            if user.is_admin:
                                if st.button("👤 הסר מנהל", key=f"demote_{user.id}"):
                                    user.is_admin = False
                                    db.commit()
                                    st.rerun()
                            else:
                                if st.button("👑 הפוך למנהל", key=f"promote_{user.id}"):
                                    user.is_admin = True
                                    db.commit()
                                    st.rerun()
                        
                        st.markdown("---")
                        delete_key = f"delete_confirm_{user.id}"
                        if delete_key not in st.session_state:
                            st.session_state[delete_key] = False
                        
                        if not st.session_state[delete_key]:
                            if st.button("🗑️ מחק משתמש", key=f"delete_{user.id}", type="secondary"):
                                st.session_state[delete_key] = True
                                st.rerun()
                        else:
                            st.warning(f"⚠️ האם אתה בטוח שברצונך למחוק את {user.full_name}?")
                            col_yes, col_no = st.columns(2)
                            with col_yes:
                                if st.button("✅ כן, מחק", key=f"confirm_delete_{user.id}", type="primary"):
                                    try:
                                        db.delete(user)
                                        db.commit()
                                        st.success(f"✅ המשתמש {user.full_name} נמחק!")
                                        st.session_state[delete_key] = False
                                        st.rerun()
                                    except Exception as e:
                                        db.rollback()
                                        st.error(f"❌ שגיאה במחיקה: {str(e)}")
                            with col_no:
                                if st.button("❌ ביטול", key=f"cancel_delete_{user.id}"):
                                    st.session_state[delete_key] = False
                                    st.rerun()
            db.close()
        
        st.markdown('</div>', unsafe_allow_html=True)

def page_change_password():
    """Page for users to change their own password"""
    st.markdown("""
    <div class="header-container">
        <h1>🔑 שינוי סיסמה</h1>
        <p>שנה את הסיסמה שלך</p>
    </div>
    """, unsafe_allow_html=True)
    
    if st.button("⬅️ חזרה לתפריט"):
        st.session_state.admin_page = None
        st.rerun()
    
    st.markdown("---")
    
    user = st.session_state.get('user', {})
    user_id = user.get('id')
    
    if not user_id:
        st.error("❌ לא נמצא משתמש מחובר")
        return
    
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        st.markdown('<div class="form-section">', unsafe_allow_html=True)
        st.markdown(f"### 👤 {user.get('full_name', 'משתמש')}")
        st.markdown(f"📧 {user.get('email', '')}")
        
        st.markdown("---")
        st.markdown("##### 🔐 הזן סיסמה חדשה")
        
        current_password = st.text_input("סיסמה נוכחית", type="password", key="current_pass")
        new_password = st.text_input("סיסמה חדשה", type="password", key="new_pass")
        confirm_password = st.text_input("אימות סיסמה חדשה", type="password", key="confirm_pass")
        
        if st.button("✅ שנה סיסמה", use_container_width=True):
            if not current_password or not new_password or not confirm_password:
                st.warning("⚠️ נא למלא את כל השדות")
            elif new_password != confirm_password:
                st.error("❌ הסיסמאות החדשות לא תואמות")
            elif len(new_password) < 4:
                st.error("❌ הסיסמה חייבת להכיל לפחות 4 תווים")
            else:
                db = get_db()
                if db:
                    try:
                        db_user = db.query(User).filter(User.id == user_id).first()
                        if db_user and db_user.check_password(current_password):
                            db_user.set_password(new_password)
                            db.commit()
                            st.success("✅ הסיסמה שונתה בהצלחה!")
                        else:
                            st.error("❌ הסיסמה הנוכחית שגויה")
                    except Exception as e:
                        db.rollback()
                        st.error(f"❌ שגיאה: {str(e)}")
                    finally:
                        db.close()
        
        st.markdown('</div>', unsafe_allow_html=True)

def page_package_templates():
    """Page for managing package templates"""
    st.markdown("""
    <div class="header-container">
        <h1>📦 חבילות קבועות</h1>
        <p>ניהול חבילות קבועות לשימוש חוזר</p>
    </div>
    """, unsafe_allow_html=True)
    
    if st.button("⬅️ חזרה לתפריט"):
        st.session_state.admin_page = None
        st.rerun()
    
    st.markdown("---")
    
    st.info("💡 **ליצור חבילה חדשה:** מלא את טופס ההזמנה הרגיל ולחץ על '📦 שמור כחבילה קבועה' בסוף.")
    
    st.markdown("### 📋 חבילות שמורות")
    
    db = get_db()
    packages = []
    if db:
        try:
            packages = db.query(PackageTemplate).filter(PackageTemplate.is_active == True).order_by(PackageTemplate.created_at.desc()).all()
        except:
            pass
        finally:
            db.close()
    
    if not packages:
        st.info("📦 אין חבילות שמורות. לך להזמנה חדשה ושמור חבילה משם.")
    else:
        st.markdown(f"**סה\"כ: {len(packages)} חבילות**")
        
        for pkg in packages:
            pkg_data = pkg.to_dict()
            with st.container():
                st.markdown('<div class="form-section">', unsafe_allow_html=True)
                
                col1, col2, col3 = st.columns([4, 2, 2])
                
                with col1:
                    event_emoji = "🎤" if pkg_data.get('event_type') == 'concert' else "⚽" if pkg_data.get('event_type') == 'football' else "🎭"
                    st.markdown(f"### {event_emoji} {pkg_data.get('name', 'ללא שם')}")
                    if pkg_data.get('event_name'):
                        st.markdown(f"🎫 {pkg_data.get('event_name')}")
                    if pkg_data.get('venue'):
                        st.markdown(f"📍 {pkg_data.get('venue')}")
                
                with col2:
                    if pkg_data.get('event_date'):
                        st.markdown(f"📅 {pkg_data.get('event_date')}")
                    if pkg_data.get('ticket_category'):
                        st.markdown(f"🎟️ {pkg_data.get('ticket_category')}")
                    if pkg_data.get('package_price_euro'):
                        st.markdown(f"💶 {pkg_data.get('package_price_euro'):.0f}€ לאדם")
                
                with col3:
                    btn_col1, btn_col2 = st.columns(2)
                    with btn_col1:
                        if st.button("📋 שכפל", key=f"dup_pkg_{pkg.id}", use_container_width=True):
                            dup_db = get_db()
                            if dup_db:
                                try:
                                    new_pkg = PackageTemplate(
                                        name=f"{pkg.name} (עותק)",
                                        event_type=pkg.event_type,
                                        product_type=pkg.product_type,
                                        event_name=pkg.event_name,
                                        event_date=pkg.event_date,
                                        event_time=pkg.event_time,
                                        venue=pkg.venue,
                                        ticket_description=pkg.ticket_description,
                                        ticket_category=pkg.ticket_category,
                                        price_per_ticket_euro=pkg.price_per_ticket_euro,
                                        hotel_data=pkg.hotel_data,
                                        flight_data=pkg.flight_data,
                                        package_price_euro=pkg.package_price_euro,
                                        stadium_map_data=pkg.stadium_map_data,
                                        stadium_map_mime=pkg.stadium_map_mime,
                                        notes=pkg.notes
                                    )
                                    dup_db.add(new_pkg)
                                    dup_db.commit()
                                    st.success("✅ החבילה שוכפלה!")
                                    st.rerun()
                                except:
                                    dup_db.rollback()
                                finally:
                                    dup_db.close()
                    with btn_col2:
                        if st.button("🗑️", key=f"del_pkg_{pkg.id}", use_container_width=True):
                            del_db = get_db()
                            if del_db:
                                try:
                                    del_db.query(PackageTemplate).filter(PackageTemplate.id == pkg.id).update({'is_active': False})
                                    del_db.commit()
                                    st.success("✅ החבילה נמחקה!")
                                    st.rerun()
                                except:
                                    del_db.rollback()
                                finally:
                                    del_db.close()
                
                with st.expander("📄 פרטים מלאים"):
                    col1, col2 = st.columns(2)
                    with col1:
                        st.markdown("**פרטי אירוע:**")
                        st.write(f"- סוג: {pkg_data.get('event_type')}")
                        st.write(f"- תאריך: {pkg_data.get('event_date')} {pkg_data.get('event_time')}")
                        st.write(f"- מקום: {pkg_data.get('venue')}")
                        st.write(f"- קטגוריה: {pkg_data.get('ticket_category')}")
                        if pkg_data.get('ticket_description'):
                            st.write(f"- תיאור: {pkg_data.get('ticket_description')}")
                    
                    with col2:
                        hotel = pkg_data.get('hotel', {})
                        if hotel:
                            st.markdown("**פרטי מלון:**")
                            st.write(f"- שם: {hotel.get('name', 'לא הוגדר')}")
                            st.write(f"- צ'ק-אין: {hotel.get('check_in', '')}")
                            st.write(f"- צ'ק-אאוט: {hotel.get('check_out', '')}")
                        
                        flights = pkg_data.get('flights', {})
                        if flights:
                            st.markdown("**טיסות:**")
                            outbound = flights.get('outbound', {})
                            if outbound:
                                st.write(f"- הלוך: {outbound.get('date')} {outbound.get('time')} | {outbound.get('flight_number')}")
                            ret = flights.get('return', {})
                            if ret:
                                st.write(f"- חזור: {ret.get('date')} {ret.get('time')} | {ret.get('flight_number')}")
                
                st.markdown('</div>', unsafe_allow_html=True)
                st.markdown("")

def page_saved_concerts():
    """Page for managing saved concerts and artists"""
    st.markdown("""
    <div class="header-container">
        <h1>⭐ אמנים והופעות שמורים</h1>
        <p>ניהול אמנים והופעות שנשמרו לשימוש חוזר</p>
    </div>
    """, unsafe_allow_html=True)
    
    if st.button("⬅️ חזרה לתפריט"):
        st.session_state.admin_page = None
        st.rerun()
    
    st.markdown("---")
    
    tab1, tab2 = st.tabs(["🎤 אמנים שמורים", "🎵 הופעות שמורות"])
    
    with tab1:
        saved_artists = get_saved_artists()
        
        if not saved_artists:
            st.info("🎤 אין אמנים שמורים. חפש אמן והוסף אותו לרשימה שלך.")
        else:
            st.markdown(f"### 🎤 {len(saved_artists)} אמנים שמורים")
            
            for artist in saved_artists:
                with st.container():
                    st.markdown('<div class="form-section">', unsafe_allow_html=True)
                    
                    col1, col2 = st.columns([4, 1])
                    
                    with col1:
                        st.markdown(f"**⭐ {artist.get('name_he', artist.get('name_en', 'לא ידוע'))}**")
                        if artist.get('genre'):
                            st.caption(f"🎸 {artist.get('genre')}")
                    
                    with col2:
                        if st.button("🗑️ מחק", key=f"delete_artist_{artist.get('db_id')}", use_container_width=True):
                            if delete_saved_artist(artist.get('db_id')):
                                st.success("✅ אמן הוסר!")
                                st.rerun()
                            else:
                                st.error("❌ שגיאה במחיקה")
                    
                    st.markdown('</div>', unsafe_allow_html=True)
                    st.markdown("")
    
    with tab2:
        saved_concerts = get_saved_concerts()
        
        if not saved_concerts:
            st.info("🎵 אין הופעות שמורות. הופעות שתשמור מההזמנה יופיעו כאן.")
        else:
            st.markdown(f"### 📋 {len(saved_concerts)} הופעות שמורות")
            
            for concert in saved_concerts:
                with st.container():
                    st.markdown('<div class="form-section">', unsafe_allow_html=True)
                    
                    col1, col2, col3 = st.columns([3, 2, 1])
                    
                    with col1:
                        st.markdown(f"**🎤 {concert.get('artist', 'לא ידוע')}**")
                        st.markdown(f"📍 {concert.get('venue', 'לא ידוע')}")
                        if concert.get('city') or concert.get('country'):
                            st.markdown(f"🌍 {concert.get('city', '')} {concert.get('country', '')}")
                    
                    with col2:
                        if concert.get('date'):
                            st.markdown(f"📅 {concert.get('date')}")
                        if concert.get('time'):
                            st.markdown(f"🕐 {concert.get('time')}")
                        if concert.get('category'):
                            st.markdown(f"🏷️ {concert.get('category')}")
                    
                    with col3:
                        if st.button("🗑️ מחק", key=f"delete_concert_{concert.get('id')}", use_container_width=True):
                            if delete_saved_concert(concert.get('id')):
                                st.success("✅ הופעה נמחקה!")
                                st.rerun()
                            else:
                                st.error("❌ שגיאה במחיקה")
                    
                    st.markdown('</div>', unsafe_allow_html=True)
                    st.markdown("")

def page_beginner_guide():
    """Beginner's guide page with detailed step-by-step instructions"""
    st.markdown("""
    <div class="header-container">
        <h1>📖 מדריך למתחיל</h1>
        <p>מדריך צעד-אחר-צעד ליצירת הזמנה ראשונה במערכת TikTik</p>
    </div>
    """, unsafe_allow_html=True)
    
    if st.button("⬅️ חזרה לתפריט"):
        st.session_state.admin_page = None
        st.rerun()
    
    st.markdown("---")
    
    st.success("🎯 **המטרה:** ליצור הזמנה מקצועית ולשלוח אותה ללקוח בוואטסאפ תוך דקות ספורות!")
    
    st.markdown("---")
    
    st.markdown("## 📋 השלבים ליצירת הזמנה")
    
    with st.expander("🔐 שלב 1: התחברות למערכת", expanded=True):
        st.markdown("""
        1. הזן את **שם המשתמש** שלך
        2. הזן את **הסיסמה** שלך
        3. לחץ על **"התחבר"**
        
        💡 **טיפ:** אם שכחת סיסמה - פנה למנהל המערכת
        """)
    
    with st.expander("📦 שלב 2: בחירת סוג המוצר"):
        st.markdown("""
        בחר את סוג ההזמנה:
        
        | סוג | מה כולל? |
        |-----|----------|
        | **🎫 כרטיסים בלבד** | רק כרטיסים לאירוע |
        | **📦 חבילה מלאה** | כרטיסים + מלון + טיסות + העברות |
        
        ⚡ **המערכת תתאים את הטופס אוטומטית לפי הבחירה שלך**
        """)
    
    with st.expander("🎭 שלב 3: בחירת סוג האירוע"):
        st.markdown("""
        בחר את סוג האירוע:
        
        - **⚽ כדורגל** - משחקי ליגות, מוקדמות מונדיאל, גביעים
        - **🎤 הופעה** - הופעות מוזיקה, פסטיבלים
        - **🎭 אחר** - אירועים אחרים
        
        💡 **לכדורגל:** המערכת תציג רשימת קבוצות עם מפות אצטדיון אוטומטיות!
        """)
    
    with st.expander("⚽ שלב 4: פרטי האירוע"):
        st.markdown("""
        ### לאירועי כדורגל:
        1. בחר **קבוצת בית** מהרשימה
        2. בחר **קבוצה אורחת** מהרשימה
        3. ✨ המערכת תמלא אוטומטית:
           - שם האצטדיון
           - העיר
           - מפת מושבים
        4. בחר **קטגוריית מושבים** (אם יש)
        5. הזן **תאריך ושעה**
        
        ### להופעות:
        1. חפש את **האמן** או בחר מהופעות שמורות
        2. הזן את **מקום ההופעה**
        3. הזן **תאריך ושעה**
        
        💡 **טיפ:** ניתן להעלות מפת מושבים ידנית אם אין מפה אוטומטית
        """)
    
    with st.expander("👥 שלב 5: הוספת נוסעים - ⚠️ שלב חשוב!"):
        st.markdown("""
        ### אפשרות מומלצת - סריקת דרכון 📸
        
        **שלב א:** העלה תמונת דרכון
        - לחץ על "העלה תמונת דרכון"
        - בחר תמונה מהמחשב
        
        **שלב ב: ⚠️ חובה ללחוץ על הכפתור!**
        """)
        st.error("👆 לחץ על הכפתור: **🔍 סרוק דרכון** ← אחרי שהעלית תמונה!")
        st.markdown("""
        **שלב ג:** המתן לסריקה
        - המערכת תזהה אוטומטית:
          - ✅ שם מלא
          - ✅ מספר דרכון
          - ✅ תאריך לידה
          - ✅ תוקף דרכון
        
        **שלב ד:** בדוק שהפרטים נכונים ולחץ "הוסף נוסע"
        
        ---
        
        ### אפשרות חלופית - הזנה ידנית
        - מלא את כל השדות ידנית
        - לחץ "הוסף נוסע"
        
        🔄 **חזור על השלב עבור כל נוסע נוסף**
        """)
    
    with st.expander("🏨 שלב 6: פרטי מלון (בחבילה מלאה)"):
        st.markdown("""
        1. הקלד את **שם המלון + עיר** (לדוגמה: "Hilton Madrid")
        2. לחץ על **🔍 חפש מלון**
        3. ✨ המערכת תביא אוטומטית:
           - שם מלא
           - כתובת
           - דירוג כוכבים
           - 2 תמונות למסמך PDF
        
        4. הזן את מספר **הלילות**
        5. בחר את סוג **הארוחות** (אם יש)
        
        💡 **טיפ:** ככל שהחיפוש יותר מדויק - התוצאה יותר טובה
        """)
    
    with st.expander("✈️ שלב 7: פרטי טיסות (בחבילה מלאה)"):
        st.markdown("""
        ### אפשרות מומלצת - סריקת טיסה 📸
        
        **שלב א:** העלה צילום מסך של פרטי הטיסה
        
        **שלב ב: ⚠️ חובה ללחוץ על הכפתור!**
        """)
        st.error("👆 לחץ על הכפתור: **🔍 סרוק טיסה** ← אחרי שהעלית צילום!")
        st.markdown("""
        **שלב ג:** המערכת תזהה:
        - ✅ מספר טיסה
        - ✅ חברת תעופה
        - ✅ שדות המראה ונחיתה
        - ✅ תאריכים ושעות
        
        ---
        
        ### אפשרות חלופית - הזנה ידנית
        - מלא את פרטי הטיסות ידנית
        - בחר שדות תעופה מהרשימה
        
        📦 **כבודה:** סמן אם כלול טרולי / מזוודה
        """)
    
    with st.expander("💰 שלב 8: תמחור"):
        st.markdown("""
        1. הזן **מחיר לכרטיס** (ביורו)
        2. הזן **מספר כרטיסים**
        
        ✨ **המערכת מחשבת אוטומטית:**
        - סה"כ ביורו
        - שער המרה (בנק ישראל + 5 אגורות)
        - סה"כ בשקלים
        
        💡 **שער ההמרה מתעדכן אוטומטית - אין צורך להזין!**
        """)
    
    with st.expander("👤 שלב 9: פרטי לקוח"):
        st.markdown("""
        מלא את פרטי הלקוח:
        
        - **שם מלא** - שם הלקוח שיופיע על ההזמנה
        - **מספר טלפון** - לקשר עם הלקוח
        - **אימייל** - לשליחת ההזמנה
        - **מספר ת.ז.** - לרישום
        """)
    
    with st.expander("📄 שלב 10: יצירת PDF ושליחה"):
        st.markdown("""
        ### יצירת ה-PDF
        1. לחץ על **📄 צור PDF והורד**
        2. המתן ליצירת המסמך
        3. הקובץ יורד אוטומטית למחשב
        
        ### שליחה ללקוח 📱
        1. פתח את הוואטסאפ
        2. בחר את השיחה עם הלקוח
        3. צרף את קובץ ה-PDF שהורדת
        4. שלח! 🚀
        
        ✅ **ההזמנה נשמרה אוטומטית בהיסטוריה**
        """)
    
    st.markdown("---")
    
    st.info("""
    ### 🆘 צריך עזרה נוספת?
    - לחץ על **💬 שאל שאלה** בתפריט הצד לשיחה עם העוזר החכם
    - לחץ על **❓ עזרה** לצפייה במדריך המלא
    """)

def page_help():
    """Help page with comprehensive Hebrew guide for using the system"""
    st.markdown("""
    <div class="header-container">
        <h1>❓ עזרה ומדריך למשתמש</h1>
        <p>מדריך מקיף לשימוש במערכת TikTik להזמנות חכמות</p>
    </div>
    """, unsafe_allow_html=True)
    
    if st.button("⬅️ חזרה לתפריט"):
        st.session_state.admin_page = None
        st.rerun()
    
    st.markdown("---")
    
    with st.expander("📋 סקירה כללית של המערכת", expanded=True):
        st.markdown("""
        ### מה זו מערכת TikTik?
        
        מערכת TikTik היא כלי חכם ליצירת הצעות מחיר והזמנות מקצועיות עבור כרטיסים לאירועי ספורט והופעות באירופה.
        
        **יכולות המערכת:**
        - ✅ יצירת הזמנות מקצועיות בקליק
        - ✅ יצירת PDF מעוצב עם תנאי ביטול
        - ✅ שליחת הזמנות במייל ללקוחות
        - ✅ מעקב אחר סטטוס הזמנות
        - ✅ ייצוא דוחות לאקסל
        - ✅ זיהוי אוטומטי של פרטי דרכון
        - ✅ מפות אצטדיון אוטומטיות
        - ✅ חיפוש מלונות אוטומטי
        """)
    
    with st.expander("🆕 איך ליצור הזמנה חדשה?"):
        st.markdown("""
        ### שלב 1: בחירת סוג המוצר
        בחר בין **חבילה מלאה** (כולל מלון, טיסות, העברות וכרטיסים) או **כרטיסים בלבד**.
        
        ### שלב 2: בחירת סוג האירוע
        - **⚽ כדורגל** - משחקי ליגות אירופאיות, מוקדמות מונדיאל, גביעים
        - **🎤 הופעה** - הופעות מוזיקה ופסטיבלים
        - **🎭 אחר** - אירועים נוספים
        
        ### שלב 3: מילוי פרטי האירוע
        - **לכדורגל**: בחר קבוצת בית ואורחת - המערכת תמלא אוטומטית את שם האצטדיון ומפת המושבים
        - **להופעות**: חפש אמן או בחר מהופעות שמורות
        
        ### שלב 4: הוספת נוסעים
        **חשוב מאוד!** יש שתי דרכים להוסיף נוסע:
        
        **אפשרות 1 - סריקת דרכון (מומלץ):**
        1. לחץ על "העלה תמונת דרכון" ובחר תמונה
        2. **לחץ על כפתור "🔍 סרוק דרכון"** ← זה הכפתור החשוב!
        3. המתן שהמערכת תזהה את הפרטים
        4. הפרטים יתמלאו אוטומטית: שם, מספר דרכון, תאריך לידה, תוקף
        
        **אפשרות 2 - הזנה ידנית:**
        - מלא את השדות ידנית
        
        ### שלב 5: פרטי מלון (בחבילה מלאה)
        - הקלד שם מלון והמערכת תחפש אוטומטית פרטים מלאים
        - תמונות המלון ייכללו ב-PDF
        
        ### שלב 6: פרטי טיסות (בחבילה מלאה)
        **אפשרות 1 - סריקת צילום מסך (מומלץ):**
        1. צלם מסך מאתר הטיסות (כל האתרים עובדים)
        2. העלה את התמונה לשדה "העלה צילום מסך טיסה"
        3. **לחץ על "🔍 סרוק טיסה"**
        4. המערכת תזהה: חברה, מספר טיסה, מוצא, יעד, תאריך ושעה
        
        **אפשרות 2 - הזנה ידנית:**
        - מלא את כל השדות: חברת תעופה, מספר טיסה, שדות, תאריכים
        
        ### שלב 7: תמחור
        - הזן מחיר לכרטיס באירו
        - שער ההמרה מתעדכן **אוטומטית** (שער בנק ישראל + 5 אגורות)
        - המערכת תחשב את הסכום הכולל בשקלים
        
        ### שלב 8: פרטי הלקוח
        - שם מלא, טלפון, אימייל, תעודת זהות
        
        ### שלב 9: יצירת ההזמנה
        - לחץ **צור PDF והורד** - הורדה למחשב שלך
        - שלח את ה-PDF ללקוח דרך **וואטסאפ**
        """)
    
    with st.expander("⚽ עבודה עם אירועי כדורגל"):
        st.markdown("""
        ### בחירת קבוצות
        בחר קבוצת בית וקבוצה אורחת מהרשימה. המערכת תמצא אוטומטית:
        - 📍 שם האצטדיון
        - 🗺️ מפת מושבים (אם קיימת)
        - 📅 תאריך ושעת המשחק (מלוח המשחקים)
        
        ### קטגוריות כרטיסים
        הזן את קטגוריית הכרטיסים (למשל: CAT1, CAT2, VIP)
        
        ### מונדיאל 2026
        המערכת כוללת את כל אצטדיוני המונדיאל 2026 בארה"ב, קנדה ומקסיקו עם מפות מושבים.
        """)
    
    with st.expander("🎤 עבודה עם הופעות"):
        st.markdown("""
        ### חיפוש הופעות
        - חפש לפי שם אמן - המערכת מחפשת ב-Ticketmaster ומקורות נוספים
        - בחר הופעה מהתוצאות - פרטי המקום והתאריך יתמלאו אוטומטית
        
        ### הופעות שמורות
        - שמור הופעות שחוזרות לשימוש מהיר
        - גש להופעות שמורות מכפתור ⭐ בתפריט הצד
        
        ### מפות מקום
        - העלה מפת מושבים או הדבק לינק לתמונה
        - המערכת תשמור את המפה לשימוש עתידי
        """)
    
    with st.expander("🛂 זיהוי דרכון אוטומטי (OCR)"):
        st.markdown("""
        ### איך זה עובד? (שלב אחר שלב)
        1. העלה תמונה של דרכון (צילום ברור של עמוד הפרטים)
        2. או הדבק תמונה מהלוח (Ctrl+V)
        3. **⚠️ חשוב! לחץ על כפתור "🔍 סרוק דרכון"**
        4. המתן כמה שניות
        5. הפרטים יתמלאו אוטומטית:
           - שם פרטי ושם משפחה
           - מספר דרכון
           - תאריך לידה
           - תאריך תוקף
        
        ### ❓ העליתי דרכון ולא קורה כלום?
        **הסיבה הנפוצה ביותר:** שכחת ללחוץ על כפתור **"🔍 סרוק דרכון"**!
        לאחר העלאת התמונה, חייבים ללחוץ על הכפתור כדי להפעיל את הסריקה.
        
        ### טיפים לצילום טוב
        - ודא שהתמונה ברורה ולא מטושטשת
        - ודא תאורה טובה
        - צלם את כל עמוד הפרטים
        - הימנע מהחזרי אור על הדרכון
        """)
    
    with st.expander("🏨 חיפוש מלונות אוטומטי"):
        st.markdown("""
        ### איך לחפש מלון?
        1. הקלד את שם המלון בשדה המתאים
        2. לחץ על "חפש מלון"
        3. המערכת תביא אוטומטית:
           - כתובת מלאה
           - דירוג כוכבים
           - אתר אינטרנט
           - תמונות המלון
        
        ### שעות כניסה ויציאה
        - Check-in: 15:00
        - Check-out: 11:00
        (ניתן לשנות לפי הצורך)
        """)
    
    with st.expander("✈️ פרטי טיסות"):
        st.markdown("""
        ### הזנת טיסות ידנית
        מלא את הפרטים הבאים לכל טיסה:
        - חברת תעופה
        - מספר טיסה
        - שדה תעופה מוצא
        - שדה תעופה יעד
        - תאריך ושעה
        
        ### זיהוי טיסה מצילום מסך
        1. צלם מסך מאתר הטיסות
        2. העלה או הדבק את התמונה
        3. המערכת תזהה אוטומטית את כל פרטי הטיסה
        
        ### כבודה
        סמן את אפשרויות הכבודה הכלולות בחבילה:
        - 🧳 טרולי (יד)
        - 🛄 מזוודה רשומה (בחר גודל)
        """)
    
    with st.expander("📋 היסטוריית הזמנות"):
        st.markdown("""
        ### צפייה בהזמנות קודמות
        - כל ההזמנות נשמרות אוטומטית
        - חפש לפי מספר הזמנה, שם לקוח או אירוע
        - סנן לפי סטטוס או תאריכים
        
        ### סטטוסי הזמנה
        - 📝 **טיוטה** - הזמנה נוצרה אך לא נשלחה
        - 📧 **נשלח** - הזמנה נשלחה ללקוח
        - 👁️ **נצפה** - הלקוח פתח את ההזמנה
        - ✍️ **נחתם** - הלקוח חתם על ההזמנה
        - ❌ **בוטל** - הזמנה בוטלה
        
        ### פעולות על הזמנות
        - הורד PDF מחדש
        - שלח שוב במייל
        - עדכן סטטוס
        """)
    
    with st.expander("📊 ייצוא דוחות"):
        st.markdown("""
        ### ייצוא לאקסל
        - בחר טווח תאריכים
        - סנן לפי סטטוס או סוג אירוע
        - הורד קובץ Excel עם כל הנתונים
        
        ### מה כלול בדוח?
        - מספר הזמנה
        - פרטי לקוח
        - פרטי אירוע
        - מחירים
        - סטטוס
        - תאריכים
        """)
    
    with st.expander("🔧 כלים נוספים"):
        st.markdown("""
        ### ⭐ הופעות שמורות
        שמור הופעות קבועות לשימוש מהיר חוזר.
        
        ### 🗺️ הורדת מפות
        הורד מפות אצטדיון מאתר TikTik והוסף למערכת.
        
        ### 🖼️ ניהול תמונות
        נהל תמונות אווירה לפי קטגוריות (כדורגל, הופעות וכו').
        
        ### 🔑 שינוי סיסמה
        שנה את הסיסמה שלך בכל עת.
        """)
    
    with st.expander("❓ שאלות נפוצות"):
        st.markdown("""
        ### למה אני לא רואה מפת אצטדיון?
        לא לכל הקבוצות יש מפה במערכת. ניתן להעלות מפה ידנית או להשתמש בכלי "הורדת מפות".
        
        ### איך משנים סטטוס הזמנה?
        בהיסטוריית ההזמנות, לחץ על ההזמנה ובחר סטטוס חדש.
        
        ### האם ניתן לערוך הזמנה קיימת?
        לא, אבל ניתן ליצור הזמנה חדשה ולבטל את הישנה.
        
        ### מה לעשות אם ה-OCR לא מזהה נכון?
        ודא שהתמונה ברורה. ניתן תמיד לתקן ידנית את הפרטים.
        
        ### איך לשלוח הזמנה שוב?
        בהיסטוריית ההזמנות, לחץ על "שלח שוב" בהזמנה הרצויה.
        """)
    
    st.markdown("---")
    st.markdown("""
    <div style="text-align: center; color: #888; padding: 1rem;">
        💡 לתמיכה נוספת, פנה למנהל המערכת
    </div>
    """, unsafe_allow_html=True)

def page_stadium_map_scraper():
    """Page for scraping stadium maps from TikTik website"""
    import requests
    import re
    
    st.markdown("""
    <div class="header-container">
        <h1>🗺️ הורדת מפות אצטדיון</h1>
        <p>הורד מפת אצטדיון מלינק TikTik והוסף לקבוצה</p>
    </div>
    """, unsafe_allow_html=True)
    
    if st.button("⬅️ חזרה לתפריט"):
        st.session_state.admin_page = None
        st.rerun()
    
    st.markdown("---")
    
    col1, col2 = st.columns([2, 1])
    
    with col1:
        st.markdown('<div class="form-section">', unsafe_allow_html=True)
        st.markdown("### 🔗 הדבק לינק מהאתר")
        
        tiktik_url = st.text_input("לינק לעמוד המוצר", placeholder="https://www.tiktik-online.co.il/product/...")
        
        if tiktik_url and st.button("🔍 חפש תמונות", use_container_width=True):
            try:
                response = requests.get(tiktik_url, timeout=10)
                content = response.text
                
                img_pattern = r'https://www\.tiktik-online\.co\.il/wp-content/uploads/[^\s"\'<>]+\.(svg|jpg|jpeg|png|webp)'
                all_matches = re.findall(img_pattern, content)
                
                full_pattern = r'(https://www\.tiktik-online\.co\.il/wp-content/uploads/[^\s"\'<>]+\.(?:svg|jpg|jpeg|png|webp))'
                full_urls = re.findall(full_pattern, content)
                
                exclude_words = ['logo', 'icon', 'fav', 'cropped', 'button', 'facebook', 'google', 'whatsapp', 'phone', 'search', 'arrow', 'ticket', 'airplane', 'flight', 'youtube']
                filtered_urls = []
                for url in full_urls:
                    url_lower = url.lower()
                    if not any(word in url_lower for word in exclude_words):
                        if '-300x' not in url and '-150x' not in url and '-100x' not in url:
                            filtered_urls.append(url)
                
                unique_urls = list(dict.fromkeys(filtered_urls))
                
                if unique_urls:
                    st.session_state.found_images = unique_urls
                    st.success(f"✅ נמצאו {len(unique_urls)} תמונות!")
                else:
                    st.warning("⚠️ לא נמצאו תמונות בלינק זה")
            except Exception as e:
                st.error(f"❌ שגיאה בטעינת הלינק: {str(e)}")
        
        if st.session_state.get('found_images'):
            st.markdown("### 🖼️ בחר את מפת האצטדיון")
            for i, url in enumerate(st.session_state.found_images):
                col_img, col_btn = st.columns([3, 1])
                with col_img:
                    try:
                        st.image(url, caption=url.split('/')[-1], use_container_width=True)
                    except:
                        st.write(url)
                with col_btn:
                    if st.button("✅ בחר", key=f"select_img_{i}"):
                        st.session_state.found_map_url = url
                        st.session_state.found_images = None
                        st.rerun()
        
        if st.session_state.get('found_map_url') and not st.session_state.get('found_images'):
            st.markdown("### ✅ תמונה נבחרה")
            st.image(st.session_state.found_map_url, caption="מפת האצטדיון שנבחרה", use_container_width=True)
            st.info("👈 עכשיו בחר קבוצה בצד ימין ולחץ 'שמור מפה'")
            if st.button("🔄 בחר תמונה אחרת", key="change_image"):
                st.session_state.found_map_url = None
                st.rerun()
        
        st.markdown('</div>', unsafe_allow_html=True)
    
    with col2:
        st.markdown('<div class="form-section">', unsafe_allow_html=True)
        st.markdown("### 🏟️ בחר קבוצה/מקום")
        
        from sports_api import LEAGUES, get_teams_by_league, get_hebrew_name
        
        try:
            with open('teams_stadiums_mapping.json', 'r', encoding='utf-8') as f:
                teams_data = json.load(f)
                existing_teams = {t['id']: t for t in teams_data.get('teams', [])}
        except:
            teams_data = {'teams': []}
            existing_teams = {}
        
        category = st.radio("קטגוריה", ["⚽ כדורגל", "🎤 הופעות"], horizontal=True, key="scraper_category")
        
        team_options = {}
        selected_team_display = None
        
        if "כדורגל" in category:
            league_options = ["-- בחר ליגה --"] + list(LEAGUES.keys())
            selected_league = st.selectbox("ליגה", league_options, key="scraper_league")
            
            if selected_league and selected_league != "-- בחר ליגה --":
                english_league = LEAGUES.get(selected_league, "")
                teams = get_teams_by_league(english_league)
                
                if teams:
                    team_display_list = []
                    for t in teams:
                        team_id = t['name'].replace(" ", "_").replace("'", "").lower()
                        english_name = t['name']
                        has_map = team_id in existing_teams and existing_teams[team_id].get('map_filename')
                        status = "✅" if has_map else "❌"
                        hebrew_name = get_hebrew_name(t['name'])
                        display = f"{status} {hebrew_name} ({t['name']})"
                        team_display_list.append((display, team_id, hebrew_name, english_league, english_name))
                    
                    team_options = {t[0]: (t[1], t[2], t[3], t[4]) for t in team_display_list}
                    selected_team_display = st.selectbox("קבוצה", ["-- בחר קבוצה --"] + list(team_options.keys()), key="scraper_team")
                    if selected_team_display == "-- בחר קבוצה --":
                        selected_team_display = None
            else:
                st.info("בחר ליגה כדי לראות את הקבוצות")
        else:
            concert_venues = [
                ("תל אביב - היכל התרבות", "tel_aviv_heichal", "היכל התרבות", "Israel"),
                ("תל אביב - בלומפילד", "tel_aviv_bloomfield", "בלומפילד", "Israel"),
                ("לונדון - O2 Arena", "london_o2", "O2 Arena", "UK"),
                ("לונדון - Wembley", "london_wembley", "Wembley Stadium", "UK"),
                ("ברלין - Mercedes-Benz Arena", "berlin_mercedes", "Mercedes-Benz Arena", "Germany"),
                ("פריז - Accor Arena", "paris_accor", "Accor Arena", "France"),
                ("אמסטרדם - Ziggo Dome", "amsterdam_ziggo", "Ziggo Dome", "Netherlands"),
                ("ברצלונה - Palau Sant Jordi", "barcelona_palau", "Palau Sant Jordi", "Spain"),
                ("מילאן - San Siro", "milan_san_siro", "San Siro", "Italy"),
            ]
            all_venues = []
            for venue in concert_venues:
                venue_id = venue[1]
                has_map = venue_id in existing_teams and existing_teams[venue_id].get('map_filename')
                status = "✅" if has_map else "❌"
                all_venues.append((f"{status} {venue[0]}", venue[1], venue[0], venue[3], venue[2]))
            team_options = {v[0]: (v[1], v[2], v[3], v[4]) for v in all_venues}
            selected_team_display = st.selectbox("מקום", ["-- בחר מקום --"] + list(team_options.keys()), key="scraper_venue")
            if selected_team_display == "-- בחר מקום --":
                selected_team_display = None
        
        if st.session_state.get('found_map_url') and selected_team_display:
            team_info = team_options[selected_team_display]
            selected_team_id = team_info[0]
            team_name_he = team_info[1]
            team_league = team_info[2]
            team_name_en = team_info[3] if len(team_info) > 3 else team_name_he
            
            if st.button("💾 שמור מפה", use_container_width=True, type="primary"):
                try:
                    map_url = st.session_state.found_map_url
                    ext = map_url.split('.')[-1].split('?')[0]
                    filename = f"{selected_team_id}.{ext}"
                    filepath = f"stadium_maps/{filename}"
                    
                    response = requests.get(map_url, timeout=30)
                    with open(filepath, 'wb') as f:
                        f.write(response.content)
                    
                    if selected_team_id in existing_teams:
                        for team in teams_data['teams']:
                            if team['id'] == selected_team_id:
                                team['map_filename'] = filename
                                break
                    else:
                        new_team = {
                            'id': selected_team_id,
                            'name_en': team_name_en,
                            'name_he': team_name_he,
                            'stadium': '',
                            'stadium_he': '',
                            'city': '',
                            'city_he': '',
                            'country': '',
                            'league': team_league,
                            'map_filename': filename
                        }
                        teams_data['teams'].append(new_team)
                    
                    with open('teams_stadiums_mapping.json', 'w', encoding='utf-8') as f:
                        json.dump(teams_data, f, ensure_ascii=False, indent=2)
                    
                    st.session_state.map_save_success = f"מפת האצטדיון נשמרה בהצלחה עבור: {team_name_he}"
                    st.session_state.found_map_url = None
                    st.rerun()
                except Exception as e:
                    st.error(f"❌ שגיאה בשמירה: {str(e)}")
        
        if st.session_state.get('map_save_success'):
            st.balloons()
            st.success(f"🎉 {st.session_state.map_save_success}")
            st.session_state.map_save_success = None
        
        st.markdown('</div>', unsafe_allow_html=True)
    
    st.markdown("---")
    st.markdown("### 📊 סיכום מפות")
    
    try:
        with open('teams_stadiums_mapping.json', 'r', encoding='utf-8') as f:
            summary_data = json.load(f)
            saved_teams = summary_data.get('teams', [])
    except:
        saved_teams = []
    
    if saved_teams:
        teams_with_maps = [t for t in saved_teams if t.get('map_filename')]
        
        st.markdown(f"**✅ יש מפה ({len(teams_with_maps)} קבוצות/מקומות)**")
        for t in teams_with_maps:
            st.markdown(f"- {t.get('name_he', t.get('id'))} ({t.get('league', '')})")

def main():
    # Try to restore session from token if not logged in
    if not st.session_state.get('logged_in'):
        restored_user = restore_session_from_token()
        if restored_user:
            st.session_state.user = restored_user
            st.session_state.logged_in = True
    
    # Check if user is logged in
    if not st.session_state.get('logged_in'):
        page_login()
        st.stop()
    
    user = st.session_state.get('user', {})
    is_admin = user.get('is_admin', False)
    
    # Sidebar navigation
    st.sidebar.markdown(f"### 👤 {user.get('full_name', 'משתמש')}")
    if is_admin:
        st.sidebar.markdown("👑 מנהל מערכת")
    st.sidebar.markdown("---")
    st.sidebar.markdown("## 📌 תפריט")
    
    page = st.sidebar.radio(
        "בחר עמוד",
        ["🆕 הזמנה חדשה", "📋 היסטוריית הזמנות", "📊 ייצוא דוחות"],
        label_visibility="collapsed"
    )
    
    st.sidebar.markdown("---")
    st.sidebar.markdown("##### 🔧 כלים")
    if st.sidebar.button("📦 חבילות קבועות", use_container_width=True):
        st.session_state.admin_page = "packages"
    if st.sidebar.button("📖 מדריך למתחיל", use_container_width=True):
        st.session_state.admin_page = "beginner_guide"
    if st.sidebar.button("❓ עזרה", use_container_width=True):
        st.session_state.admin_page = "help"
    if st.sidebar.button("⭐ הופעות שמורות", use_container_width=True):
        st.session_state.admin_page = "saved_concerts"
    if st.sidebar.button("🗺️ הורדת מפות", use_container_width=True):
        st.session_state.admin_page = "maps"
    if st.sidebar.button("🔑 שינוי סיסמה", use_container_width=True):
        st.session_state.admin_page = "change_password"
    
    if is_admin:
        st.sidebar.markdown("---")
        st.sidebar.markdown("##### 👑 ניהול")
        if st.sidebar.button("🖼️ ניהול תמונות", use_container_width=True):
            st.session_state.admin_page = "images"
        if st.sidebar.button("👥 ניהול משתמשים", use_container_width=True):
            st.session_state.admin_page = "users"
    
    st.sidebar.markdown("---")
    if st.sidebar.button("🚪 התנתק", use_container_width=True):
        st.session_state.logged_in = False
        st.session_state.user = None
        st.session_state.admin_page = None
        clear_session_token()
        st.rerun()
    
    render_ai_chatbot()
    
    # Determine which page to show
    if st.session_state.get("admin_page") == "packages":
        page_package_templates()
    elif st.session_state.get("admin_page") == "beginner_guide":
        page_beginner_guide()
    elif st.session_state.get("admin_page") == "help":
        page_help()
    elif st.session_state.get("admin_page") == "images":
        page_image_gallery()
    elif st.session_state.get("admin_page") == "maps":
        page_stadium_map_scraper()
    elif st.session_state.get("admin_page") == "saved_concerts":
        page_saved_concerts()
    elif st.session_state.get("admin_page") == "change_password":
        page_change_password()
    elif st.session_state.get("admin_page") == "users" and is_admin:
        page_user_management()
    elif page == "📋 היסטוריית הזמנות":
        page_order_history()
    elif page == "📊 ייצוא דוחות":
        page_export()
    else:
        page_new_order()

if __name__ == "__main__":
    main()
